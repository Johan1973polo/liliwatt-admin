from flask import Flask, render_template, request, jsonify, session, redirect, url_for
import os
import requests
import json
import uuid
import re
import hmac
import hashlib
import html as html_mod
import threading
import time
from functools import wraps
from datetime import datetime, timedelta
from google_meet_service import create_referent_meet_room

# ── Newsletter : constantes ────────────────────────────────────────────────
MENTIONS_LEGALES = {
    "raison_sociale": "LILISTRAT STRATÉGIE SAS",
    "marque":         "LILIWATT",
    "adresse":        "59 rue de Ponthieu, Bureau 326, 75008 Paris",
    "email":          "contact@liliwatt.fr",
    "telephone":      "01 84 16 08 56",
    "forme":          "au capital de 10 000 €",
    "siren":          "SIREN 103 572 947",
}

SOCIAL_LINKS = [
    ("linkedin",  "https://www.linkedin.com/company/liliwatt/",              "linkedin.png"),
    ("instagram", "https://www.instagram.com/liliwatt.fr/",                  "instagram.png"),
    ("x",         "https://x.com/liliwattfrance",                            "x.png"),
    ("youtube",   "https://www.youtube.com/@liliwattfrance",                 "youtube.png"),
    ("facebook",  "https://www.facebook.com/profile.php?id=61577269553280",  "facebook.png"),
]

FOOTER_NAV = [
    ("Nos offres",        "https://liliwatt.fr/offres.html"),
    ("Qui sommes-nous ?", "https://liliwatt.fr/apropos.html"),
    ("Actualités",        "https://liliwatt.fr/blog"),
    ("Contact",           "https://liliwatt.fr/contact.html"),
]

NL_ASSETS_BASE = "https://liliwatt-admin.onrender.com/static/newsletter"
# Cache-busting des images de newsletter.
# Gmail met en cache les images par URL exacte, indefiniment.
# A INCREMENTER a chaque fois qu'un asset de static/newsletter/
# est modifie, sinon les destinataires verront l'ancienne version.
NL_ASSETS_VERSION = "6"

def _nl_asset(nom):
    """Retourne l'URL complète d'un asset newsletter avec cache-busting."""
    return f"{NL_ASSETS_BASE}/{nom}?v={NL_ASSETS_VERSION}"

NL_INTRO = {
    "newsletter":    "Votre newsletter mensuelle avec les dernières actualités du marché de l'énergie, nos conseils et nos services pour vous accompagner au mieux.",
    "communication": "Nous avons une information importante à vous transmettre.",
    "bienvenue":     "Nous sommes heureux de vous compter parmi nos clients.",
}

BIENVENUE_DEFAUT = {
    "objet": "Bienvenue chez LILIWATT",
    "titre": "Bienvenue chez LILIWATT",
    "message": """Merci de votre confiance.

Vous avez choisi LILIWATT pour votre contrat d\u2019\xe9nergie, et nous en sommes heureux. Notre r\xf4le ne s\u2019arr\xeate pas \xe0 la signature \u2014 il commence.

\xc0 partir d\u2019aujourd\u2019hui, nous suivons deux choses pour vous\u00a0: votre contrat, et le march\xe9 de l\u2019\xe9nergie. Les prix bougent, les offres changent, les \xe9ch\xe9ances arrivent. Nous surveillons tout cela en continu, et nous revenons vers vous \xe0 chaque fois que votre situation peut \xeatre am\xe9lior\xe9e. Vous n\u2019avez rien \xe0 surveiller de votre c\xf4t\xe9.

Nous sommes votre interlocuteur unique pour tout ce qui touche \xe0 votre \xe9nergie\u00a0: votre contrat, vos consommations, vos factures, une question sur le march\xe9 ou simplement un doute. Une seule adresse, un seul num\xe9ro, et quelqu\u2019un qui conna\xeet d\xe9j\xe0 votre dossier.

Nous travaillons en toute transparence. Vous saurez toujours pourquoi nous vous recommandons une offre plut\xf4t qu\u2019une autre, et vous resterez libre de votre d\xe9cision.

N\u2019h\xe9sitez jamais \xe0 nous solliciter \u2014 c\u2019est exactement pour cela que nous sommes l\xe0.

Deux choses encore, juste en dessous de ce message. Si notre accompagnement vous convient, quelques mots d\u2019avis sur Google nous aident beaucoup \xe0 nous faire conna\xeetre. Et si vous connaissez une entreprise qui gagnerait \xe0 \xeatre mieux conseill\xe9e sur son \xe9nergie, notre programme de parrainage vous permet de la recommander \u2014 et d\u2019\xeatre remerci\xe9 pour cela.

\xc0 tr\xe8s bient\xf4t,
L\u2019\xe9quipe LILIWATT""",
}

GOOGLE_AVIS_URL = os.environ.get("GOOGLE_AVIS_URL", "https://g.page/r/CUzpowIihy_ZEBM/review")
PARRAINAGE_URL  = os.environ.get("PARRAINAGE_URL", "https://liliwatt-parrainage.onrender.com/")
NEWSLETTER_SECRET = os.environ.get("NEWSLETTER_SECRET", "nl-liliwatt-secret-2026")
NL_DELAI_S = 3
NL_LOT = 50

_nl_status = {"en_cours": False, "total": 0, "envoyes": 0, "erreurs": [], "objet": ""}

def parse_float(val):
    if not val:
        return 0.0
    try:
        return float(str(val).replace(',', '.').replace(' ', '').replace('\u202f', ''))
    except:
        return 0.0

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'liliwatt-admin-secret-2026')
app.json.ensure_ascii = False

ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', '')
ZOHO_CLIENT_ID = os.environ.get('ZOHO_CLIENT_ID', '')
ZOHO_CLIENT_SECRET = os.environ.get('ZOHO_CLIENT_SECRET', '')
ZOHO_REFRESH_TOKEN = os.environ.get('ZOHO_REFRESH_TOKEN', '')
ZOHO_ORG_ID = os.environ.get('ZOHO_ORG_ID', '')

def get_zoho_token():
    r = requests.post('https://accounts.zoho.eu/oauth/v2/token', data={
        'refresh_token': ZOHO_REFRESH_TOKEN,
        'client_id': ZOHO_CLIENT_ID,
        'client_secret': ZOHO_CLIENT_SECRET,
        'grant_type': 'refresh_token'
    })
    return r.json().get('access_token')



def save_to_sheet(prenom, nom, email, password, poste, drive_folder_id='', referent_email='', token_rgpd='', role='vendeur'):
    """Enregistre le commercial dans Google Sheets"""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        import json
        from datetime import datetime

        creds_json = os.environ.get('GOOGLE_CREDS_JSON', '')
        if not creds_json:
            print("⚠️ GOOGLE_CREDS_JSON non défini")
            return False

        creds_dict = json.loads(creds_json)
        creds = Credentials.from_service_account_info(
            creds_dict,
            scopes=[
                'https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive'
            ]
        )
        gc = gspread.authorize(creds)
        sheet_id = os.environ.get('GOOGLE_SHEET_ID', '')
        sh = gc.open_by_key(sheet_id)
        ws = sh.sheet1
        rgpd_link = f'https://liliwatt-courtier.onrender.com/rgpd/{token_rgpd}' if token_rgpd else ''
        ws.append_row([
            nom.upper(),
            prenom.capitalize(),
            password,
            email,
            poste,
            drive_folder_id,
            referent_email,
            token_rgpd,
            rgpd_link,
            role,
            'actif'
        ])
        print(f"✅ {nom} {prenom} enregistré dans Google Sheets (token RGPD: {token_rgpd})")
        return True
    except Exception as e:
        print(f"⚠️ Erreur Google Sheets : {e}")
        return False

def orpheliner_vendeurs(ancien_referent_email):
    """Détache tous les vendeurs de leur ancien référent."""
    try:
        gc = get_sheets_client()
        sheet_id = os.environ.get('GOOGLE_SHEET_ID', '')
        ws = gc.open_by_key(sheet_id).sheet1
        rows = ws.get_all_values()
        orphelins = []
        for i, row in enumerate(rows):
            if len(row) > 6 and row[6].strip().lower() == ancien_referent_email.lower():
                ws.update_cell(i + 1, 7, '')
                orphelins.append(row[3].strip().lower())
                print(f'  → Vendeur orphelin : {row[3]}')
        print(f'✅ {len(orphelins)} vendeur(s) orphelin(s) après départ de {ancien_referent_email}')
        if orphelins:
            crm_url = os.environ.get('CRM_URL', 'https://liliwatt-crm-8ofi.vercel.app')
            crm_key = os.environ.get('CRM_API_KEY', 'liliwatt-crm-api-key-2026')
            for email_v in orphelins:
                try:
                    requests.post(f'{crm_url}/api/crm/assign-referent',
                        headers={'X-API-Key': crm_key, 'Content-Type': 'application/json'},
                        json={'vendeur_email': email_v, 'referent_email': None}, timeout=10)
                except: pass
            print(f'✅ CRM Neon : {len(orphelins)} vendeurs mis à referentId=null')
        return orphelins
    except Exception as e:
        print(f'❌ Erreur orpheliner_vendeurs : {e}')
        return []


def get_sheets_client():
    """Retourne un client gspread authentifié"""
    import gspread
    from google.oauth2.service_account import Credentials
    creds_json = os.environ.get('GOOGLE_CREDS_JSON', '')
    if not creds_json:
        return None
    creds_dict = json.loads(creds_json)
    creds = Credentials.from_service_account_info(
        creds_dict,
        scopes=['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    )
    return gspread.authorize(creds)

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def make_signature(prenom, nom, poste, telephone, email):
    return f"""<table cellpadding="0" cellspacing="0" border="0" style="font-family:Arial,sans-serif;max-width:480px;">
  <tr><td colspan="2" style="border-top:4px solid #7c3aed;padding-bottom:12px;"></td></tr>
  <tr><td colspan="2" style="font-size:16px;font-weight:700;color:#1e1b4b;padding-bottom:2px;">{prenom} {nom}</td></tr>
  <tr><td colspan="2" style="font-size:10px;font-weight:600;color:#7c3aed;letter-spacing:1.5px;text-transform:uppercase;padding-bottom:2px;">{poste}</td></tr>
  <tr><td colspan="2" style="font-size:10px;font-weight:600;color:#7c3aed;letter-spacing:1px;text-transform:uppercase;padding-bottom:12px;">LILIWATT &mdash; Courtage &Eacute;nergie B2B &amp; B2C</td></tr>
  <tr>
    <td style="padding:3px 12px 3px 0;color:#9ca3af;font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;vertical-align:top;">T&eacute;l</td>
    <td style="font-size:13px;"><a href="tel:{telephone.replace(' ','')}" style="color:#1e1b4b;text-decoration:none;font-weight:600;">{telephone}</a></td>
  </tr>
  <tr>
    <td style="padding:3px 12px 3px 0;color:#9ca3af;font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;vertical-align:top;">Email</td>
    <td style="font-size:13px;"><a href="mailto:{email}" style="color:#7c3aed;text-decoration:none;font-weight:600;">{email}</a></td>
  </tr>
  <tr>
    <td style="padding:3px 12px 3px 0;color:#9ca3af;font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;vertical-align:top;">Web</td>
    <td style="font-size:13px;"><a href="https://liliwatt.fr" style="color:#7c3aed;text-decoration:none;font-weight:600;">www.liliwatt.fr</a></td>
  </tr>
  <tr>
    <td style="padding:3px 12px 3px 0;color:#9ca3af;font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;vertical-align:top;">Adresse</td>
    <td style="font-size:13px;color:#374151;">59 rue de Ponthieu, Bureau 326<br>75008 Paris</td>
  </tr>
  <tr><td colspan="2" style="padding-top:12px;border-bottom:1px solid #7c3aed;"></td></tr>
  <tr><td colspan="2" style="font-size:11px;color:#6b7280;padding-top:10px;">Courtier &Eacute;nergie B2B &amp; B2C<br><span style="color:#7c3aed;font-weight:700;">18% d'&eacute;conomies en moyenne</span> &mdash; Sans engagement &bull; Sans coupure &bull; 18+ fournisseurs compar&eacute;s</td></tr>
  <tr><td colspan="2" style="border-bottom:2px solid #7c3aed;padding-top:12px;"></td></tr>
</table>"""


def generate_password():
    """Génère un mot de passe sécurisé automatiquement"""
    import random
    import string
    majuscules = random.choices(string.ascii_uppercase, k=2)
    minuscules = random.choices(string.ascii_lowercase, k=4)
    chiffres = random.choices(string.digits, k=3)
    speciaux = random.choices('@#$!%&', k=2)
    all_chars = majuscules + minuscules + chiffres + speciaux
    random.shuffle(all_chars)
    return ''.join(all_chars)

def send_welcome_email(prenom, nom, email, password, poste='', telephone='', email_perso='', account_id_zoho='', token_rgpd='', referent_email=''):
    """Envoie l'email de bienvenue via API Zoho Mail"""
    try:
        destinataire = email_perso if email_perso else email
        email_liliwatt = email

        token = get_zoho_token()
        if not token:
            print("⚠️ Token Zoho non obtenu pour email")
            return False

        # Chercher le nom du référent
        referent_name = None
        if referent_email:
            try:
                gc = get_sheets_client()
                if gc:
                    sheet_id = os.environ.get('GOOGLE_SHEET_ID', '')
                    ws = gc.open_by_key(sheet_id).sheet1
                    for row in ws.get_all_values()[1:]:
                        if len(row) > 3 and row[3].strip().lower() == referent_email.strip().lower():
                            referent_name = f"{row[1]} {row[0]}".strip()
                            break
            except:
                referent_name = referent_email

        referent_block = ''
        if referent_name:
            referent_block = f"""<div style="background:#f0fdf4;border-radius:16px;padding:28px;margin-bottom:24px;border:1px solid #bbf7d0;">
    <h3 style="color:#166534;margin:0 0 8px;font-size:17px;">👤 Votre référent</h3>
    <p style="color:#166534;margin:0;font-size:14px;line-height:1.8;">
      <strong>{referent_name}</strong> est votre référent dédié chez LILIWATT. Il est disponible via la messagerie interne du CRM pour vous accompagner dans la prise en main de vos outils, répondre à vos questions et vous guider lors de vos premières semaines d'activité.<br><br>
      N'hésitez pas à le contacter dès votre connexion.
    </p>
  </div>"""

        html_body = f"""<!DOCTYPE html><html><body style="margin:0;padding:0;background:#f5f3ff;font-family:Arial,sans-serif;">
<div style="max-width:600px;margin:0 auto;padding:40px 20px;">
  <div style="background:linear-gradient(135deg,#7c3aed,#d946ef);border-radius:16px;padding:40px;text-align:center;margin-bottom:32px;">
    <h1 style="color:white;margin:0;font-size:32px;font-weight:700;letter-spacing:2px;">⚡ LILIWATT</h1>
    <p style="color:rgba(255,255,255,0.85);margin:8px 0 0;font-size:15px;">Cabinet de courtage en énergie B2B</p>
  </div>
  <div style="background:white;border-radius:16px;padding:32px;margin-bottom:24px;border:1px solid #e9d5ff;">
    <h2 style="color:#7c3aed;margin:0 0 16px;font-size:22px;">Bienvenue {prenom} ! 🎉</h2>
    <p style="color:#374151;line-height:1.8;margin:0;font-size:15px;">
      Nous sommes ravis de vous accueillir dans l'équipe LILIWATT.<br><br>
      Votre espace de travail est prêt. <strong>Votre point d'entrée unique est le CRM LILIWATT</strong> — depuis ce tableau de bord personnel, vous accédez directement à tous vos outils métier, vos modules de formation et votre messagerie interne.<br><br>
      Dès votre première connexion, vous trouverez vos accès pré-configurés et vos modules de formation disponibles pour vous permettre de démarrer rapidement.
    </p>
  </div>
  <div style="background:linear-gradient(135deg,#7c3aed,#d946ef);border-radius:16px;padding:36px;margin-bottom:24px;text-align:center;">
    <h3 style="color:white;margin:0 0 8px;font-size:22px;">🖥️ Votre espace CRM LILIWATT</h3>
    <p style="color:rgba(255,255,255,0.85);margin:0 0 6px;font-size:14px;">Tableau de bord — Formation — Messagerie — Outils métier</p>
    <p style="color:rgba(255,255,255,0.75);margin:0 0 24px;font-size:13px;">Connectez-vous dès maintenant pour découvrir votre espace et commencer votre formation.</p>
    <a href="https://liliwatt-crm-8ofi.vercel.app" style="background:white;color:#7c3aed;padding:16px 40px;border-radius:50px;text-decoration:none;font-weight:700;font-size:16px;display:inline-block;letter-spacing:0.5px;">Accéder à mon espace →</a>
    <div style="margin-top:28px;background:rgba(0,0,0,0.2);border-radius:12px;padding:18px;">
      <p style="color:rgba(255,255,255,0.9);margin:0 0 8px;font-size:14px;">📧 Identifiant : <strong>{email_liliwatt}</strong></p>
      <p style="color:rgba(255,255,255,0.9);margin:0;font-size:14px;">🔑 Mot de passe : <strong>{password}</strong></p>
    </div>
  </div>
  <div style="background:white;border-radius:16px;padding:32px;margin-bottom:24px;border:1px solid #e9d5ff;">
    <h3 style="color:#374151;margin:0 0 6px;font-size:17px;">Vos outils accessibles depuis le CRM</h3>
    <p style="color:#6b7280;margin:0 0 20px;font-size:13px;">Tous vos outils sont accessibles en un clic depuis votre tableau de bord personnel.</p>
    <div style="padding:16px;background:#f5f3ff;border-radius:12px;margin-bottom:12px;">
      <p style="margin:0 0 4px;font-weight:700;color:#374151;font-size:15px;">📞 Base Prospection</p>
      <p style="margin:0;font-size:13px;color:#6b7280;line-height:1.6;">Vos prospects qualifiés avec coordonnées téléphoniques, secteur d'activité et statut d'appel. Commencez à prospecter dès le premier jour.</p>
    </div>
    <div style="padding:16px;background:#f5f3ff;border-radius:12px;">
      <p style="margin:0 0 4px;font-weight:700;color:#374151;font-size:15px;">⚡ Outil Courtage — Extracteur &amp; Comparatif</p>
      <p style="margin:0;font-size:13px;color:#6b7280;line-height:1.6;">Déposez la facture de votre prospect, notre outil extrait automatiquement les données et génère un comparatif personnalisé en moins de 2 minutes.</p>
    </div>
  </div>
  <div style="background:white;border-radius:16px;padding:28px;margin-bottom:24px;border:1px solid #e9d5ff;">
    <h3 style="color:#374151;margin:0 0 6px;font-size:17px;">✉️ Votre boîte mail professionnelle</h3>
    <p style="color:#6b7280;margin:0 0 16px;font-size:13px;">Votre adresse email professionnelle est active. Utilisez-la pour toutes vos communications clients.</p>
    <div style="background:#f5f3ff;border-radius:12px;padding:16px;">
      <p style="margin:0 0 8px;font-size:14px;color:#374151;">🌐 Accès webmail : <a href="https://mail.zoho.eu" style="color:#7c3aed;font-weight:600;">mail.zoho.eu</a></p>
      <p style="margin:0 0 8px;font-size:14px;color:#374151;">📧 Email : <strong>{email_liliwatt}</strong></p>
      <p style="margin:0;font-size:14px;color:#374151;">🔑 Mot de passe : <strong>{password}</strong></p>
    </div>
  </div>
  {referent_block}
  <div style="background:white;border-radius:16px;padding:28px;margin-bottom:24px;border:1px solid #e9d5ff;">
    <h3 style="color:#374151;margin:0 0 16px;font-size:17px;">🚀 Vos prochaines étapes</h3>
    <div style="display:flex;gap:16px;margin-bottom:14px;align-items:flex-start;"><div style="background:#7c3aed;color:white;width:28px;height:28px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-weight:700;font-size:13px;flex-shrink:0;">1</div><p style="margin:0;color:#374151;font-size:14px;line-height:1.6;"><strong>Connectez-vous au CRM</strong> avec vos identifiants ci-dessus</p></div>
    <div style="display:flex;gap:16px;margin-bottom:14px;align-items:flex-start;"><div style="background:#7c3aed;color:white;width:28px;height:28px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-weight:700;font-size:13px;flex-shrink:0;">2</div><p style="margin:0;color:#374151;font-size:14px;line-height:1.6;"><strong>Commencez votre formation</strong> — les modules sont déverrouillés progressivement par votre référent</p></div>
    <div style="display:flex;gap:16px;margin-bottom:14px;align-items:flex-start;"><div style="background:#7c3aed;color:white;width:28px;height:28px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-weight:700;font-size:13px;flex-shrink:0;">3</div><p style="margin:0;color:#374151;font-size:14px;line-height:1.6;"><strong>Découvrez vos outils</strong> — prospection et courtage accessibles depuis votre tableau de bord</p></div>
    <div style="display:flex;gap:16px;align-items:flex-start;"><div style="background:#d946ef;color:white;width:28px;height:28px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-weight:700;font-size:13px;flex-shrink:0;">4</div><p style="margin:0;color:#374151;font-size:14px;line-height:1.6;"><strong>Lancez vos premiers appels</strong> — votre base de prospects qualifiés vous attend</p></div>
  </div>
  <div style="text-align:center;padding:24px 0;">
    <div style="background:linear-gradient(135deg,#7c3aed,#d946ef);border-radius:12px;padding:20px;margin-bottom:20px;">
      <p style="color:white;margin:0;font-size:14px;font-weight:600;">Une question ? Contactez-nous</p>
      <p style="color:rgba(255,255,255,0.85);margin:8px 0 0;font-size:13px;">contact@liliwatt.fr — 01 84 16 08 56</p>
    </div>
    <p style="color:#9ca3af;font-size:12px;margin:0;">LILIWATT — LILISTRAT STRATÉGIE SAS</p>
    <p style="color:#9ca3af;font-size:12px;margin:4px 0 0;">59 rue de Ponthieu, Bureau 326 — 75008 Paris</p>
    <p style="color:#9ca3af;font-size:12px;margin:4px 0 0;">www.liliwatt.fr</p>
  </div>
</div>
</body></html>"""

        account_id = os.environ.get('ZOHO_ACCOUNT_ID', '8439060000000002002')
        print(f"📧 Envoi email bienvenue à {destinataire} depuis account_id: {account_id}")

        send_r = requests.post(
            f'https://mail.zoho.eu/api/accounts/{account_id}/messages',
            headers={'Authorization': f'Zoho-oauthtoken {token}', 'Content-Type': 'application/json'},
            json={
                'fromAddress': 'bo@liliwatt.fr',
                'toAddress': destinataire,
                'subject': f'Bienvenue chez LILIWATT — Vos accès {prenom} 🎉',
                'content': html_body,
                'mailFormat': 'html'
            },
            timeout=15
        )

        result = send_r.json()
        print(f"✅ Email bienvenue envoyé - status: {send_r.status_code} - response: {str(result)[:100]}")
        return True

    except Exception as e:
        print(f"⚠️ Erreur email bienvenue : {e}")
        import traceback; traceback.print_exc()
        return False


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        if request.form.get('password') == ADMIN_PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('index'))
        error = 'Mot de passe incorrect'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/api/create-drive-folder', methods=['POST'])
@login_required
def create_drive_folder():
    try:
        from googleapiclient.discovery import build
        from google.oauth2.service_account import Credentials as SACredentials
        import base64

        d = request.get_json()
        prenom = d.get('prenom', '').strip()
        nom = d.get('nom', '').strip()
        if not prenom or not nom:
            return jsonify({'success': False, 'error': 'Prénom et nom requis'})

        VENDEURS_PARENT_ID = '157Sol6u32W0loIEv8CmYT3uoDaGyZ7q6'
        SHARED_DRIVE_ID = '0ACKaJQqRlmwgUk9PVA'
        folder_name = f"{prenom.capitalize()} {nom.upper()}"

        # Charger credentials Drive (3 sources possibles)
        creds_b64 = os.environ.get('GOOGLE_DRIVE_CREDS_BASE64', '')
        creds_json_env = os.environ.get('GOOGLE_CREDS_JSON', '')
        if creds_b64:
            creds_dict = json.loads(base64.b64decode(creds_b64).decode())
        elif creds_json_env:
            creds_dict = json.loads(creds_json_env)
        else:
            with open(os.path.join(os.path.dirname(__file__), 'liliwatt-eddcc0bc9e18.json')) as f:
                creds_dict = json.load(f)

        creds = SACredentials.from_service_account_info(
            creds_dict, scopes=['https://www.googleapis.com/auth/drive']
        )
        drive = build('drive', 'v3', credentials=creds)

        # Créer le dossier principal du vendeur dans le Shared Drive
        vendeur_folder = drive.files().create(
            body={'name': folder_name, 'mimeType': 'application/vnd.google-apps.folder', 'parents': [VENDEURS_PARENT_ID], 'driveId': SHARED_DRIVE_ID},
            fields='id', supportsAllDrives=True
        ).execute()
        vendeur_id = vendeur_folder['id']

        # Créer les 3 sous-dossiers dans le Shared Drive
        for sub in ['CLIENT EN ATTENTE', 'CLIENTS SIGNÉS', 'CLIENTS PERDUS']:
            drive.files().create(
                body={'name': sub, 'mimeType': 'application/vnd.google-apps.folder', 'parents': [vendeur_id], 'driveId': SHARED_DRIVE_ID},
                fields='id', supportsAllDrives=True
            ).execute()

        link = f"https://drive.google.com/drive/folders/{vendeur_id}"
        print(f"✅ Dossier Drive créé : {folder_name} → {vendeur_id}")
        return jsonify({'success': True, 'drive_folder_id': vendeur_id, 'link': link})

    except Exception as e:
        import traceback
        print(f"⚠️ Erreur création dossier Drive : {e}")
        return jsonify({'success': False, 'error': str(e), 'trace': traceback.format_exc()})

@app.route('/api/update-rgpd-links')
@login_required
def update_rgpd_links():
    try:
        gc = get_sheets_client()
        if not gc:
            return jsonify({'success': False, 'error': 'Google Sheets non configuré'})
        sheet_id = os.environ.get('GOOGLE_SHEET_ID', '')
        ws = gc.open_by_key(sheet_id).sheet1
        rows = ws.get_all_values()
        updated = []
        batch = []
        for i, row in enumerate(rows):
            token = row[7] if len(row) > 7 else ''
            link = row[8] if len(row) > 8 else ''
            email = row[3] if len(row) > 3 else ''
            if token and not link and '@' in email:
                rgpd_link = f'https://liliwatt-courtier.onrender.com/rgpd/{token}'
                batch.append({'range': f'I{i+1}', 'values': [[rgpd_link]]})
                updated.append({'email': email, 'link': rgpd_link})
        if batch:
            ws.batch_update(batch, value_input_option='RAW')
        return jsonify({'success': True, 'updated': len(updated), 'details': updated})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/signature/<token_rgpd>')
def signature_page(token_rgpd):
    try:
        gc = get_sheets_client()
        if not gc:
            return 'Erreur configuration', 500
        sheet_id = os.environ.get('GOOGLE_SHEET_ID', '')
        ws = gc.open_by_key(sheet_id).sheet1
        rows = ws.get_all_values()
        vendeur = None
        for row in rows:
            if len(row) > 7 and row[7] == token_rgpd and '@' in (row[3] or ''):
                vendeur = {'nom': row[0], 'prenom': row[1], 'email': row[3], 'poste': row[4], 'telephone': ''}
                break
        if not vendeur:
            return 'Vendeur introuvable.', 404
        sig_html = make_signature(vendeur['prenom'], vendeur['nom'], vendeur['poste'], vendeur['telephone'], vendeur['email'])
        return f"""<!DOCTYPE html>
<html lang="fr"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Signature — {vendeur['prenom']} {vendeur['nom']}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',system-ui,sans-serif;background:#f5f3ff;min-height:100vh;display:flex;align-items:center;justify-content:center;padding:20px}}
.card{{background:#fff;border-radius:16px;padding:40px;max-width:560px;width:100%;box-shadow:0 4px 24px rgba(124,58,237,.1)}}
h1{{color:#1e1b4b;font-size:20px;margin-bottom:6px}}
.sub{{color:#6b7280;font-size:13px;margin-bottom:24px}}
.sig-box{{border:1.5px solid #e9d5ff;border-radius:10px;padding:20px;margin-bottom:20px;background:#faf5ff}}
.btn{{display:block;width:100%;padding:14px;background:linear-gradient(135deg,#7c3aed,#d946ef);color:#fff;border:none;border-radius:10px;font-size:15px;font-weight:700;cursor:pointer;transition:transform .15s}}
.btn:hover{{transform:translateY(-1px)}}
.msg{{text-align:center;margin-top:12px;font-size:14px;color:#16a34a;font-weight:600;display:none}}
</style></head><body>
<div class="card">
<h1>Signature email de {vendeur['prenom']} {vendeur['nom']}</h1>
<p class="sub">Cliquez sur le bouton pour copier la signature dans votre presse-papier, puis collez-la dans Zoho Mail &rarr; Param&egrave;tres &rarr; Signatures.</p>
<div class="sig-box" id="sigBox">{sig_html}</div>
<button class="btn" onclick="copySig()">&#128203; Copier la signature</button>
<p class="msg" id="msg">&#10003; Signature copi&eacute;e !</p>
</div>
<script>
async function copySig(){{
  const box=document.getElementById('sigBox');
  try{{
    const blob=new Blob([box.innerHTML],{{type:'text/html'}});
    await navigator.clipboard.write([new ClipboardItem({{'text/html':blob}})]);
  }}catch(e){{
    const r=document.createRange();r.selectNodeContents(box);
    const s=window.getSelection();s.removeAllRanges();s.addRange(r);
    document.execCommand('copy');s.removeAllRanges();
  }}
  const m=document.getElementById('msg');m.style.display='block';
  setTimeout(()=>m.style.display='none',3000);
}}
</script></body></html>"""
    except Exception as e:
        return f'Erreur : {e}', 500

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/api/vendeurs')
@login_required
def list_vendeurs_api():
    try:
        gc = get_sheets_client()
        if not gc:
            return jsonify({'success': False, 'error': 'Sheets non configuré'})
        sheet_id = os.environ.get('GOOGLE_SHEET_ID', '')
        ws = gc.open_by_key(sheet_id).sheet1
        rows = ws.get_all_values()
        vendeurs = []
        for row in rows[1:]:  # skip header
            if len(row) > 3 and '@' in row[3]:
                statut = (row[10] if len(row) > 10 else 'actif').strip().lower()
                if statut == 'supprime':
                    continue
                vendeurs.append({
                    'nom': row[0],
                    'prenom': row[1],
                    'email': row[3],
                    'referent_email': (row[6] if len(row) > 6 else '').strip(),
                    'role': (row[9] if len(row) > 9 else 'vendeur').strip().lower(),
                    'statut': statut
                })
        return jsonify({'success': True, 'vendeurs': vendeurs})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/referents-avec-equipe')
@login_required
def referents_avec_equipe():
    try:
        gc = get_sheets_client()
        sheet_id = os.environ.get('GOOGLE_SHEET_ID', '')
        ws = gc.open_by_key(sheet_id).sheet1
        rows = ws.get_all_values()
        # Collecter les référents par colonne J (role) ET colonne G (fallback)
        ref_emails = set()
        for row in rows[1:]:
            # Méthode 1 : colonne J = referent
            if len(row) > 9 and row[9].strip().lower() == 'referent' and len(row) > 3 and '@' in row[3]:
                ref_emails.add(row[3].strip().lower())
            # Méthode 2 : email en colonne G
            if len(row) > 6 and row[6].strip() and '@' in row[6]:
                ref_emails.add(row[6].strip().lower())
        # Construire la liste avec équipes
        referents = []
        for row in rows[1:]:
            if len(row) > 3 and row[3].strip().lower() in ref_emails:
                email = row[3].strip()
                equipe = [r[3] for r in rows[1:] if len(r) > 6 and r[6].strip().lower() == email.lower() and r[3] != email]
                referents.append({'email': email, 'nom': row[0], 'prenom': row[1], 'vendeurs': equipe})
        return jsonify({'success': True, 'referents': referents})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/referents')
@login_required
def get_referents():
    try:
        gc = get_sheets_client()
        if not gc:
            return jsonify({'success': False, 'error': 'Google Sheets non configuré'})
        sheet_id = os.environ.get('GOOGLE_SHEET_ID', '')
        ws = gc.open_by_key(sheet_id).sheet1
        rows = ws.get_all_values()

        referents = []
        referent_emails_from_col_g = set()

        for row in rows[1:]:  # skip header
            if len(row) < 4 or not row[3].strip():
                continue
            email = row[3].strip()
            role = row[9].strip().lower() if len(row) > 9 else ''
            statut = row[10].strip().lower() if len(row) > 10 else 'actif'

            # Collecter les emails référents depuis colonne G
            if len(row) > 6 and row[6].strip() and '@' in row[6]:
                referent_emails_from_col_g.add(row[6].strip().lower())

            # Méthode 1 : colonne J = "referent"
            if role == 'referent' and statut != 'supprime':
                referents.append({
                    'email': email,
                    'nom': row[0].strip(),
                    'prenom': row[1].strip()
                })

        # Méthode 2 fallback : emails trouvés en colonne G mais pas encore dans la liste
        existing_emails = {r['email'].lower() for r in referents}
        for row in rows[1:]:
            if len(row) > 3:
                email = row[3].strip()
                if email.lower() in referent_emails_from_col_g and email.lower() not in existing_emails:
                    referents.append({
                        'email': email,
                        'nom': row[0].strip(),
                        'prenom': row[1].strip()
                    })
                    existing_emails.add(email.lower())

        return jsonify({'success': True, 'referents': referents})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/users', methods=['GET'])
@login_required
def list_users():
    try:
        token = get_zoho_token()
        r = requests.get(
            f'https://mail.zoho.eu/api/organization/{ZOHO_ORG_ID}/accounts',
            headers={'Authorization': f'Zoho-oauthtoken {token}'}
        )
        data = r.json()
        users = data.get('data', [])
        return jsonify({'success': True, 'users': users})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/users', methods=['POST'])
@login_required
def create_user():
    try:
        d = request.get_json()
        prenom = d.get('prenom', '').strip()
        nom = d.get('nom', '').strip()
        poste = d.get('poste', '').strip()
        telephone = d.get('telephone', '').strip()
        password_input = d.get('password', '').strip()
        email_perso = d.get('email_perso', '').strip()
        drive_folder_id = d.get('drive_folder_id', '').strip()
        referent_email = d.get('referent_email', '').strip()
        role = d.get('role', 'vendeur').strip().lower()
        # Générer automatiquement si vide
        password = password_input if password_input else generate_password()

        email_local = f"{prenom.lower()}.{nom.lower()}@liliwatt.fr"
        email_local = email_local.replace('é','e').replace('è','e').replace('ê','e').replace('ë','e')
        email_local = email_local.replace('à','a').replace('â','a').replace('ù','u').replace('û','u')
        email_local = email_local.replace(' ','.').replace("'",'')

        token = get_zoho_token()

        # Créer l'utilisateur
        r = requests.post(
            f'https://mail.zoho.eu/api/organization/{ZOHO_ORG_ID}/accounts',
            headers={'Authorization': f'Zoho-oauthtoken {token}', 'Content-Type': 'application/json'},
            json={
                'firstName': prenom,
                'lastName': nom,
                'primaryEmailAddress': email_local,
                'password': password
            }
        )
        result = r.json()
        print(f"Création utilisateur: {result}")

        if result.get('status', {}).get('code') in [200, 201, '200', '201'] or 'data' in result:
            account_id = result.get('data', {}).get('accountId', '')

            # Appliquer la signature
            if account_id:
                sig_html = make_signature(prenom, nom, poste, telephone, email_local)
                sig_r = requests.post(
                    f'https://mail.zoho.eu/api/accounts/{account_id}/signatures',
                    headers={'Authorization': f'Zoho-oauthtoken {token}', 'Content-Type': 'application/json'},
                    json={'signatureName': 'LILIWATT', 'signature': sig_html, 'isDefault': True},
                    timeout=15
                )
                sig_result = sig_r.json()
                print(f"📝 Signature API response: {sig_result}")
                # Récupérer l'ID de la signature pour la définir par défaut
                sig_id = sig_result.get('data', {}).get('signatureId', '')
                if sig_id:
                    requests.put(
                        f'https://mail.zoho.eu/api/accounts/{account_id}/signatures/{sig_id}',
                        headers={'Authorization': f'Zoho-oauthtoken {token}', 'Content-Type': 'application/json'},
                        json={'signatureName': 'LILIWATT', 'signature': sig_html, 'isDefault': True},
                        timeout=15
                    )
                    print(f"✅ Signature appliquée pour {email_local}")

            # Configurer redirection vers contact@liliwatt.fr
            try:
                token_fwd = get_zoho_token()
                account_id = result.get('data', {}).get('accountId', '')
                if account_id:
                    requests.post(
                        f'https://mail.zoho.eu/api/organization/{ZOHO_ORG_ID}/accounts/{account_id}/settings/forwardingaddress',
                        headers={
                            'Authorization': f'Zoho-oauthtoken {token_fwd}',
                            'Content-Type': 'application/json'
                        },
                        json={
                            'forwardingAddress': 'contact@liliwatt.fr',
                            'keepCopy': True
                        },
                        timeout=15
                    )
                    print(f"✅ Redirection configurée : {email_local} → contact@liliwatt.fr")
            except Exception as e:
                print(f"⚠️ Erreur redirection : {e}")

            # Générer un token RGPD unique
            token_rgpd = uuid.uuid4().hex[:12]

            # Enregistrer dans Google Sheets
            save_to_sheet(prenom, nom, email_local, password, poste, drive_folder_id, referent_email, token_rgpd, role)

            # Créer l'utilisateur dans courtier-energie
            try:
                import jwt as pyjwt
                courtier_url = os.environ.get('COURTIER_API_URL', 'https://liliwatt-courtier.onrender.com')
                courtier_secret = os.environ.get('COURTIER_JWT_SECRET', 'liliwatt-jwt-secret-2026')
                admin_token = pyjwt.encode(
                    {'id': 'admin_liliwatt', 'email': 'johan.mallet@liliwatt.fr', 'role': 'admin', 'exp': datetime.utcnow() + timedelta(hours=2)},
                    courtier_secret, algorithm='HS256'
                )
                courtier_r = requests.post(
                    f'{courtier_url}/api/auth/create-user',
                    headers={'Authorization': f'Bearer {admin_token}', 'Content-Type': 'application/json'},
                    json={
                        'email': email_local,
                        'password': password,
                        'role': 'vendeur',
                        'drive_folder_id': drive_folder_id
                    },
                    timeout=10
                )
                print(f"✅ Utilisateur créé dans courtier-energie: {courtier_r.json()}")
            except Exception as e:
                print(f"⚠️ Erreur création courtier-energie: {e}")

            # Créer l'utilisateur dans le CRM LILIWATT (Neon/Vercel)
            try:
                CRM_URL = os.environ.get('CRM_URL', 'https://liliwatt-crm-8ofi.vercel.app')
                CRM_API_KEY = os.environ.get('CRM_API_KEY', 'liliwatt-crm-api-key-2026')
                lien_rgpd = f'https://liliwatt-courtier.onrender.com/rgpd/{token_rgpd}'
                # Numéro courtier auto
                courtier_number = get_next_courtier_number()
                crm_r = requests.post(
                    f'{CRM_URL}/api/crm/create-user',
                    headers={'X-API-Key': CRM_API_KEY, 'Content-Type': 'application/json'},
                    json={
                        'email': email_local,
                        'firstName': prenom,
                        'lastName': nom,
                        'role': role.upper(),
                        'password': password,
                        'referentEmail': referent_email,
                        'token_rgpd': token_rgpd,
                        'lien_rgpd': lien_rgpd,
                        'zoho_password': password,
                        'courtierNumber': courtier_number
                    },
                    timeout=10
                )
                print(f"✅ CRM LILIWATT: {crm_r.status_code} — {crm_r.json()}")
            except Exception as e:
                print(f"⚠️ Erreur CRM LILIWATT: {e}")

            # Envoyer email de bienvenue
            created_account_id = result.get('data', {}).get('accountId', '')
            send_welcome_email(prenom, nom, email_local, password, poste, telephone, email_perso, created_account_id, token_rgpd, referent_email)

            # Notifier bo@liliwatt.fr avec la signature HTML prête à copier
            try:
                sig_html = make_signature(prenom, nom, poste, telephone, email_local)
                rgpd_link = f'https://liliwatt-courtier.onrender.com/rgpd/{token_rgpd}'
                bo_body = f"""<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;">
  <div style="background:linear-gradient(135deg,#1e1b4b,#7c3aed);padding:24px;border-radius:12px 12px 0 0;text-align:center;">
    <h1 style="color:white;font-size:24px;font-weight:800;letter-spacing:3px;margin:0;">LILIWATT</h1>
    <p style="color:rgba(255,255,255,0.8);margin:6px 0 0;font-size:12px;">Nouveau commercial cr&eacute;&eacute;</p>
  </div>
  <div style="background:#f5f3ff;padding:28px;border-radius:0 0 12px 12px;">
    <p style="font-size:15px;color:#1e1b4b;margin-bottom:20px;"><strong>{prenom} {nom}</strong> a &eacute;t&eacute; ajout&eacute; &agrave; l'&eacute;quipe.</p>

    <div style="background:#fef3c7;border:2px solid #fbbf24;border-radius:10px;padding:20px;margin-bottom:20px;">
      <p style="margin:0 0 12px;font-weight:700;color:#92400e;font-size:14px;">&#128272; Identifiants Zoho Mail</p>
      <table style="width:100%;font-size:14px;border-collapse:collapse;">
        <tr><td style="padding:6px 0;color:#92400e;font-weight:700;width:130px;">Email</td><td style="color:#1e1b4b;font-weight:700;">{email_local}</td></tr>
        <tr><td style="padding:6px 0;color:#92400e;font-weight:700;">Mot de passe</td><td style="color:#1e1b4b;font-weight:700;font-size:16px;">{password}</td></tr>
        <tr><td style="padding:6px 0;color:#92400e;font-weight:700;">Connexion</td><td><a href="https://mail.zoho.eu" style="color:#7c3aed;font-weight:700;text-decoration:none;">mail.zoho.eu</a></td></tr>
      </table>
    </div>

    <div style="background:white;border-radius:10px;padding:20px;margin-bottom:20px;border-left:4px solid #7c3aed;">
      <table style="width:100%;font-size:13px;border-collapse:collapse;">
        <tr><td style="padding:6px 0;color:#6b7280;font-weight:700;width:130px;">Poste</td><td style="color:#1e1b4b;">{poste}</td></tr>
        <tr><td style="padding:6px 0;color:#6b7280;font-weight:700;">T&eacute;l&eacute;phone</td><td style="color:#1e1b4b;">{telephone}</td></tr>
        <tr><td style="padding:6px 0;color:#6b7280;font-weight:700;">R&eacute;f&eacute;rent</td><td style="color:#1e1b4b;">{referent_email or '&mdash;'}</td></tr>
        <tr><td style="padding:6px 0;color:#6b7280;font-weight:700;">Lien RGPD</td><td><a href="{rgpd_link}" style="color:#7c3aed;word-break:break-all;">{rgpd_link}</a></td></tr>
        <tr><td style="padding:6px 0;color:#6b7280;font-weight:700;">Drive</td><td><a href="https://drive.google.com/drive/folders/{drive_folder_id}" style="color:#7c3aed;">Ouvrir le dossier</a></td></tr>
      </table>
    </div>

    <div style="background:#ede9fe;border-radius:10px;padding:16px;margin-bottom:16px;">
      <p style="margin:0 0 10px;font-weight:700;color:#1e1b4b;font-size:13px;">&#9999;&#65039; Signature email pr&ecirc;te &agrave; copier dans Zoho :</p>
      <div style="background:white;border-radius:8px;padding:16px;border:1px solid #e9d5ff;">{sig_html}</div>
    </div>
  </div>
</div>"""
                bo_token = get_zoho_token()
                if bo_token:
                    account_id = os.environ.get('ZOHO_ACCOUNT_ID', '8439060000000002002')
                    requests.post(
                        f'https://mail.zoho.eu/api/accounts/{account_id}/messages',
                        headers={'Authorization': f'Zoho-oauthtoken {bo_token}', 'Content-Type': 'application/json'},
                        json={
                            'fromAddress': 'bo@liliwatt.fr',
                            'toAddress': 'bo@liliwatt.fr',
                            'subject': f'Nouveau commercial : {prenom} {nom} — {poste}',
                            'content': bo_body,
                            'mailFormat': 'html'
                        },
                        timeout=15
                    )
                    print(f"✅ Notification bo@liliwatt.fr envoyée pour {prenom} {nom}")
            except Exception as e:
                print(f"⚠️ Erreur notification bo@: {e}")

            return jsonify({
                'success': True,
                'email': email_local,
                'password': password,
                'message': f'Utilisateur {prenom} {nom} créé avec succès'
            })
        else:
            return jsonify({'success': False, 'error': str(result)})

    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e), 'trace': traceback.format_exc()})

@app.route('/api/users/drive-folder', methods=['GET'])
@login_required
def get_drive_folder():
    email = request.args.get('email', '').strip()
    if not email:
        return jsonify({'success': False, 'error': 'Email manquant'})
    try:
        creds = ServiceAccountCredentials.from_json_keyfile_name(
            CREDENTIALS_FILE,
            ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        )
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(SPREADSHEET_ID)
        ws = sh.sheet1
        rows = ws.get_all_values()
        for row in rows:
            if len(row) >= 4 and row[3].lower() == email.lower():
                drive_folder_id = row[5] if len(row) > 5 else ''
                return jsonify({'success': True, 'drive_folder_id': drive_folder_id})
        return jsonify({'success': False, 'error': 'Vendeur non trouvé'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

RECRUTEMENT_SHEET_ID = '11A-aJIqtm0JZ01lU43GpWudWDNFtknIr-4sYgYD-6ck'

@app.route('/api/recrutement/candidats')
@login_required
def list_candidats():
    try:
        gc = get_sheets_client()
        if not gc:
            return jsonify({'success': False, 'error': 'Sheets non configuré'})
        ws = gc.open_by_key(RECRUTEMENT_SHEET_ID).sheet1
        rows = ws.get_all_values()
        candidats = []
        for i, row in enumerate(rows):
            if i == 0 or not row[2] or '@' not in row[2]:
                continue
            candidats.append({
                'row': i + 1,
                'nom': row[0] if len(row) > 0 else '',
                'prenom': row[1] if len(row) > 1 else '',
                'email': row[2] if len(row) > 2 else '',
                'telephone': row[3] if len(row) > 3 else '',
                'siren': row[4] if len(row) > 4 else '',
                'qualite': row[5] if len(row) > 5 else '',
                'date': row[6] if len(row) > 6 else '',
                'drive_link': row[7] if len(row) > 7 else '',
                'statut': row[8] if len(row) > 8 else 'EN COURS',
                'referant': row[10] if len(row) > 10 else ''
            })
        return jsonify({'success': True, 'candidats': candidats})
    except Exception as e:
        import traceback
        print(f"⚠️ Erreur candidats: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/recrutement/statut', methods=['POST'])
@login_required
def update_statut_candidat():
    try:
        d = request.get_json()
        row_num = d.get('row')
        statut = d.get('statut', '')
        gc = get_sheets_client()
        ws = gc.open_by_key(RECRUTEMENT_SHEET_ID).sheet1
        ws.update_cell(row_num, 9, statut)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/recrutement/referant', methods=['POST'])
@login_required
def update_referant_candidat():
    try:
        d = request.get_json()
        row_num = d.get('row')
        referant = d.get('referant', '')
        gc = get_sheets_client()
        ws = gc.open_by_key(RECRUTEMENT_SHEET_ID).sheet1
        ws.update_cell(row_num, 11, referant)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/recrutement/envoyer-contrat', methods=['POST'])
@login_required
def envoyer_contrat():
    try:
        from generate_contrats import generate_contrats
        d = request.get_json()
        email = d.get('email', '').strip()
        if not email:
            return jsonify({'success': False, 'error': 'Email requis'})

        gc = get_sheets_client()
        ws = gc.open_by_key(RECRUTEMENT_SHEET_ID).sheet1
        rows = ws.get_all_values()
        candidat = None
        for row in rows:
            if len(row) > 2 and row[2].lower() == email.lower():
                candidat = {'nom': row[0], 'prenom': row[1], 'email': row[2],
                           'siren': row[4] if len(row) > 4 else '',
                           'qualite': row[5] if len(row) > 5 else '',
                           'drive_link': row[7] if len(row) > 7 else ''}
                break
        if not candidat:
            return jsonify({'success': False, 'error': 'Candidat non trouvé'})

        # Extraire le folder ID du lien Drive
        import re
        match = re.search(r'folders/([a-zA-Z0-9_-]+)', candidat['drive_link'])
        if not match:
            return jsonify({'success': False, 'error': 'Dossier Drive non trouvé'})
        folder_id = match.group(1)

        print(f"📄 Génération contrats pour {candidat['prenom']} {candidat['nom']}")
        files = generate_contrats(
            candidat['prenom'], candidat['nom'],
            candidat['siren'], candidat['qualite'], folder_id
        )
        print(f"✅ {len(files)} contrat(s) générés")
        return jsonify({'success': True, 'files': files, 'drive_link': candidat['drive_link']})
    except Exception as e:
        import traceback
        print(f"⚠️ Erreur contrat: {e}")
        return jsonify({'success': False, 'error': str(e), 'trace': traceback.format_exc()})

# ===== PHASE 1 — Import CV + Profils =====

@app.route('/api/recrutement/candidats-phase1')
@login_required
def list_candidats_phase1():
    try:
        gc = get_sheets_client()
        if not gc:
            return jsonify({'success': False, 'error': 'Sheets non configuré'})
        sh = gc.open_by_key(RECRUTEMENT_SHEET_ID)
        try:
            ws = sh.worksheet('PHASE 1')
        except Exception:
            ws = sh.add_worksheet(title='PHASE 1', rows=500, cols=10)
            ws.update('A1:J1', [['NOM', 'PRENOM', 'EMAIL', 'TEL', 'ADRESSE', 'STATUT', 'NOTE', 'DATE', 'SESSION', 'LIEN_CV']])
        rows = ws.get_all_values()
        candidats = []
        for i, row in enumerate(rows):
            if i == 0:
                # Log les headers pour debug
                print(f"📋 PHASE 1 headers ({len(row)} cols): {row}")
                continue
            if len(row) < 3 or not row[2]:
                continue
            lien_cv = row[9] if len(row) > 9 else ''
            if not lien_cv and len(row) > 8:
                # Chercher un lien Drive dans toutes les colonnes restantes
                for col_idx in range(8, len(row)):
                    if row[col_idx] and ('drive.google.com' in row[col_idx] or row[col_idx].startswith('http')):
                        lien_cv = row[col_idx]
                        break
            candidats.append({
                'row': i + 1,
                'nom': row[0], 'prenom': row[1], 'email': row[2],
                'telephone': row[3] if len(row) > 3 else '',
                'adresse': row[4] if len(row) > 4 else '',
                'statut': row[5] if len(row) > 5 else 'NON CONTACTÉ',
                'note': row[6] if len(row) > 6 else '',
                'date': row[7] if len(row) > 7 else '',
                'session': row[8] if len(row) > 8 else '',
                'lien_cv': lien_cv
            })
            if lien_cv:
                print(f"  📄 CV trouvé pour {row[1]} {row[0]}: {lien_cv[:60]}...")
        return jsonify({'success': True, 'candidats': candidats})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/recrutement/referents-liste')
@login_required
def referents_liste():
    try:
        gc = get_sheets_client()
        sheet_id = os.environ.get('GOOGLE_SHEET_ID', '')
        ws = gc.open_by_key(sheet_id).sheet1
        rows = ws.get_all_values()
        refs = []
        for row in rows:
            if len(row) > 9 and row[9] in ('referent', 'admin') and '@' in (row[3] or ''):
                refs.append({'nom': row[0], 'prenom': row[1], 'email': row[3]})
        return jsonify({'success': True, 'referents': refs})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/recrutement/phase1/envoyer-referent', methods=['POST'])
@login_required
def envoyer_referent_phase1():
    try:
        d = request.get_json()
        ref_email = d.get('referent_email', '')
        candidat = d.get('candidat', {})
        token = get_zoho_token()
        if not token:
            return jsonify({'success': False, 'error': 'Zoho token non obtenu'})
        mail_html = f"""<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;">
<div style="background:linear-gradient(135deg,#1e1b4b,#7c3aed);padding:24px;border-radius:12px 12px 0 0;text-align:center;">
<h1 style="color:#fff;font-size:24px;letter-spacing:3px;margin:0;">LILIWATT</h1>
<p style="color:rgba(255,255,255,.8);font-size:12px;margin:4px 0 0;">Profil candidat à évaluer</p>
</div>
<div style="background:#f5f3ff;padding:28px;border-radius:0 0 12px 12px;">
<p style="font-size:15px;color:#1e1b4b;">Bonjour,</p>
<p style="color:#374151;line-height:1.7;">Un nouveau profil candidat vous est transmis pour évaluation :</p>
<div style="background:#fff;border-radius:10px;padding:20px;margin:16px 0;border-left:4px solid #7c3aed;">
<table style="width:100%;font-size:13px;border-collapse:collapse;">
<tr><td style="padding:6px 0;color:#6b7280;font-weight:700;width:100px;">Nom</td><td style="color:#1e1b4b;">{candidat.get('prenom','')} {candidat.get('nom','')}</td></tr>
<tr><td style="padding:6px 0;color:#6b7280;font-weight:700;">Email</td><td style="color:#1e1b4b;">{candidat.get('email','')}</td></tr>
<tr><td style="padding:6px 0;color:#6b7280;font-weight:700;">Tél</td><td style="color:#1e1b4b;">{candidat.get('telephone','')}</td></tr>
<tr><td style="padding:6px 0;color:#6b7280;font-weight:700;">Adresse</td><td style="color:#1e1b4b;">{candidat.get('adresse','')}</td></tr>
</table>
</div>
{('<p style="margin:12px 0;"><a href="' + candidat.get('lien_cv','') + '" style="color:#7c3aed;font-weight:600;">📄 Voir le CV</a></p>') if candidat.get('lien_cv') else ''}
<p style="color:#374151;">Lien session Meet : <a href="https://meet.google.com/tzv-pgjc-und?authuser=0" style="color:#7c3aed;font-weight:600;">Rejoindre</a></p>
<hr style="border:1px solid #e9d5ff;margin:20px 0;">
<p style="font-size:11px;color:#9ca3af;">LILIWATT — LILISTRAT STRATÉGIE SAS — 59 rue de Ponthieu, Bureau 326 — 75008 Paris</p>
</div></div>"""
        account_id = os.environ.get('ZOHO_ACCOUNT_ID', '8439060000000002002')
        requests.post(
            f'https://mail.zoho.eu/api/accounts/{account_id}/messages',
            headers={'Authorization': f'Zoho-oauthtoken {token}', 'Content-Type': 'application/json'},
            json={'fromAddress': 'recrutement@liliwatt.fr', 'toAddress': ref_email,
                  'subject': f"📋 Profil candidat — {candidat.get('prenom','')} {candidat.get('nom','')}",
                  'content': mail_html, 'mailFormat': 'html'},
            timeout=15
        )
        print(f"✅ Profil envoyé à {ref_email}: {candidat.get('prenom','')} {candidat.get('nom','')}")
        return jsonify({'success': True})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

def extract_text_from_file(file_bytes, filename):
    """Extrait le texte d'un PDF ou Word."""
    import io
    text = ''
    fname = filename.lower()
    if fname.endswith('.pdf'):
        import pdfplumber
        pdf = pdfplumber.open(io.BytesIO(file_bytes))
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text += t + '\n'
        pdf.close()
    elif fname.endswith('.docx') or fname.endswith('.doc'):
        from docx import Document
        doc = Document(io.BytesIO(file_bytes))
        text = '\n'.join(p.text for p in doc.paragraphs)
    return text.strip()

def extract_cv_with_gpt(text):
    """Appelle GPT-4o-mini pour extraire les infos du CV."""
    import openai
    client = openai.OpenAI(api_key=os.environ.get('OPENAI_API_KEY', ''))
    resp = client.chat.completions.create(
        model='gpt-4o-mini',
        messages=[
            {'role': 'system', 'content': 'Extrais ces informations du CV en JSON : nom, prenom, email, telephone, adresse. Réponds UNIQUEMENT en JSON valide.'},
            {'role': 'user', 'content': text[:4000]}
        ],
        temperature=0
    )
    raw = resp.choices[0].message.content.strip()
    if raw.startswith('```'):
        raw = raw.split('\n', 1)[1].rsplit('```', 1)[0]
    return json.loads(raw)

def save_cv_to_sheet(data, lien_cv=''):
    """Sauvegarde les données extraites dans PHASE 1."""
    gc = get_sheets_client()
    sh = gc.open_by_key(RECRUTEMENT_SHEET_ID)
    try:
        ws = sh.worksheet('PHASE 1')
    except Exception:
        ws = sh.add_worksheet(title='PHASE 1', rows=500, cols=10)
        ws.update('A1:J1', [['NOM', 'PRENOM', 'EMAIL', 'TEL', 'ADRESSE', 'STATUT', 'NOTE', 'DATE', 'SESSION', 'LIEN_CV']])
    import time
    date_str = datetime.now().strftime('%d/%m/%Y')
    tel = data.get('telephone', '') or ''
    if isinstance(tel, list):
        tel = ' / '.join(str(t) for t in tel)
    email = data.get('email', '') or ''
    if isinstance(email, list):
        email = email[0] if email else ''
    row_data = [
        (data.get('nom', '') or '').upper(),
        data.get('prenom', '') or '',
        email,
        tel,
        data.get('adresse', '') or '',
        'NON CONTACTÉ', '', date_str, '', lien_cv
    ]
    for attempt in range(3):
        try:
            ws.append_row(row_data)
            break
        except Exception as e:
            if attempt < 2:
                print(f"⚠️ Sheets retry {attempt+1}/3: {e}")
                time.sleep(2)
            else:
                raise e

RECRUTEMENT_DRIVE_PARENT = '1eQYZqexJ67EcVPe8rsmKDf6mtASf9yjA'

def upload_cv_to_drive(file_bytes, original_filename, prenom, nom):
    """Upload le CV original dans Drive et retourne le lien."""
    try:
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseUpload
        from google.oauth2.service_account import Credentials as SACredentials
        import base64, io

        creds_b64 = os.environ.get('GOOGLE_DRIVE_CREDS_BASE64', '')
        creds_json_env = os.environ.get('GOOGLE_CREDS_JSON', '')
        if creds_b64:
            creds_dict = json.loads(base64.b64decode(creds_b64).decode())
        elif creds_json_env:
            creds_dict = json.loads(creds_json_env)
        else:
            with open(os.path.join(os.path.dirname(__file__), 'liliwatt-eddcc0bc9e18.json')) as fl:
                creds_dict = json.load(fl)
        creds = SACredentials.from_service_account_info(creds_dict, scopes=['https://www.googleapis.com/auth/drive'])
        drive = build('drive', 'v3', credentials=creds)

        # Trouver/créer CANDIDATURES EN COURS
        q = f"'{RECRUTEMENT_DRIVE_PARENT}' in parents and name='CANDIDATURES EN COURS' and mimeType='application/vnd.google-apps.folder' and trashed=false"
        res = drive.files().list(q=q, fields='files(id)', supportsAllDrives=True, includeItemsFromAllDrives=True).execute()
        if res['files']:
            attente_id = res['files'][0]['id']
        else:
            f2 = drive.files().create(body={'name': 'CANDIDATURES EN COURS', 'mimeType': 'application/vnd.google-apps.folder', 'parents': [RECRUTEMENT_DRIVE_PARENT]}, fields='id', supportsAllDrives=True).execute()
            attente_id = f2['id']

        # Trouver/créer dossier candidat
        folder_name = f"{prenom} {nom.upper()}"
        q2 = f"'{attente_id}' in parents and name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
        res2 = drive.files().list(q=q2, fields='files(id)', supportsAllDrives=True, includeItemsFromAllDrives=True).execute()
        if res2['files']:
            cand_id = res2['files'][0]['id']
        else:
            f3 = drive.files().create(body={'name': folder_name, 'mimeType': 'application/vnd.google-apps.folder', 'parents': [attente_id]}, fields='id', supportsAllDrives=True).execute()
            cand_id = f3['id']

        # Upload le fichier
        mime = 'application/pdf' if original_filename.lower().endswith('.pdf') else 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        media = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype=mime)
        uploaded = drive.files().create(
            body={'name': original_filename, 'parents': [cand_id], 'mimeType': mime},
            media_body=media, fields='id, webViewLink', supportsAllDrives=True
        ).execute()
        print(f"📁 CV uploadé Drive: {original_filename} → {uploaded.get('webViewLink','')}")
        return uploaded.get('webViewLink', '')
    except Exception as e:
        print(f"⚠️ Erreur upload CV Drive: {e}")
        return ''

@app.route('/api/recrutement/upload-cv', methods=['POST'])
@login_required
def upload_cv():
    try:
        if 'cv' not in request.files:
            return jsonify({'success': False, 'error': 'Fichier CV requis'})
        f = request.files['cv']
        filename = f.filename
        file_bytes = f.read()

        # ZIP : extraire et traiter chaque fichier
        if filename.endswith('.zip'):
            import zipfile, io
            results = {'total': 0, 'ok': 0, 'errors': 0}
            with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
                all_names = zf.namelist()
                print(f"📦 ZIP ouvert: {len(all_names)} entrées")
                for name in all_names:
                    # Ignorer fichiers cachés macOS et dossiers
                    basename = name.split('/')[-1]
                    if basename.startswith('.') or basename.startswith('__') or not basename:
                        continue
                    if not basename.lower().endswith(('.pdf', '.docx', '.doc')):
                        print(f"  ⏭️ Ignoré (format): {name}")
                        continue
                    print(f"  📄 Trouvé: {name} ({zf.getinfo(name).file_size} octets)")
                    results['total'] += 1
                    try:
                        inner = zf.read(name)
                        text = extract_text_from_file(inner, basename)
                        if not text:
                            print(f"  ⚠️ Pas de texte extrait: {name}")
                            results['errors'] += 1; continue
                        data = extract_cv_with_gpt(text)
                        lien = upload_cv_to_drive(inner, basename, data.get('prenom', ''), data.get('nom', ''))
                        save_cv_to_sheet(data, lien)
                        results['ok'] += 1
                        print(f"  ✅ {name}: {data.get('nom','')} {data.get('prenom','')} ({data.get('email','')})")
                    except Exception as e:
                        results['errors'] += 1
                        print(f"  ❌ {name}: {e}")
            print(f"📦 ZIP terminé: {results['total']} traités, {results['ok']} ok, {results['errors']} erreurs")
            return jsonify({'success': True, 'zip': True, **results})

        # Fichier unique
        text = extract_text_from_file(file_bytes, filename)
        if not text:
            return jsonify({'success': False, 'error': 'Impossible d\'extraire le texte du CV'})

        data = extract_cv_with_gpt(text)
        print(f"📄 CV extrait: {data}")
        lien_cv = upload_cv_to_drive(file_bytes, filename, data.get('prenom', ''), data.get('nom', ''))
        save_cv_to_sheet(data, lien_cv)
        return jsonify({'success': True, 'data': data, 'lien_cv': lien_cv})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/recrutement/phase1/statut', methods=['POST'])
@login_required
def update_phase1_statut():
    try:
        d = request.get_json()
        gc = get_sheets_client()
        ws = gc.open_by_key(RECRUTEMENT_SHEET_ID).worksheet('PHASE 1')
        ws.update_cell(d['row'], 6, d['statut'])
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/recrutement/phase1/note', methods=['POST'])
@login_required
def update_phase1_note():
    try:
        d = request.get_json()
        gc = get_sheets_client()
        ws = gc.open_by_key(RECRUTEMENT_SHEET_ID).worksheet('PHASE 1')
        ws.update_cell(d['row'], 7, d['note'])
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/recrutement/phase1/inviter', methods=['POST'])
@login_required
def inviter_phase1():
    try:
        d = request.get_json()
        email = d.get('email', '')
        prenom = d.get('prenom', '')
        date_session = d.get('date_session', '')
        heure_session = d.get('heure_session', '')
        row_num = d.get('row')

        # Envoyer le mail d'invitation
        mail_html = f"""<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;">
<div style="background:linear-gradient(135deg,#1e1b4b,#7c3aed);padding:32px;border-radius:12px 12px 0 0;text-align:center;">
<h1 style="color:#fff;font-size:28px;font-weight:800;letter-spacing:3px;margin:0;">LILIWATT</h1>
<p style="color:rgba(255,255,255,.8);font-size:12px;margin:6px 0 0;">Invitation session de présentation</p>
</div>
<div style="background:#f5f3ff;padding:32px;border-radius:0 0 12px 12px;">
<p style="font-size:16px;color:#1e1b4b;">Bonjour <strong>{prenom}</strong>,</p>
<p style="color:#374151;line-height:1.7;">Suite à notre échange, nous avons le plaisir de vous inviter à rejoindre notre session de présentation LILIWATT.</p>
<div style="background:#fff;border-radius:10px;padding:24px;margin:24px 0;border-left:4px solid #7c3aed;">
<table style="width:100%;font-size:14px;border-collapse:collapse;">
<tr><td style="padding:8px 0;color:#6b7280;font-weight:700;width:100px;">Date</td><td style="color:#1e1b4b;font-weight:700;">{date_session}</td></tr>
<tr><td style="padding:8px 0;color:#6b7280;font-weight:700;">Heure</td><td style="color:#1e1b4b;font-weight:700;">{heure_session}</td></tr>
</table>
</div>
<a href="https://meet.google.com/tzv-pgjc-und?authuser=0" style="display:inline-block;background:linear-gradient(135deg,#7c3aed,#d946ef);color:#fff;padding:14px 32px;border-radius:50px;text-decoration:none;font-weight:700;font-size:14px;">Rejoindre la session Google Meet</a>
<p style="color:#374151;margin-top:20px;line-height:1.7;">À très bientôt !</p>
<p style="color:#6b7280;font-size:13px;">L'équipe LILIWATT<br>recrutement@liliwatt.fr</p>
<hr style="border:1px solid #e9d5ff;margin:24px 0;">
<p style="font-size:11px;color:#9ca3af;">LILIWATT — LILISTRAT STRATÉGIE SAS — 59 rue de Ponthieu, Bureau 326 — 75008 Paris</p>
</div></div>"""

        token = get_zoho_token()
        if token:
            account_id = os.environ.get('ZOHO_ACCOUNT_ID', '8439060000000002002')
            requests.post(
                f'https://mail.zoho.eu/api/accounts/{account_id}/messages',
                headers={'Authorization': f'Zoho-oauthtoken {token}', 'Content-Type': 'application/json'},
                json={'fromAddress': 'recrutement@liliwatt.fr', 'toAddress': email,
                      'subject': f'Invitation session LILIWATT — {date_session} à {heure_session}',
                      'content': mail_html, 'mailFormat': 'html'},
                timeout=15
            )
            print(f"✅ Invitation envoyée à {email}")

        # Mettre à jour Sheets
        gc = get_sheets_client()
        ws = gc.open_by_key(RECRUTEMENT_SHEET_ID).worksheet('PHASE 1')
        ws.update_cell(row_num, 6, 'CONTACTÉ')
        ws.update_cell(row_num, 9, f'{date_session} {heure_session}')

        return jsonify({'success': True})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/recrutement/profil/<int:row_id>', methods=['DELETE'])
@login_required
def delete_phase1_profil(row_id):
    try:
        d = request.get_json() or {}
        drive_file_id = d.get('drive_file_id', '')

        # 1. Supprimer la ligne dans le Sheet PHASE 1
        gc = get_sheets_client()
        ws = gc.open_by_key(RECRUTEMENT_SHEET_ID).worksheet('PHASE 1')
        ws.delete_rows(row_id)
        print(f"✅ Ligne {row_id} supprimée du Sheet PHASE 1")

        # 2. Supprimer le fichier CV dans le Drive si fourni
        if drive_file_id:
            try:
                from googleapiclient.discovery import build
                from google.oauth2.service_account import Credentials as SACredentials
                import base64

                creds_b64 = os.environ.get('GOOGLE_DRIVE_CREDS_BASE64', '')
                creds_json_env = os.environ.get('GOOGLE_CREDS_JSON', '')
                if creds_b64:
                    creds_dict = json.loads(base64.b64decode(creds_b64).decode())
                elif creds_json_env:
                    creds_dict = json.loads(creds_json_env)
                else:
                    creds_dict = None

                if creds_dict:
                    creds = SACredentials.from_service_account_info(
                        creds_dict,
                        scopes=['https://www.googleapis.com/auth/drive']
                    )
                    drive_service = build('drive', 'v3', credentials=creds)
                    drive_service.files().delete(fileId=drive_file_id).execute()
                    print(f"✅ Fichier Drive {drive_file_id} supprimé")
            except Exception as e:
                print(f"⚠️ Erreur suppression Drive: {e}")

        return jsonify({'success': True})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

# ===== DEROGATION OHM =====
SUIVI_VENTES_SHEET_ID = os.environ.get('SUIVI_VENTES_SHEET_ID', '1Ld1Zl3qVzdVZsyksdfxYfL1LiVcFd5BEbrPV6NYLfcA')

@app.route('/api/derogation/soumettre', methods=['POST'])
@login_required
def derogation_soumettre():
    try:
        data = request.json
        nom = data.get('nom', '').strip()
        siren = data.get('siren', '').strip()
        date_demarrage = data.get('date', '').strip()
        energie = data.get('energie', '').strip()
        volume = data.get('volume', '').strip()
        score = data.get('score', '').strip()
        commentaire = data.get('commentaire', '').strip()

        if not all([nom, siren, date_demarrage, energie, volume, score]):
            return jsonify({'success': False, 'error': 'Champs manquants'}), 400

        date_obj = datetime.strptime(date_demarrage, '%Y-%m-%d')
        date_fr = date_obj.strftime('%d/%m/%Y')
        date_soumission = datetime.now().strftime('%d/%m/%Y %H:%M')

        # 1. Append Sheet DEROGATIONS (best effort)
        try:
            sh = get_sheets_client().open_by_key(SUIVI_VENTES_SHEET_ID)
            ws = sh.worksheet('DEROGATIONS')
            ws.append_row([
                date_soumission, nom, siren, date_fr, energie,
                volume, score, commentaire, 'En attente'
            ])
            print(f'[DEROGATION] Sheet OK — {nom} ({siren})')
        except Exception as e:
            print(f'[DEROGATION] Erreur Sheet: {e}')

        # 2. Mail Zoho a bo@liliwatt.fr
        token = get_zoho_token()
        if not token:
            return jsonify({'success': False, 'error': 'Token Zoho indisponible'}), 500

        commentaire_block = f'<div style="background:#fef3c7;border:1px solid #fbbf24;padding:14px 18px;margin:14px 0;border-radius:6px;"><p style="margin:0 0 6px;font-weight:700;color:#92400e;font-size:13px;">COMMENTAIRE</p><p style="margin:0;color:#92400e;font-size:13px;line-height:1.5;">{commentaire}</p></div>' if commentaire else ''

        html_body = f'''<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;background:#fff;">
<div style="background:linear-gradient(135deg,#7c3aed,#d946ef);padding:24px;text-align:center;color:white;border-radius:8px 8px 0 0;">
<h2 style="margin:0;font-size:20px;">🔍 Demande de dérogation</h2>
<p style="margin:6px 0 0;font-size:13px;opacity:0.9;">Score client en dessous du seuil</p>
</div>
<div style="padding:24px;background:#fafafa;border:1px solid #e5e7eb;border-top:none;border-radius:0 0 8px 8px;">
<p style="margin:0 0 16px;color:#374151;">Bonjour Aya,</p>
<p style="margin:0 0 20px;color:#374151;">Suite à un score client défavorable, nous demandons une dérogation pour le dossier suivant :</p>
<div style="background:white;border-left:3px solid #7c3aed;padding:14px 18px;margin:14px 0;border-radius:4px;">
<p style="margin:0 0 8px;font-weight:700;color:#7c3aed;font-size:13px;">ENTREPRISE</p>
<p style="margin:3px 0;color:#374151;font-size:14px;">Nom : <strong>{nom}</strong></p>
<p style="margin:3px 0;color:#374151;font-size:14px;">SIREN : {siren}</p>
<p style="margin:3px 0;color:#374151;font-size:14px;">Date de démarrage : {date_fr}</p>
</div>
<div style="background:white;border-left:3px solid #d946ef;padding:14px 18px;margin:14px 0;border-radius:4px;">
<p style="margin:0 0 8px;font-weight:700;color:#d946ef;font-size:13px;">CONTRAT</p>
<p style="margin:3px 0;color:#374151;font-size:14px;">Énergie : <strong>{energie}</strong></p>
<p style="margin:3px 0;color:#374151;font-size:14px;">Volume annuel : {volume} kWh</p>
<p style="margin:3px 0;color:#374151;font-size:14px;">Score : {score}</p>
</div>
{commentaire_block}
<p style="margin:20px 0 0;color:#374151;">Merci de revenir vers moi pour validation ou refus.</p>
<div style="margin-top:24px;padding-top:16px;border-top:1px solid #e5e7eb;">
<p style="margin:0;color:#374151;font-size:14px;font-weight:600;">Bien cordialement,</p>
<p style="margin:4px 0 0;color:#1e1b4b;font-size:14px;font-weight:700;">Johan MALLET</p>
<p style="margin:0;color:#6b7280;font-size:12px;">Directeur Général · LILIWATT</p>
<p style="margin:0;color:#6b7280;font-size:12px;">📧 johan.mallet@liliwatt.fr</p>
</div>
<p style="margin:20px 0 0;color:#9ca3af;font-size:11px;text-align:center;">— Système LILIWATT</p>
</div>
</div>'''

        account_id = os.environ.get('ZOHO_ACCOUNT_ID', '8439060000000002002')
        resp = requests.post(
            f'https://mail.zoho.eu/api/accounts/{account_id}/messages',
            headers={
                'Authorization': f'Zoho-oauthtoken {token}',
                'Content-Type': 'application/json'
            },
            json={
                'fromAddress': 'bo@liliwatt.fr',
                'toAddress': 'aya.benchikar@ohm-energie.com',
                'ccAddress': 'bo@liliwatt.fr,johan.mallet@liliwatt.fr',
                'replyTo': 'johan.mallet@liliwatt.fr',
                'subject': f'🔍 Demande de dérogation — {nom} (SIREN {siren})',
                'content': html_body,
                'mailFormat': 'html'
            },
            timeout=15
        )

        if resp.status_code not in (200, 201):
            print(f'[DEROGATION] Mail erreur Zoho: {resp.status_code} {resp.text}')
            return jsonify({'success': False, 'error': 'Envoi mail echec'}), 500

        print(f'[DEROGATION] Mail OK — {nom} ({siren})')
        return jsonify({'success': True})

    except Exception as e:
        print(f'[DEROGATION] Erreur: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500

# ===== GENERATION CONTRATS =====

VENDEURS_PARENT_ID = '157Sol6u32W0loIEv8CmYT3uoDaGyZ7q6'

CONTRAT_TEMPLATES = {
    'VENDEUR': [
        ('Contrat Partenariat', '1ll4JmsEKBusTafQfaxbe96dU2u4kmE7gxVG9ShPXj-k'),
        ('Avenant 1 Remuneration', '1o3lGOSX4BUGywBXMJewbhRfDLnbB_5SvqZc3lhQBxik'),
    ],
    'REFERENT': [
        ('Avenant 2 Referent', '1XvLhgnMGo4ep8O2UgAZKs1PaIPfxwmeemLR2FLSdWMk'),
    ]
}

MOIS_FR = ['janvier', 'février', 'mars', 'avril', 'mai', 'juin',
           'juillet', 'août', 'septembre', 'octobre', 'novembre', 'décembre']

def get_drive_docs_services():
    """Build Drive v3 + Docs v1 services avec les bons scopes"""
    import base64
    from googleapiclient.discovery import build
    from google.oauth2.service_account import Credentials as SACredentials

    creds_b64 = os.environ.get('GOOGLE_DRIVE_CREDS_BASE64', '')
    creds_json_env = os.environ.get('GOOGLE_CREDS_JSON', '')
    if creds_b64:
        creds_dict = json.loads(base64.b64decode(creds_b64).decode())
    elif creds_json_env:
        creds_dict = json.loads(creds_json_env)
    else:
        with open(os.path.join(os.path.dirname(__file__), 'liliwatt-eddcc0bc9e18.json')) as f:
            creds_dict = json.load(f)

    creds = SACredentials.from_service_account_info(creds_dict, scopes=[
        'https://www.googleapis.com/auth/drive',
        'https://www.googleapis.com/auth/documents',
        'https://www.googleapis.com/auth/spreadsheets',
    ])
    drive_svc = build('drive', 'v3', credentials=creds)
    docs_svc = build('docs', 'v1', credentials=creds)
    return drive_svc, docs_svc

def find_or_create_folder(drive, name, parent_id):
    """Cherche un dossier par nom sous parent_id, le cree si absent"""
    q = f"name='{name}' and '{parent_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
    r = drive.files().list(q=q, supportsAllDrives=True, includeItemsFromAllDrives=True, fields='files(id,name)').execute()
    if r.get('files'):
        return r['files'][0]['id']
    new_folder = drive.files().create(body={
        'name': name,
        'mimeType': 'application/vnd.google-apps.folder',
        'parents': [parent_id]
    }, supportsAllDrives=True, fields='id').execute()
    return new_folder['id']

@app.route('/api/contrats/generer', methods=['POST'])
@login_required
def contrats_generer():
    try:
        import io
        from googleapiclient.http import MediaIoBaseUpload

        data = request.json
        type_contrat = data.get('type', '').strip()
        civilite = data.get('civilite', '').strip()
        prenom = data.get('prenom', '').strip()
        nom = data.get('nom', '').strip().upper()
        entreprise = data.get('entreprise', '').strip().upper()
        adresse = data.get('adresse', '').strip()
        siren = data.get('siren', '').strip()
        date_signature = data.get('date_signature', '').strip()
        email_signataire = data.get('email_signataire', '').strip()

        if not all([type_contrat, civilite, prenom, nom, entreprise, siren, date_signature, email_signataire]):
            return jsonify({'success': False, 'error': 'Champs obligatoires manquants'}), 400
        if type_contrat == 'VENDEUR' and not adresse:
            return jsonify({'success': False, 'error': 'Adresse requise pour un contrat vendeur'}), 400

        templates = CONTRAT_TEMPLATES.get(type_contrat)
        if not templates:
            return jsonify({'success': False, 'error': 'Type invalide'}), 400

        # Format date FR : "29 avril 2026"
        date_obj = datetime.strptime(date_signature, '%Y-%m-%d')
        date_fr = f"{date_obj.day} {MOIS_FR[date_obj.month - 1]} {date_obj.year}"

        nom_complet = f"{prenom.capitalize()} {nom}"

        drive, docs = get_drive_docs_services()

        # Trouver/creer dossier vendeur + sous-dossier Contrats
        vendeur_folder_id = find_or_create_folder(drive, f"{prenom.capitalize()} {nom}", VENDEURS_PARENT_ID)
        contrats_folder_id = find_or_create_folder(drive, 'Contrats', vendeur_folder_id)

        liens_pdf = []

        for nom_template, template_id in templates:
            # 1. Copier le template Google Doc directement dans Contrats/
            copy_name = f"{nom_template} - {prenom.capitalize()} {nom}"
            copy = drive.files().copy(
                fileId=template_id,
                body={'name': copy_name, 'parents': [contrats_folder_id]},
                supportsAllDrives=True,
                fields='id, webViewLink'
            ).execute()
            copy_id = copy['id']
            gdoc_link = copy.get('webViewLink', '')

            # 2. Remplacer les marqueurs
            replacements = [
                {'replaceAllText': {'containsText': {'text': '[[CIVILITE]]', 'matchCase': True}, 'replaceText': civilite}},
                {'replaceAllText': {'containsText': {'text': '[[PRENOM_NOM]]', 'matchCase': True}, 'replaceText': nom_complet}},
                {'replaceAllText': {'containsText': {'text': '[[ENTREPRISE]]', 'matchCase': True}, 'replaceText': entreprise}},
                {'replaceAllText': {'containsText': {'text': '[[ADRESSE]]', 'matchCase': True}, 'replaceText': adresse or ''}},
                {'replaceAllText': {'containsText': {'text': '[[SIREN]]', 'matchCase': True}, 'replaceText': siren}},
                {'replaceAllText': {'containsText': {'text': '[[DATE]]', 'matchCase': True}, 'replaceText': date_fr}},
            ]
            docs.documents().batchUpdate(documentId=copy_id, body={'requests': replacements}).execute()

            # 3. Export PDF
            pdf_bytes = drive.files().export(fileId=copy_id, mimeType='application/pdf').execute()

            # 4. Upload PDF dans Contrats/
            pdf_filename = f"{nom_template} - {prenom.capitalize()} {nom}.pdf"
            media = MediaIoBaseUpload(io.BytesIO(pdf_bytes), mimetype='application/pdf')
            pdf_file = drive.files().create(
                body={'name': pdf_filename, 'parents': [contrats_folder_id]},
                media_body=media,
                fields='id, webViewLink',
                supportsAllDrives=True
            ).execute()

            liens_pdf.append({
                'nom': nom_template,
                'lien_pdf': pdf_file.get('webViewLink', ''),
                'lien_gdoc': gdoc_link
            })
            print(f'[CONTRATS] PDF + Google Doc crees: {copy_name}')

        # 6. Log Sheet CONTRATS_GENERES (best effort)
        try:
            sh = get_sheets_client().open_by_key(SUIVI_VENTES_SHEET_ID)
            ws = sh.worksheet('CONTRATS_GENERES')
            liens_str = ' | '.join([f"{l['nom']}: GDoc={l['lien_gdoc']} | PDF={l['lien_pdf']}" for l in liens_pdf])
            ws.append_row([
                datetime.now().strftime('%d/%m/%Y %H:%M'),
                type_contrat, civilite, prenom.capitalize(), nom, entreprise,
                siren, email_signataire, date_fr, liens_str
            ])
        except Exception as sheet_err:
            print(f'[CONTRATS] Sheet log warning: {sheet_err}')

        # 7. Mail a johan.mallet@liliwatt.fr
        token = get_zoho_token()
        if token:
            docs_links = ''.join([
                f'<tr><td style="padding:8px 0 4px;"><p style="margin:0 0 4px;font-weight:700;color:#1e1b4b;font-size:14px;">{l["nom"]}</p><p style="margin:0 0 4px;"><a href="{l["lien_gdoc"]}" style="color:#7c3aed;font-weight:600;text-decoration:none;font-size:13px;">🖋️ Signer le document (Google Doc)</a></p><p style="margin:0 0 8px;"><a href="{l["lien_pdf"]}" style="color:#6b7280;font-weight:500;text-decoration:none;font-size:12px;">📄 PDF (archive)</a></p></td></tr>'
                for l in liens_pdf
            ])
            html_mail = f'''<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;background:#fff;">
<div style="background:linear-gradient(135deg,#7c3aed,#d946ef);padding:24px;text-align:center;color:white;border-radius:8px 8px 0 0;">
<h2 style="margin:0;font-size:20px;">📄 Contrats générés</h2>
<p style="margin:6px 0 0;font-size:13px;opacity:0.9;">Prêts à envoyer pour signature</p>
</div>
<div style="padding:24px;background:#fafafa;border:1px solid #e5e7eb;border-top:none;border-radius:0 0 8px 8px;">
<div style="background:white;border-left:3px solid #7c3aed;padding:14px 18px;margin:0 0 16px;border-radius:4px;">
<p style="margin:0 0 8px;font-weight:700;color:#7c3aed;font-size:13px;">DESTINATAIRE</p>
<p style="margin:3px 0;color:#374151;font-size:14px;">{civilite} {nom_complet}</p>
<p style="margin:3px 0;color:#374151;font-size:14px;">Entreprise : {entreprise}</p>
<p style="margin:3px 0;color:#374151;font-size:14px;">Email : {email_signataire}</p>
</div>
<div style="background:white;border-left:3px solid #d946ef;padding:14px 18px;margin:0 0 16px;border-radius:4px;">
<p style="margin:0 0 8px;font-weight:700;color:#d946ef;font-size:13px;">DOCUMENTS GÉNÉRÉS</p>
<table style="width:100%;">{docs_links}</table>
</div>
<div style="background:#fef3c7;border:1px solid #fbbf24;padding:14px 18px;margin:0 0 16px;border-radius:6px;">
<p style="margin:0 0 6px;font-weight:700;color:#92400e;font-size:13px;">⚠️ ACTION REQUISE</p>
<p style="margin:0;color:#92400e;font-size:13px;line-height:1.5;">Cliquer sur 'Signer le document' pour ouvrir le Google Doc, puis menu Outils &gt; Signatures électroniques &gt; envoyer au signataire : <strong>{email_signataire}</strong></p>
</div>
<div style="margin-top:24px;padding-top:16px;border-top:1px solid #e5e7eb;">
<p style="margin:0;color:#374151;font-size:14px;font-weight:600;">Bien cordialement,</p>
<p style="margin:4px 0 0;color:#1e1b4b;font-size:14px;font-weight:700;">Johan MALLET</p>
<p style="margin:0;color:#6b7280;font-size:12px;">Directeur Général · LILIWATT</p>
<p style="margin:0;color:#6b7280;font-size:12px;">📧 johan.mallet@liliwatt.fr</p>
</div>
<p style="margin:20px 0 0;color:#9ca3af;font-size:11px;text-align:center;">— Système LILIWATT</p>
</div>
</div>'''

            try:
                account_id = os.environ.get('ZOHO_ACCOUNT_ID', '8439060000000002002')
                requests.post(
                    f'https://mail.zoho.eu/api/accounts/{account_id}/messages',
                    headers={'Authorization': f'Zoho-oauthtoken {token}', 'Content-Type': 'application/json'},
                    json={
                        'fromAddress': 'bo@liliwatt.fr',
                        'toAddress': 'johan.mallet@liliwatt.fr',
                        'ccAddress': 'bo@liliwatt.fr',
                        'replyTo': 'johan.mallet@liliwatt.fr',
                        'subject': f'📄 Contrats générés pour {prenom.capitalize()} {nom} — prêts à signer',
                        'content': html_mail,
                        'mailFormat': 'html'
                    },
                    timeout=15
                )
                print(f'[CONTRATS] Mail envoye a johan.mallet@liliwatt.fr')
            except Exception as mail_err:
                print(f'[CONTRATS] Mail warning: {mail_err}')

        return jsonify({'success': True, 'liens': liens_pdf})

    except Exception as e:
        print(f'[CONTRATS] Erreur: {e}')
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# ===== EXTRACTEUR DE CONTRATS PDF =====

_CONTRAT_SYSTEM_PROMPT = """Tu es un extracteur de données contractuelles pour des contrats d'énergie (gaz ou électricité).
Tu reçois le texte brut et les tableaux sérialisés d'un PDF de contrat.
Réponds UNIQUEMENT avec un objet JSON valide. Pas de texte autour, pas de balise markdown, pas de backtick.
Un champ introuvable vaut null. Ne pas inventer de valeur.

Règles d'extraction :

ref_client : null. La référence client (MB-...) provient du back-office, pas du CPV. Ne pas la chercher dans le contrat.

ref_vente : Numéro MIB, format "MIB-XXXXXXXXXX". Il apparaît souvent sous la forme "MIB-XXXXXXXXXX: Offre du..." — extraire uniquement le numéro, sans les deux-points ni ce qui suit.

siren : Numéro SIREN à 9 chiffres du CLIENT (pas du fournisseur). Se trouve dans le paragraphe décrivant le client, après "immatriculée sous le SIREN". NE PAS confondre avec le SIRET (14 chiffres) qui apparaît dans le tableau des sites.

adresse_client : Adresse du siège social du CLIENT. Se trouve dans le paragraphe "dont le siège social est situé à". Prendre l'adresse complète.

date_signature : Date de signature du contrat au format AAAA-MM-JJ. Chercher "Fait à ... le <jour> <mois en lettres> <année>". Convertir le mois français en numéro.

societe : Raison sociale du CLIENT. Dans le contrat, chercher le paragraphe qui commence par "Et" ou "d'autre part" — c'est le nom qui suit immédiatement, avant la virgule. Ne PAS prendre le nom du fournisseur (qui est entre "Entre" et "D'une part").

periode : Mois de SIGNATURE du contrat, format AAAA-MM (ex: 2026-07). Chercher "Fait à ... le <jour> <mois> <année>" dans le corps du contrat. Convertir le nom du mois français en numéro (janvier=01, février=02, mars=03, avril=04, mai=05, juin=06, juillet=07, août=08, septembre=09, octobre=10, novembre=11, décembre=12).

date_debut : Date de DÉBUT de fourniture au format AAAA-MM-JJ (ex: 2026-08-10). Se trouve dans le tableau des sites, colonne "Date de début". Convertir depuis JJ/MM/AAAA si nécessaire.

date_fin : Date de FIN de fourniture au format AAAA-MM-JJ (ex: 2030-12-31). Chercher la phrase "se termine le JJ/MM/AAAA". Convertir depuis JJ/MM/AAAA si nécessaire.

ATTENTION : periode, date_debut et date_fin sont trois dates différentes — ne jamais les confondre.

type_energie : "gaz" si c'est un contrat de gaz naturel, "elec" si c'est un contrat d'électricité. Déduire du titre du contrat.

pdl_pce : Numéro(s) à 14 chiffres figurant dans la colonne "N° du Point de Comptage et d'Estimation (PCE)". ATTENTION : le SIRET fait aussi 14 chiffres mais se trouve dans une colonne distincte intitulée "SIRET du site" — ne pas les confondre. Si plusieurs sites, concaténer tous les PCE avec " / ".

fournisseur : Nom du fournisseur d'énergie. Se trouve entre "Entre" et "D'une part" dans l'en-tête du contrat.

segment : null. Le segment est déduit automatiquement par l'application, ne pas le remplir.

nom_client : NOM (en majuscules) de la ligne du tableau Contacts dont la colonne Type contient exactement "Le signataire du contrat". Identifier la ligne par ce libellé, pas par sa position dans le tableau.

prenom_client : PRÉNOM de cette même ligne.

tel_client : Téléphone de cette même ligne (chiffres uniquement, sans espace ni tiret).

email_client : Email de cette même ligne.

volume_gaz : Si type_energie est "gaz" — volume annuel de fourniture en MWh (nombre décimal), lu dans la colonne TOTAL du tableau des sous-périodes.
Compare la date de début de fourniture à chaque sous-période. Toute sous-période dont l'année correspond à une fourniture partielle (début en cours d'année) doit être ignorée. Retiens la première sous-période couvrant une année civile entière.
Exemple : début 10/08/2026 → la sous-période 2026 est partielle (seulement quelques mois), IGNORER. Retenir la sous-période 2027 (première année complète).
Si plusieurs sites, additionner les TOTAL de chaque site pour cette même sous-période. Renvoyer uniquement la valeur numérique (ex: 14.4). null si type_energie est "elec".

volume_elec : Même règle exacte pour un contrat d'électricité — colonne TOTAL, ignorer la sous-période partielle, retenir la première année complète. Exemple : début 06/08/2026 → ignorer 2026 (38.3 MWh, partiel), retenir 2027 (95.1 MWh, année complète). null si type_energie est "gaz".

Tu ne dois jamais inventer, deviner ou approximer une valeur absente du contrat. Si une donnée ne figure pas dans le document, renvoie null et ajoute un avertissement.
En revanche, additionner des volumes qui figurent LITTÉRALEMENT dans le contrat est autorisé et attendu : sur un contrat multi-sites, le volume retenu est la somme des volumes annuels de tous les sites, sur une année complète. Cette somme porte uniquement sur des valeurs lues dans le document.

Champs toujours null (ne pas extraire) : vendeur, referent, montant_ht, commission_vendeur, commission_referent, statut_paiement, date_paiement_1, date_paiement_2, lien_drive."""


def _contrat_extraire_contenu(pdf_bytes):
    """Extrait texte brut + tableaux des 8 premières pages (traitement en mémoire).

    Renvoie une chaîne structurée. Lève ValueError si le PDF est sans texte.
    """
    import pdfplumber
    import io as _io

    parties = []
    total_chars = 0

    with pdfplumber.open(_io.BytesIO(pdf_bytes)) as pdf:
        for i, page in enumerate(pdf.pages[:8], start=1):
            texte = page.extract_text() or ''
            total_chars += len(texte)
            parties.append(f'=== PAGE {i} / TEXTE ===\n{texte}')

            for j, table in enumerate(page.extract_tables(), start=1):
                lignes = []
                for row in table:
                    cells = [str(c).strip() if c else '' for c in row]
                    lignes.append(' | '.join(cells))
                parties.append(f'=== PAGE {i} / TABLEAU {j} ===\n' + '\n'.join(lignes))

    if total_chars < 200:
        raise ValueError('PDF sans texte exploitable (document scanné ?) — saisie manuelle nécessaire.')

    return '\n\n'.join(parties)


def _contrat_analyser_tableaux(pdf_bytes):
    """Analyse les tableaux du CPV pour déduire le SEGMENT et l'adresse du premier site.

    Retourne {'segment': str|None, 'adresse_site': str|None}.
    """
    import pdfplumber
    import io as _io

    segments_sites = []
    cadrans_volumes = set()
    est_gaz = False
    adresse_premier_site = None

    with pdfplumber.open(_io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages[:8]:
            for table in page.extract_tables():
                if not table or len(table) < 2:
                    continue
                headers = [str(c or '').strip().upper() for c in table[0]]

                # Détecter le tableau des sites (colonne "Nom et adresse du Site")
                addr_col = None
                for ci, h in enumerate(headers):
                    if 'NOM' in h and 'ADRESSE' in h and 'SITE' in h:
                        addr_col = ci
                        break

                if addr_col is not None and adresse_premier_site is None:
                    for row in table[1:]:
                        val = str(row[addr_col] or '').strip()
                        if val:
                            adresse_premier_site = ' '.join(val.replace('\n', ' ').split())
                            break

                # Tableau des sites ELEC : colonne "Segment (C4 ou C5)"
                seg_col = None
                for ci, h in enumerate(headers):
                    if 'SEGMENT' in h and ('C4' in h or 'C5' in h):
                        seg_col = ci
                        break

                if seg_col is not None:
                    for row in table[1:]:
                        val = str(row[seg_col] or '').strip().upper()
                        if val in ('C5', 'C4', 'C2'):
                            segments_sites.append(val)

                # Tableau des sites GAZ : colonne "Profil"
                profil_col = None
                for ci, h in enumerate(headers):
                    if 'PROFIL' in h and 'SEGMENT' not in h:
                        profil_col = ci
                        break

                if profil_col is not None:
                    for row in table[1:]:
                        val = str(row[profil_col] or '').strip().upper()
                        if val.startswith('P0'):
                            est_gaz = True

                # Tableau des volumes : détecter les cadrans
                if any('SEGMENT C' in h or 'PROFILE P' in h for h in headers):
                    for h in headers:
                        h_clean = h.strip().upper()
                        if h_clean == 'HP':
                            cadrans_volumes.add('HP')
                        elif h_clean == 'HC':
                            cadrans_volumes.add('HC')
                        elif h_clean == 'BASE':
                            cadrans_volumes.add('BASE')

    # Déduction segment
    segment = None
    if est_gaz and not segments_sites:
        segment = 'GAZ'
    elif segments_sites:
        unique = set(segments_sites)
        if len(unique) > 1:
            segment = 'MULTISITE'
        else:
            seg = unique.pop()
            if seg in ('C4', 'C2'):
                segment = seg
            elif seg == 'C5':
                has_hp = 'HP' in cadrans_volumes
                has_hc = 'HC' in cadrans_volumes
                has_base = 'BASE' in cadrans_volumes
                if has_hp and has_hc and has_base:
                    segment = 'MULTISITE'
                elif has_hp or has_hc:
                    segment = 'C5-HPHC'
                elif has_base:
                    segment = 'C5-BASE'

    return {'segment': segment, 'adresse_site': adresse_premier_site}


def _contrat_extraire_champs(contenu):
    """Envoie le contenu structuré à GPT-4o-mini et retourne (champs_dict, usage)."""
    import openai

    _CHAMPS_ATTENDUS = [
        'ref_client', 'ref_vente', 'societe', 'periode', 'date_debut', 'date_fin',
        'type_energie', 'pdl_pce', 'fournisseur', 'segment',
        'nom_client', 'prenom_client', 'tel_client', 'email_client',
        'volume_elec', 'volume_gaz',
        'siren', 'adresse_client', 'date_signature',
        'vendeur', 'referent', 'montant_ht', 'commission_vendeur',
        'commission_referent', 'statut_paiement', 'date_paiement_1',
        'date_paiement_2', 'lien_drive',
    ]

    client = openai.OpenAI(api_key=os.environ.get('OPENAI_API_KEY', ''))
    resp = client.chat.completions.create(
        model='gpt-4o-mini',
        messages=[
            {'role': 'system', 'content': _CONTRAT_SYSTEM_PROMPT},
            {'role': 'user',   'content': contenu},
        ],
        temperature=0,
        response_format={'type': 'json_object'},
    )

    raw = resp.choices[0].message.content.strip()
    extracted = json.loads(raw)

    # Garantir tous les champs attendus (null par défaut)
    champs = {k: None for k in _CHAMPS_ATTENDUS}
    champs.update({k: v for k, v in extracted.items() if k in champs})

    return champs, raw, resp.usage


@app.route('/ventes/extraire-contrat', methods=['POST'])
@login_required
def extraire_contrat():
    """Reçoit un PDF en multipart, extrait les champs contractuels, renvoie du JSON.
    Ne stocke jamais le PDF sur disque."""
    if 'pdf' not in request.files:
        return jsonify({'success': False, 'error': 'Aucun fichier reçu'}), 400

    f = request.files['pdf']
    if not f.filename.lower().endswith('.pdf'):
        return jsonify({'success': False, 'error': 'Extension invalide — PDF uniquement'}), 400

    pdf_bytes = f.read()
    if len(pdf_bytes) > 10 * 1024 * 1024:
        return jsonify({'success': False, 'error': 'Fichier trop volumineux (max 10 Mo)'}), 400
    if len(pdf_bytes) == 0:
        return jsonify({'success': False, 'error': 'Fichier vide'}), 400

    try:
        contenu = _contrat_extraire_contenu(pdf_bytes)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 422
    except Exception as e:
        print(f'[CONTRAT] Erreur pdfplumber: {e}')
        return jsonify({'success': False, 'error': "Ce fichier n'est pas un PDF valide."}), 500

    try:
        champs, _raw_json, usage = _contrat_extraire_champs(contenu)
    except Exception as e:
        print(f'[CONTRAT] Erreur OpenAI: {e}')
        return jsonify({'success': False, 'error': "Erreur lors de l'analyse du contrat — réessayez ou saisissez manuellement."}), 500

    # Déduire segment et adresse depuis les tableaux (pas le LLM)
    try:
        tableaux = _contrat_analyser_tableaux(pdf_bytes)
        if tableaux.get('segment'):
            champs['segment'] = tableaux['segment']
        if tableaux.get('adresse_site'):
            champs['adresse_client'] = tableaux['adresse_site']
    except Exception as e:
        print(f'[CONTRAT] Erreur analyse tableaux: {e}')

    # Nettoyer ref_vente : "MIB-" sans chiffres = null
    import re as _re_clean
    if champs.get('ref_vente') and not _re_clean.match(r'^MIB-\d{4,}', str(champs['ref_vente'])):
        champs['ref_vente'] = None

    avertissements = []
    if not champs.get('pdl_pce'):
        avertissements.append('PDL/PCE non trouvé — vérification manuelle obligatoire')
    if not champs.get('ref_vente'):
        # Vérifier si la page 1 contient une ligne "Offre du..." sans MIB
        import re as _re_mib
        _has_offre_sans_mib = False
        try:
            import pdfplumber as _mib_plumber, io as _mib_io
            with _mib_plumber.open(_mib_io.BytesIO(pdf_bytes)) as _mib_pdf:
                p1 = (_mib_pdf.pages[0].extract_text() or '')
                for line in p1.split('\n'):
                    if _re_mib.search(r'Offre du', line, _re_mib.IGNORECASE) and 'MIB' not in line:
                        _has_offre_sans_mib = True
                        break
        except Exception:
            pass
        if _has_offre_sans_mib:
            avertissements.append('Ce CPV ne contient pas de MIB (format antérieur à juillet 2026) — utilisez le collage des données génériques.')
        else:
            avertissements.append('REF_VENTE (MIB) non trouvé')

    print(f'[CONTRAT] OK — {usage.total_tokens} tokens '
          f'(prompt: {usage.prompt_tokens}, completion: {usage.completion_tokens})')

    # Debug : 6 premières lignes de la page 1
    debug_page1 = []
    try:
        import pdfplumber, io as _dbg_io
        with pdfplumber.open(_dbg_io.BytesIO(pdf_bytes)) as _dbg_pdf:
            txt = (_dbg_pdf.pages[0].extract_text() or '').strip()
            debug_page1 = [l.strip() for l in txt.split('\n') if l.strip()][:6]
    except Exception:
        pass

    return jsonify({
        'success': True,
        'champs': champs,
        'avertissements': avertissements,
        'debug_page1': debug_page1,
        'usage': {
            'prompt_tokens':     usage.prompt_tokens,
            'completion_tokens': usage.completion_tokens,
            'total_tokens':      usage.total_tokens,
        },
    })


# ===== INVITATION RECRUTEMENT (script CA) =====

@app.route('/api/recrutement/inviter-candidat', methods=['POST'])
@login_required
def inviter_candidat_script():
    """Envoie l'invitation au candidat puis crée l'événement dans le CRM."""
    d = request.get_json()
    email = d.get('email', '').strip()
    prenom = d.get('prenom', '').strip()
    nom = d.get('nom', '').strip()
    date_session = d.get('date_session', '').strip()
    heure_session = d.get('heure_session', '').strip()
    description_crm = d.get('description_crm', '')

    if not email or not date_session or not heure_session:
        return jsonify({'success': False, 'error': 'Email, date et heure requis'}), 400

    # 1. Envoyer l'email d'invitation (PRIORITAIRE)
    mail_html = f"""<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;">
<div style="background:linear-gradient(135deg,#1e1b4b,#7c3aed);padding:32px;border-radius:12px 12px 0 0;text-align:center;">
<h1 style="color:#fff;font-size:28px;font-weight:800;letter-spacing:3px;margin:0;">LILIWATT</h1>
<p style="color:rgba(255,255,255,.8);font-size:12px;margin:6px 0 0;">Invitation session de présentation</p>
</div>
<div style="background:#f5f3ff;padding:32px;border-radius:0 0 12px 12px;">
<p style="font-size:16px;color:#1e1b4b;">Bonjour <strong>{prenom or 'Monsieur/Madame'}</strong>,</p>
<p style="color:#374151;line-height:1.7;">Suite à notre échange, nous avons le plaisir de vous inviter à rejoindre notre session de présentation LILIWATT.</p>
<div style="background:#fff;border-radius:10px;padding:24px;margin:24px 0;border-left:4px solid #7c3aed;">
<table style="width:100%;font-size:14px;border-collapse:collapse;">
<tr><td style="padding:8px 0;color:#6b7280;font-weight:700;width:100px;">Date</td><td style="color:#1e1b4b;font-weight:700;">{date_session}</td></tr>
<tr><td style="padding:8px 0;color:#6b7280;font-weight:700;">Heure</td><td style="color:#1e1b4b;font-weight:700;">{heure_session}</td></tr>
</table>
</div>
<a href="https://meet.google.com/tzv-pgjc-und?authuser=0" style="display:inline-block;background:linear-gradient(135deg,#7c3aed,#d946ef);color:#fff;padding:14px 32px;border-radius:50px;text-decoration:none;font-weight:700;font-size:14px;">Rejoindre la session Google Meet</a>
<p style="color:#374151;margin-top:20px;line-height:1.7;">À très bientôt !</p>
<p style="color:#6b7280;font-size:13px;">Carole Andria<br>carole.andria@liliwatt.fr</p>
<hr style="border:1px solid #e9d5ff;margin:24px 0;">
<p style="font-size:11px;color:#9ca3af;">LILIWATT — LILISTRAT STRATÉGIE SAS — 59 rue de Ponthieu, Bureau 326 — 75008 Paris</p>
</div></div>"""

    mail_ok = False
    try:
        token = get_zoho_token()
        if not token:
            return jsonify({'success': False, 'error': 'Impossible de se connecter à Zoho Mail'}), 500
        account_id = os.environ.get('ZOHO_ACCOUNT_ID', '8439060000000002002')
        resp = requests.post(
            f'https://mail.zoho.eu/api/accounts/{account_id}/messages',
            headers={'Authorization': f'Zoho-oauthtoken {token}', 'Content-Type': 'application/json'},
            json={'fromAddress': 'contact@liliwatt.fr', 'sender': 'Carole Andria — LILIWATT <contact@liliwatt.fr>',
                  'replyTo': 'carole.andria@liliwatt.fr', 'toAddress': email,
                  'subject': f'Invitation session LILIWATT — {date_session} à {heure_session}',
                  'content': mail_html, 'mailFormat': 'html'},
            timeout=15
        )
        print(f'[RECRUTEMENT-CA] Zoho response: status={resp.status_code} body={resp.text[:500]} account_id={account_id} fromAddress=contact@liliwatt.fr')
        if resp.status_code < 300:
            mail_ok = True
            print(f'[RECRUTEMENT-CA] Email envoyé à {email}')
        else:
            print(f'[RECRUTEMENT-CA] Zoho erreur {resp.status_code}: {resp.text[:200]}')
            return jsonify({'success': False, 'error': f'Échec envoi email (Zoho {resp.status_code})'}), 500
    except Exception as e:
        print(f'[RECRUTEMENT-CA] Erreur email: {e}')
        return jsonify({'success': False, 'error': f'Erreur envoi email : {e}'}), 500

    # 2. Créer l'événement dans le CRM (secondaire — échec non bloquant)
    crm_ok = False
    crm_error = None
    crm_api_url = os.environ.get('CRM_API_URL', '')
    crm_api_token = os.environ.get('CRM_API_TOKEN', '')
    if crm_api_url and crm_api_token:
        try:
            # Construire les dates ISO
            # date_session = "20/08/2026", heure_session = "14:00"
            import re as _re_inv
            dm = _re_inv.match(r'(\d{2})/(\d{2})/(\d{4})', date_session)
            if dm:
                iso_date = f'{dm.group(3)}-{dm.group(2)}-{dm.group(1)}'
            else:
                iso_date = date_session  # fallback
            start_iso = f'{iso_date}T{heure_session}:00+02:00'
            # Durée par défaut : 1 heure
            h, m = int(heure_session.split(':')[0]), int(heure_session.split(':')[1])
            h_end = h + 1
            end_iso = f'{iso_date}T{h_end:02d}:{m:02d}:00+02:00'

            crm_resp = requests.post(
                f'{crm_api_url}/api/calendar/external',
                headers={'Authorization': f'Bearer {crm_api_token}', 'Content-Type': 'application/json'},
                json={
                    'referentEmail': 'kevin.moreau@liliwatt.fr',
                    'title': f'Visio recrutement — {prenom} {nom}'.strip(),
                    'description': description_crm or f'Candidat : {prenom} {nom}\nEmail : {email}',
                    'startTime': start_iso,
                    'endTime': end_iso,
                },
                timeout=15
            )
            crm_data = crm_resp.json()
            if crm_data.get('success'):
                crm_ok = True
                print(f'[RECRUTEMENT-CA] CRM event créé: {crm_data.get("eventId")}')
            else:
                crm_error = crm_data.get('error', 'Erreur inconnue')
                print(f'[RECRUTEMENT-CA] CRM erreur: {crm_error}')
        except Exception as e:
            crm_error = str(e)
            print(f'[RECRUTEMENT-CA] CRM exception: {e}')
    else:
        crm_error = 'CRM_API_URL ou CRM_API_TOKEN non configuré'

    return jsonify({
        'success': True,
        'mail_ok': mail_ok,
        'crm_ok': crm_ok,
        'crm_error': crm_error,
    })


# ===== RECHERCHE MEC =====

@app.route('/api/mec/chercher')
@login_required
def mec_chercher():
    """Cherche des MEC dans l'onglet SUIVI COMMISSIONS ANNONCÉE par PDL ou nom."""
    import unicodedata, re

    pdl_query = request.args.get('pdl', '').strip()
    societe_query = request.args.get('societe', '').strip()
    vendeur_query = request.args.get('vendeur', '').strip()

    if not pdl_query and not societe_query:
        return jsonify({'success': False, 'error': 'Paramètre pdl ou societe requis'}), 400

    def _norm_pdl(s):
        return re.sub(r'[^0-9]', '', str(s or ''))

    _FORMES_JURIDIQUES = {'sci', 'sarl', 'sas', 'sasu', 'eurl', 'scea', 'earl', 'snc', 'scm', 'selarl'}

    def _norm_name(s):
        s = unicodedata.normalize('NFD', str(s))
        s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        return re.sub(r'[^a-z0-9 ]', '', s.lower()).strip()

    def _strip_juridique(s):
        """Retire les formes juridiques d'un nom normalisé."""
        words = s.split()
        filtered = [w for w in words if w not in _FORMES_JURIDIQUES]
        return ' '.join(filtered).strip()

    def _names_match(a, b):
        """Compare deux noms après retrait des formes juridiques."""
        a = _strip_juridique(a)
        b = _strip_juridique(b)
        if not a or not b:
            return False
        shorter = a if len(a) <= len(b) else b
        longer = b if len(a) <= len(b) else a
        if len(shorter) >= 6:
            return shorter in longer
        # < 6 chars : correspondance de mots entiers
        shorter_words = set(shorter.split())
        longer_words = set(longer.split())
        return bool(shorter_words & longer_words) and shorter_words <= longer_words

    gc = get_sheets_client()
    if not gc:
        return jsonify({'success': False, 'error': 'Google Sheets non configuré'}), 500

    ws = gc.open_by_key(SUIVI_VENTES_SHEET_ID).worksheet('SUIVI COMMISSIONS ANNONCÉE')
    rows = ws.get_all_values()

    def g(row, i):
        return row[i].strip() if len(row) > i else ''

    # PDL de la requête
    query_pdls = set()
    if pdl_query:
        for part in pdl_query.replace('/', ' ').split():
            n = _norm_pdl(part)
            if len(n) >= 10:
                query_pdls.add(n)

    query_societe_norm = _norm_name(societe_query) if societe_query else ''
    query_vendeur_words = _norm_name(vendeur_query.split('@')[0].replace('.', ' ')).split() if vendeur_query else []

    resultats = []
    for row_idx, row in enumerate(rows[1:], start=2):
        vendeur = g(row, 0)
        client = g(row, 1)
        comm_v = g(row, 2)
        comm_r = g(row, 3)
        horodatage = g(row, 4)
        fournisseur = g(row, 5)
        type_val = g(row, 6)
        pdl_mec = g(row, 7)
        siren_mec = g(row, 8)
        statut = g(row, 10) if len(row) > 10 else ''

        fiabilite = None

        # 1. Correspondance PDL (HAUTE)
        if query_pdls and pdl_mec:
            mec_pdls = set()
            for part in pdl_mec.replace('/', ' ').split():
                n = _norm_pdl(part)
                if len(n) >= 10:
                    mec_pdls.add(n)
            if query_pdls & mec_pdls:
                fiabilite = 'HAUTE'

        # 2. Correspondance nom (BASSE) — formes juridiques retirées, seuil 6 chars
        if not fiabilite and query_societe_norm and client:
            client_norm = _norm_name(client)
            if _names_match(query_societe_norm, client_norm):
                # Filtre vendeur si précisé
                if query_vendeur_words:
                    vendeur_norm = _norm_name(vendeur)
                    if not all(w in vendeur_norm for w in query_vendeur_words):
                        continue
                fiabilite = 'BASSE'

        if fiabilite:
            resultats.append({
                'row_idx': row_idx,
                'vendeur': vendeur,
                'client': client,
                'comm_vendeur': comm_v,
                'comm_referent': comm_r,
                'horodatage': horodatage,
                'fournisseur': fournisseur,
                'type': type_val,
                'pdl': pdl_mec,
                'siren': siren_mec,
                'statut': statut,
                'fiabilite': fiabilite,
            })

    # Tri : vendeur correspondant en tête, autres marqués
    if query_vendeur_words:
        for r in resultats:
            v_norm = _norm_name(r['vendeur'])
            r['vendeur_match'] = all(w in v_norm for w in query_vendeur_words)
        resultats.sort(key=lambda r: (0 if r.get('vendeur_match') else 1))

    return jsonify({'success': True, 'resultats': resultats, 'nb': len(resultats)})


# ===== RATTACHEMENT MEC =====

@app.route('/api/mec/rattacher', methods=['POST'])
@login_required
def mec_rattacher():
    """Écrit REF_VENTE et Signé sur les lignes MEC cochées."""
    d = request.get_json()
    rows = d.get('rows', [])
    ref_vente = d.get('ref_vente', '')
    if not rows or not ref_vente:
        return jsonify({'success': False, 'error': 'Paramètres manquants'}), 400

    try:
        gc = get_sheets_client()
        if not gc:
            return jsonify({'success': False, 'error': 'Google Sheets non configuré'}), 500
        ws = gc.open_by_key(SUIVI_VENTES_SHEET_ID).worksheet('SUIVI COMMISSIONS ANNONCÉE')

        cells = []
        for row_idx in rows:
            cells.append({'range': f'J{row_idx}', 'values': [[ref_vente]]})
            cells.append({'range': f'K{row_idx}', 'values': [['Signé']]})

        ws.batch_update(cells, value_input_option='RAW')
        return jsonify({'success': True, 'nb': len(rows)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ===== REPRISE — RAPPROCHEMENT =====

@app.route('/api/reprise/rapprocher', methods=['POST'])
@login_required
def reprise_rapprocher():
    """Extrait un CPV PDF et le rapproche d'une ligne du Sheet. Lecture seule."""
    import unicodedata, re

    if 'pdf' not in request.files:
        return jsonify({'success': False, 'error': 'Aucun fichier reçu'}), 400
    f = request.files['pdf']
    if not f.filename.lower().endswith('.pdf'):
        return jsonify({'success': False, 'error': 'PDF requis'}), 400
    pdf_bytes = f.read()
    if len(pdf_bytes) > 10 * 1024 * 1024:
        return jsonify({'success': False, 'error': 'Fichier trop volumineux'}), 400

    # Extraction du CPV
    try:
        contenu = _contrat_extraire_contenu(pdf_bytes)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 422
    except Exception:
        return jsonify({'success': False, 'error': "Ce fichier n'est pas un PDF valide."}), 500

    try:
        champs, _raw, usage = _contrat_extraire_champs(contenu)
    except Exception:
        return jsonify({'success': False, 'error': "Erreur lors de l'analyse du contrat."}), 500

    try:
        tableaux = _contrat_analyser_tableaux(pdf_bytes)
        if tableaux.get('segment'):
            champs['segment'] = tableaux['segment']
        if tableaux.get('adresse_site'):
            champs['adresse_client'] = tableaux['adresse_site']
    except Exception:
        pass

    # Debug page 1
    debug_page1 = []
    try:
        import pdfplumber as _dbg_plumber, io as _dbg_io2
        with _dbg_plumber.open(_dbg_io2.BytesIO(pdf_bytes)) as _dbg_pdf:
            txt = (_dbg_pdf.pages[0].extract_text() or '').strip()
            debug_page1 = [l.strip() for l in txt.split('\n') if l.strip()][:6]
    except Exception:
        pass

    # Normalisation des PDL
    def _norm_pdl(s):
        return re.sub(r'[^0-9]', '', str(s or ''))

    def _norm_name(s):
        s = unicodedata.normalize('NFD', str(s))
        s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        return re.sub(r'[^a-z0-9 ]', '', s.lower()).strip()

    # PDL extraits du CPV
    cpv_pdls_raw = str(champs.get('pdl_pce') or '')
    cpv_pdls = set()
    for part in cpv_pdls_raw.replace('/', ' ').split():
        n = _norm_pdl(part)
        if len(n) >= 10:
            cpv_pdls.add(n)

    if not cpv_pdls:
        return jsonify({'success': True, 'champs': champs, 'match': 'absent',
                        'raison': 'Aucun PDL extrait du contrat', 'debug_page1': debug_page1})

    # Lire le Sheet
    gc = get_sheets_client()
    if not gc:
        return jsonify({'success': False, 'error': 'Google Sheets non configuré'}), 500
    ws = gc.open_by_key(SUIVI_VENTES_SHEET_ID).sheet1
    rows = ws.get_all_values()

    def g(row, i):
        return row[i].strip() if len(row) > i else ''

    # Compteur REF_VENTE pour info
    nb_ref_vente = sum(1 for row in rows[1:] if g(row, 26))

    # Rapprochement par PDL
    matches = []
    for row_idx, row in enumerate(rows[1:], start=2):
        sheet_pdls_raw = g(row, 9)  # col J = PDL_PCE
        sheet_pdls = set()
        for part in sheet_pdls_raw.replace('/', ' ').split():
            n = _norm_pdl(part)
            if len(n) >= 10:
                sheet_pdls.add(n)
        if cpv_pdls & sheet_pdls:
            matches.append({
                'row_idx': row_idx,
                'ref': g(row, 0), 'societe': g(row, 2), 'periode': g(row, 5),
                'pdl_pce': sheet_pdls_raw, 'vendeur': g(row, 3),
            })

    # Champs à comparer (fournisseur exclu — le CPV donne "OHM ENERGIE", le Sheet l'offre)
    COMPARE_MAP = {
        'ref_vente':     26, 'societe':       2,  'siren':         27,
        'adresse_client':28, 'date_debut':    6,  'date_fin':      7,
        'type_energie':  8,  'segment':       18, 'pdl_pce':       9,
        'nom_client':    19, 'prenom_client': 20,
        'tel_client':    21, 'email_client':  22, 'volume_elec':   23,
        'volume_gaz':    24, 'date_signature':34, 'periode':       5,
    }

    def _norm_type(t):
        t = str(t).lower().strip()
        if 'gaz' in t: return 'gaz'
        if 'lec' in t or 'élec' in t or 'elec' in t: return 'elec'
        return t

    def _pdl_set(raw):
        s = set()
        for part in str(raw).replace('/', ' ').split():
            n = _norm_pdl(part)
            if len(n) >= 10: s.add(n)
        return s

    def _build_comparaison(row):
        comparaison = []
        for champ, col_idx in COMPARE_MAP.items():
            val_sheet = g(row, col_idx)
            val_cpv = champs.get(champ)
            if val_cpv is None:
                val_cpv = ''
            else:
                val_cpv = str(val_cpv).strip()

            if not val_sheet and not val_cpv:
                continue
            if not val_sheet and val_cpv:
                statut = 'a_completer'
            elif val_sheet and not val_cpv:
                statut = 'identique'
            elif champ == 'type_energie':
                statut = 'identique' if _norm_type(val_sheet) == _norm_type(val_cpv) else 'divergent'
            elif champ == 'pdl_pce':
                statut = 'identique' if _pdl_set(val_sheet) == _pdl_set(val_cpv) else 'divergent'
            elif champ in ('volume_elec', 'volume_gaz'):
                vs = parse_float(val_sheet)
                vc = parse_float(val_cpv)
                statut = 'identique' if abs(vs - vc) <= 0.5 else 'divergent'
            elif val_sheet == val_cpv:
                statut = 'identique'
            else:
                statut = 'divergent'
            comparaison.append({
                'champ': champ, 'valeur_sheet': val_sheet,
                'valeur_cpv': val_cpv, 'statut': statut,
            })
        return comparaison

    if len(matches) == 1:
        row = rows[matches[0]['row_idx'] - 1]
        return jsonify({
            'success': True, 'match': 'unique', 'champs': champs,
            'ligne': matches[0],
            'comparaison': _build_comparaison(row),
            'nb_ref_vente': nb_ref_vente, 'nb_total': len(rows) - 1,
            'debug_page1': debug_page1,
        })
    elif len(matches) > 1:
        return jsonify({
            'success': True, 'match': 'ambigu', 'champs': champs,
            'lignes': matches,
            'nb_ref_vente': nb_ref_vente, 'nb_total': len(rows) - 1,
            'debug_page1': debug_page1,
        })

    # Aucune correspondance PDL -> tenter par nom de société
    cpv_societe = _norm_name(champs.get('societe') or '')
    if cpv_societe:
        for row_idx, row in enumerate(rows[1:], start=2):
            sheet_societe = _norm_name(g(row, 2))
            if cpv_societe and sheet_societe and (cpv_societe in sheet_societe or sheet_societe in cpv_societe):
                matches.append({
                    'row_idx': row_idx,
                    'ref': g(row, 0), 'societe': g(row, 2), 'periode': g(row, 5),
                    'pdl_pce': g(row, 9), 'vendeur': g(row, 3),
                })

    if len(matches) == 1:
        row = rows[matches[0]['row_idx'] - 1]
        return jsonify({
            'success': True, 'match': 'incertain', 'champs': champs,
            'raison': 'Correspondance par nom de société (pas de PDL commun)',
            'ligne': matches[0],
            'comparaison': _build_comparaison(row),
            'nb_ref_vente': nb_ref_vente, 'nb_total': len(rows) - 1,
            'debug_page1': debug_page1,
        })
    elif len(matches) > 1:
        return jsonify({
            'success': True, 'match': 'ambigu', 'champs': champs,
            'raison': 'Plusieurs sociétés correspondent par nom (pas de PDL commun)',
            'lignes': matches,
            'nb_ref_vente': nb_ref_vente, 'nb_total': len(rows) - 1,
            'debug_page1': debug_page1,
        })

    return jsonify({
        'success': True, 'match': 'absent', 'champs': champs,
        'raison': 'Aucune correspondance par PDL ni par nom de société',
        'nb_ref_vente': nb_ref_vente, 'nb_total': len(rows) - 1,
        'debug_page1': debug_page1,
    })


# ===== REPRISE — ÉCRITURE =====

# Colonnes modifiables par la reprise (jamais VENDEUR, REFERENT, MONTANT, etc.)
_REPRISE_COL_MAP = {
    'ref_vente': 26, 'siren': 27, 'adresse_client': 28, 'score': 29,
    'pay_rank': 30, 'typologie': 31, 'nbr_sites': 32, 'commercial_ohm': 33,
    'date_signature': 34, 'puissance_kva': 35, 'date_activation': 36,
    'offre': 37, 'code_naf': 38,
    'ref_client': 1, 'societe': 2, 'segment': 18,
    'date_debut': 6, 'date_fin': 7, 'type_energie': 8, 'pdl_pce': 9,
    'nom_client': 19, 'prenom_client': 20, 'tel_client': 21, 'email_client': 22,
    'volume_elec': 23, 'volume_gaz': 24, 'periode': 5,
}

@app.route('/api/reprise/ecrire', methods=['POST'])
@login_required
def reprise_ecrire():
    """Écrit les champs cochés dans une ligne existante du Sheet."""
    d = request.get_json()
    row_idx = d.get('row_idx')
    expected_ref = d.get('ref', '')
    updates = d.get('updates', {})  # {champ: valeur}

    if not row_idx or not updates:
        return jsonify({'success': False, 'error': 'Paramètres manquants'}), 400
    if not expected_ref:
        return jsonify({'success': False, 'error': 'Référence de la ligne manquante'}), 400

    # Vérifier qu'aucun champ interdit n'est modifié
    INTERDIT = {'vendeur', 'referent', 'montant_ht', 'commission_vendeur',
                'commission_referent', 'marge', 'statut_paiement',
                'date_paiement_1', 'date_paiement_2', 'lien_drive'}
    for champ in updates:
        if champ in INTERDIT:
            return jsonify({'success': False, 'error': f'Champ {champ} interdit en reprise'}), 400
        if champ not in _REPRISE_COL_MAP:
            return jsonify({'success': False, 'error': f'Champ {champ} inconnu'}), 400

    try:
        gc = get_sheets_client()
        if not gc:
            return jsonify({'success': False, 'error': 'Google Sheets non configuré'}), 500
        ws = gc.open_by_key(SUIVI_VENTES_SHEET_ID).sheet1

        # Garde-fou : vérifier que la REF en colonne A correspond
        actual_ref = ws.cell(row_idx, 1).value or ''
        if actual_ref.strip() != expected_ref.strip():
            return jsonify({'success': False,
                'error': f'La ligne a changé depuis l\'analyse (attendu {expected_ref}, trouvé {actual_ref}). Relancez le rapprochement.'}), 409

        # Contrôles : doublon REF_VENTE (unique) + cohérence REF_CLIENT (même société)
        doublons = []
        if updates.get('ref_client') or updates.get('ref_vente'):
            all_rows = ws.get_all_values()
            # Société de la ligne en cours
            current_soc = (all_rows[row_idx - 1][2] if len(all_rows[row_idx - 1]) > 2 else '').strip()
            for ri, row in enumerate(all_rows[1:], start=2):
                if ri == row_idx:
                    continue
                row_ref = row[0] if len(row) > 0 else ''
                row_soc = (row[2] if len(row) > 2 else '').strip()
                # REF_VENTE doit être unique
                if updates.get('ref_vente') and len(row) > 26 and row[26].strip() == updates['ref_vente'].strip():
                    doublons.append(f'{updates["ref_vente"]} est déjà présent sur {row_ref} ({row_soc})')
                # REF_CLIENT : même MB sur une société différente = alerte
                if updates.get('ref_client') and len(row) > 1 and row[1].strip() == updates['ref_client'].strip():
                    if row_soc and current_soc and row_soc != current_soc:
                        doublons.append(f'{updates["ref_client"]} est déjà utilisé par {row_soc} ({row_ref}) — vérifiez le contrat collé')

        # Construire les cellules à mettre à jour
        cells = []
        for champ, val in updates.items():
            col_idx = _REPRISE_COL_MAP[champ]
            # Convertir en lettre(s) de colonne
            if col_idx < 26:
                col_letter = chr(65 + col_idx)
            else:
                col_letter = 'A' + chr(65 + col_idx - 26)
            cells.append({'range': f'{col_letter}{row_idx}', 'values': [[str(val)]]})

        ws.batch_update(cells, value_input_option='RAW')

        return jsonify({'success': True, 'nb_champs': len(cells), 'doublons': doublons})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ===== VÉRIFICATEUR AAF =====

def _analyser_aaf_soho(wb, request, get_sheets_client, sheet_id, parse_float):
    """Analyse un AAF SOHO : contrôle global + détail par contrat."""
    import re

    # Déduire le mois depuis le nom du fichier ou des feuilles
    aaf_mois = None
    for ws in wb.worksheets:
        m = re.search(r'(\d{2})-(\d{4})', ws.title)
        if m:
            aaf_mois = f'{m.group(2)}-{m.group(1)}'
            break
    # Fallback : chercher dans le nom du fichier passé
    if not aaf_mois:
        fname = request.files.get('file')
        if fname and fname.filename:
            m = re.search(r'(\d{2})-(\d{4})', fname.filename)
            if m:
                aaf_mois = f'{m.group(2)}-{m.group(1)}'

    # Période de production = mois AAF - 1
    periode_prod = None
    if aaf_mois:
        ay, am = int(aaf_mois[:4]), int(aaf_mois[5:7])
        pm = am - 1; py = ay
        if pm < 1: pm = 12; py -= 1
        periode_prod = f'{py:04d}-{pm:02d}'

    # Feuille détail (première feuille avec contractnbr)
    contrats_aaf = []
    for ws in wb.worksheets:
        row1 = [str(c.value or '').strip().lower() for c in ws[1]]
        if 'contractnbr' not in row1:
            continue
        col_map = {}
        for idx, h in enumerate(row1):
            col_map[h] = idx
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            vals = list(row)
            if not any(v is not None for v in vals):
                continue
            def _g(name):
                i = col_map.get(name)
                return str(vals[i]).strip() if i is not None and i < len(vals) and vals[i] is not None else ''
            contrats_aaf.append({
                'ref_client': _g('customernbr'),
                'ref_vente': _g('contractnbr'),
                'societe': _g('companyname'),
                'productcode': _g('productcode'),
                'date_signature': _g('contractsubscriptiondate')[:10] if _g('contractsubscriptiondate') else '',
                'date_activation': _g('contractstartdate')[:10] if _g('contractstartdate') else '',
                'statut': _g('contractstatuscode'),
            })
        break  # une seule feuille suffit

    # Feuille "facture à établir"
    facture = None
    paiement_type = ''
    for ws in wb.worksheets:
        title_lower = ws.title.strip().lower()
        if 'facture' not in title_lower or 'contrats' in title_lower:
            continue
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            vals = list(row)
            first = str(vals[0] or '').strip().lower() if vals else ''
            if first and first != 'total' and 'type' not in first and 'facture' not in first:
                paiement_type = str(vals[0] or '').strip()
                try:
                    nb_contrats = int(vals[1]) if vals[1] else 0
                    ht = round(float(vals[2]), 2) if vals[2] else 0
                    tva = round(float(vals[3]), 2) if vals[3] else 0
                    ttc = round(float(vals[4]), 2) if vals[4] else 0
                except (ValueError, TypeError):
                    continue
                facture = {'type_paiement': paiement_type, 'nb_contrats': nb_contrats,
                           'ht': ht, 'tva': tva, 'ttc': ttc}
                break
        break

    # Charger le Sheet pour comparer
    gc = get_sheets_client()
    mes_soho = []
    if gc and periode_prod:
        ws_sheet = gc.open_by_key(sheet_id).sheet1
        rows = ws_sheet.get_all_values()
        def g(row, i):
            return row[i].strip() if len(row) > i else ''
        for row in rows[1:]:
            if g(row, 10).upper() == 'OHM ENERGIE SOHO':
                mes_soho.append({
                    'ref': g(row, 0), 'societe': g(row, 2), 'periode': g(row, 5),
                    'montant': parse_float(g(row, 11)), 'statut_paiement': g(row, 15),
                    'date_p2': g(row, 17), 'comm_vendeur': parse_float(g(row, 12)),
                    'comm_referent': parse_float(g(row, 13)),
                })

    # Calcul écart
    total_attendu = sum(v['montant'] for v in mes_soho)
    total_verse = facture['ht'] if facture else 0
    ecart = round(total_verse - total_attendu, 2) if facture else None

    # Détection "1ere et 2eme partie" = contrat soldé
    solde_alert = None
    if paiement_type and '1ere' in paiement_type.lower() and '2eme' in paiement_type.lower():
        # Chercher les lignes SOHO encore en 50-50
        lignes_5050 = [v for v in mes_soho if v['statut_paiement'] == '50-50']
        if lignes_5050:
            solde_alert = [f'{v["ref"]} ({v["societe"]}) est encore en 50-50 avec DATE_P2={v["date_p2"]} alors que l\'AAF verse les deux parties en une fois' for v in lignes_5050]

    return jsonify({
        'success': True,
        'format': 'soho',
        'aaf_mois': aaf_mois,
        'periode_prod': periode_prod,
        'contrats_aaf': contrats_aaf,
        'facture': facture,
        'mes_soho': mes_soho,
        'total_attendu': total_attendu,
        'total_verse': total_verse,
        'ecart': ecart,
        'solde_alert': solde_alert,
        'nb_contrats_aaf': len(contrats_aaf),
    })


@app.route('/api/aaf/analyser', methods=['POST'])
@login_required
def analyser_aaf():
    """Analyse un fichier AAF Excel et le rapproche du Sheet. Lecture seule."""
    import openpyxl, io as _io, unicodedata, re
    from collections import defaultdict

    tunnel = request.form.get('tunnel', 'middle').lower()  # 'middle' ou 'soho'

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Aucun fichier reçu'}), 400
    f = request.files['file']
    if not f.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': 'Fichier Excel requis (.xlsx)'}), 400
    file_bytes = f.read()
    if len(file_bytes) > 10 * 1024 * 1024:
        return jsonify({'success': False, 'error': 'Fichier trop volumineux (max 10 Mo)'}), 400

    def _norm(s):
        s = unicodedata.normalize('NFD', str(s))
        s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        return re.sub(r'[^a-z0-9 ]', '', s.lower()).strip()

    def _pf(v):
        if v is None: return 0.0
        try: return round(float(v), 2)
        except (ValueError, TypeError): return 0.0

    try:
        wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), data_only=True)
    except Exception:
        return jsonify({'success': False, 'error': "Ce fichier n'est pas un Excel valide."}), 422

    ws = wb.worksheets[0]
    sheet_name = ws.title.strip()

    # ── Détecter le format : classique ou SOHO ──
    row1 = [str(c.value or '').strip().lower() for c in ws[1]]
    is_soho = 'customernbr' in row1 or 'contractnbr' in row1

    if is_soho:
        return _analyser_aaf_soho(wb, request, get_sheets_client, SUIVI_VENTES_SHEET_ID, parse_float)

    # ── Format classique ──
    # Déduire le mois : 1) nom du FICHIER (prioritaire), 2) feuille (fallback)
    aaf_mois = None
    aaf_mois_source = None
    fname = f.filename or ''
    m_match = re.search(r'(\d{2})-(\d{4})', fname)
    if m_match:
        aaf_mois = f'{m_match.group(2)}-{m_match.group(1)}'
        aaf_mois_source = 'fichier'
    if not aaf_mois:
        m_match = re.search(r'(\d{2})-(\d{4})', sheet_name)
        if m_match:
            aaf_mois = f'{m_match.group(2)}-{m_match.group(1)}'
            aaf_mois_source = 'feuille'
    if not aaf_mois:
        return jsonify({'success': False, 'error': 'Impossible de déduire le mois depuis le fichier ou la feuille. Renommez le fichier au format AAF-MM-AAAA.'}), 422

    # ── Détecter les colonnes par en-tête (ligne 3) ──
    row3 = [str(c.value or '').strip() for c in ws[3]]
    def _norm_col(s):
        """Normalise un nom de colonne : tout en minuscules, sans accents, sans séparateurs."""
        s = unicodedata.normalize('NFD', str(s))
        s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        return re.sub(r'[^a-z0-9]', '', s.lower()).strip()

    COL_MAP = {
        'reference': None, 'raisonsociale': None, 'commentaire': None,
        'premierepartiecommission': None, 'secondepartiecommission': None,
        'decommission': None, 'totalgeneral': None,
    }
    unknown_cols = []
    for idx, header in enumerate(row3):
        if not header:
            continue
        hn = _norm_col(header)
        matched = False
        for key in COL_MAP:
            if hn.startswith(key[:10]):
                COL_MAP[key] = idx
                matched = True
                break
        if not matched and idx < 8:
            unknown_cols.append(header)

    col_ref = COL_MAP.get('reference')
    col_soc = COL_MAP.get('raisonsociale')
    col_com = COL_MAP.get('commentaire')
    col_p1 = COL_MAP.get('premierepartiecommission')
    col_p2 = COL_MAP.get('secondepartiecommission')
    col_dec = COL_MAP.get('decommission')
    col_tot = COL_MAP.get('totalgeneral')

    has_decomm_col = col_dec is not None
    has_part2_col = col_p2 is not None

    if col_ref is None or col_soc is None:
        return jsonify({'success': False, 'error': f"Colonnes Reference/Raison_Sociale introuvables. En-têtes trouvés : {row3[:8]}"}), 422

    # ── Détecter le bloc "Facture à établir" (colonnes I-K) ──
    facture = None
    for row in ws.iter_rows(min_row=4, max_row=12, values_only=True):
        vals = list(row)
        for i, v in enumerate(vals[7:14], start=7):
            if str(v or '').strip().upper() == 'TOTAL':
                # La ligne suivante devrait avoir les montants
                ht_idx = i + 1
                ttc_idx = i + 3 if len(vals) > i + 3 else None
                # Chercher dans cette même ligne
                ht = _pf(vals[ht_idx] if len(vals) > ht_idx else None)
                ttc = _pf(vals[ht_idx + 2] if len(vals) > ht_idx + 2 else None)
                if ht > 0:
                    facture = {'ht': ht, 'ttc': ttc}
                break
        if facture:
            break

    # ── Extraire les données (arrêt à "Total général") ──
    aaf_lignes = []
    total_gen = {'part1': 0, 'part2': 0, 'decomm': 0, 'total': 0}
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        vals = list(row)
        ref = str(vals[col_ref] if col_ref is not None and col_ref < len(vals) else '' or '').strip()
        if not ref:
            continue
        if ref == 'Total général':
            total_gen['part1'] = _pf(vals[col_p1] if col_p1 is not None and col_p1 < len(vals) else None)
            total_gen['part2'] = _pf(vals[col_p2] if col_p2 is not None and col_p2 < len(vals) else None)
            total_gen['decomm'] = _pf(vals[col_dec] if col_dec is not None and col_dec < len(vals) else None)
            total_gen['total'] = _pf(vals[col_tot] if col_tot is not None and col_tot < len(vals) else None)
            break  # STOP — ne pas lire au-delà
        p1 = _pf(vals[col_p1] if col_p1 is not None and col_p1 < len(vals) else None)
        p2 = _pf(vals[col_p2] if col_p2 is not None and col_p2 < len(vals) else None)
        dec = _pf(vals[col_dec] if col_dec is not None and col_dec < len(vals) else None)
        tot = _pf(vals[col_tot] if col_tot is not None and col_tot < len(vals) else None)
        aaf_lignes.append({
            'ref': ref,
            'societe': str(vals[col_soc] if col_soc is not None and col_soc < len(vals) else '' or '').strip(),
            'commentaire': str(vals[col_com] if col_com is not None and col_com < len(vals) else '' or '').strip(),
            'part1': p1, 'part2': p2, 'decomm': dec, 'total': tot,
            'verse': round(p1 + p2, 2),
        })

    # Total versé = part1 + part2
    total_gen['verse'] = round(total_gen['part1'] + total_gen['part2'], 2)

    # ── Classifier ──
    challenges = [l for l in aaf_lignes if 'challenge' in l['commentaire'].lower()]
    decomms = [l for l in aaf_lignes if has_decomm_col and l['decomm'] != 0]
    # Séparer P2 pures (part1=0, part2>0, pas challenge, pas renouvellement) = secondes parties de périodes antérieures
    # Un "renouvellement" est une P2 liée à une P1 du même MIB — elle doit rester dans les commissions
    refs_avec_p1 = set(l['ref'] for l in aaf_lignes if l['part1'] > 0)
    p2_anterieures = [l for l in aaf_lignes if l not in challenges and l not in decomms
                      and l['part1'] == 0 and l['part2'] > 0 and l['ref'] not in refs_avec_p1]
    commissions = [l for l in aaf_lignes if l not in challenges and l not in decomms and l not in p2_anterieures]

    # ── Regrouper commissions par MIB (somme part1+part2 pour les renouvellements) ──
    commissions_par_ref = defaultdict(lambda: {'ref': '', 'societe': '', 'commentaire': '', 'part1': 0, 'part2': 0, 'verse': 0, 'lignes': 0})
    for l in commissions:
        r = l['ref']
        commissions_par_ref[r]['ref'] = r
        commissions_par_ref[r]['societe'] = l['societe']
        if l['commentaire'] and not commissions_par_ref[r]['commentaire']:
            commissions_par_ref[r]['commentaire'] = l['commentaire']
        commissions_par_ref[r]['part1'] = round(commissions_par_ref[r]['part1'] + l['part1'], 2)
        commissions_par_ref[r]['part2'] = round(commissions_par_ref[r]['part2'] + l['part2'], 2)
        commissions_par_ref[r]['verse'] = round(commissions_par_ref[r]['verse'] + l['verse'], 2)
        commissions_par_ref[r]['lignes'] += 1
    # Marquer les MIB soldés (P1 et P2 dans la même AAF)
    for r in commissions_par_ref.values():
        r['solde'] = r['part1'] > 0 and r['part2'] > 0
    commissions_grouped = list(commissions_par_ref.values())

    # Période de production = AAF - 1 mois
    if aaf_mois:
        ay, am = int(aaf_mois[:4]), int(aaf_mois[5:7])
        pm = am - 1; py = ay
        if pm < 1:
            pm = 12; py -= 1
        periode_prod = f'{py:04d}-{pm:02d}'
    else:
        periode_prod = None

    # Regrouper commissions_grouped par ref et par société normalisée
    aaf_par_ref = {}
    aaf_par_societe = defaultdict(lambda: {'refs': [], 'societe': '', 'total_verse': 0.0, 'lignes': []})
    for l in commissions_grouped:
        aaf_par_ref[l['ref']] = l
        sn = _norm(l['societe'])
        if l['ref'] not in aaf_par_societe[sn]['refs']:
            aaf_par_societe[sn]['refs'].append(l['ref'])
        aaf_par_societe[sn]['societe'] = l['societe']
        aaf_par_societe[sn]['total_verse'] = round(aaf_par_societe[sn]['total_verse'] + l['verse'], 2)
        aaf_par_societe[sn]['lignes'].append(l)

    # Index des montants versés pour rapprochement par montant
    aaf_montants = defaultdict(list)
    for sn, grp in aaf_par_societe.items():
        aaf_montants[grp['total_verse']].append(sn)

    # ── Charger le Sheet ──
    gc = get_sheets_client()
    if not gc:
        return jsonify({'success': False, 'error': 'Google Sheets non configuré'}), 500
    ws_sheet = gc.open_by_key(SUIVI_VENTES_SHEET_ID).sheet1
    rows = ws_sheet.get_all_values()

    def g(row, i):
        return row[i].strip() if len(row) > i else ''

    mes_ventes = []
    exclues_autre_tunnel = []
    if periode_prod:
        for row in rows[1:]:
            if g(row, 5) != periode_prod:
                continue
            fournisseur_row = g(row, 10).upper()
            is_soho_row = fournisseur_row == 'OHM ENERGIE SOHO'
            # Filtre tunnel
            if tunnel == 'middle' and is_soho_row:
                exclues_autre_tunnel.append({'ref': g(row, 0), 'societe': g(row, 2), 'montant': parse_float(g(row, 11)), 'fournisseur': g(row, 10)})
                continue
            if tunnel == 'soho' and not is_soho_row:
                exclues_autre_tunnel.append({'ref': g(row, 0), 'societe': g(row, 2), 'montant': parse_float(g(row, 11)), 'fournisseur': g(row, 10)})
                continue
            montant = parse_float(g(row, 11))
            statut = g(row, 15)
            attendu = round(montant / 2, 2) if statut == '50-50' else round(montant, 2)
            mes_ventes.append({
                'ref': g(row, 0), 'ref_vente': g(row, 26), 'societe': g(row, 2),
                'montant': montant, 'statut': statut, 'attendu': attendu,
                'societe_norm': _norm(g(row, 2)),
            })

    # ── Rapprochement ──
    matched_aaf_refs = set()
    matched_aaf_societes = set()
    rapproches = []
    manquants = []
    nb_haute = 0

    for v in mes_ventes:
        match = None
        fiabilite = None
        aaf_montant = 0

        # Niveau 1 : REF_VENTE exacte
        if v['ref_vente'] and v['ref_vente'] in aaf_par_ref:
            entry = aaf_par_ref[v['ref_vente']]
            fiabilite = 'HAUTE'
            aaf_montant = entry['verse']
            matched_aaf_refs.add(v['ref_vente'])
            nb_haute += 1
            match = entry
        # Niveau 2 : société normalisée
        elif v['societe_norm'] in aaf_par_societe and v['societe_norm'] not in matched_aaf_societes:
            grp = aaf_par_societe[v['societe_norm']]
            fiabilite = 'BASSE'
            aaf_montant = grp['total_verse']
            for r in grp['refs']:
                matched_aaf_refs.add(r)
            matched_aaf_societes.add(v['societe_norm'])
            match = grp

        if match:
            # Si le MIB est soldé (P1+P2 dans la même AAF), comparer au montant total
            is_solde = match.get('solde', False) if isinstance(match, dict) else False
            attendu_compare = v['montant'] if is_solde else v['attendu']
            ecart = round(aaf_montant - attendu_compare, 2)
            entry_type = 'SOLDE' if is_solde else ('ECART' if abs(ecart) > 0.02 else 'OK')
            rapproches.append({
                'type': entry_type,
                'ref': v['ref'], 'ref_vente': v['ref_vente'], 'societe': v['societe'],
                'attendu': attendu_compare, 'aaf': aaf_montant, 'ecart': ecart,
                'fiabilite': fiabilite, 'statut': v['statut'], 'solde': is_solde,
            })
        else:
            manquants.append({
                'ref': v['ref'], 'ref_vente': v['ref_vente'], 'societe': v['societe'],
                'attendu': v['attendu'], 'montant': v['montant'], 'statut': v['statut'],
                'societe_norm': v['societe_norm'],
            })

    # Inconnus
    inconnus = []
    for l in commissions:
        if l['ref'] not in matched_aaf_refs:
            sn = _norm(l['societe'])
            if sn not in matched_aaf_societes:
                inconnus.append(l)

    # ── "À vérifier" : correspondance partielle OU par montant exact ──
    a_verifier = []
    manquants_restants = []
    inconnus_restants = list(inconnus)

    for mq in manquants:
        found_match = False
        # 1) Correspondance partielle par nom
        for inc in inconnus_restants:
            inc_norm = _norm(inc['societe'])
            if mq['societe_norm'] in inc_norm or inc_norm in mq['societe_norm']:
                inc_group = [x for x in inconnus_restants if _norm(x['societe']) == inc_norm]
                inc_total = round(sum(x['verse'] for x in inc_group), 2)
                a_verifier.append({
                    'sheet_ref': mq['ref'], 'sheet_societe': mq['societe'], 'sheet_attendu': mq['attendu'],
                    'aaf_societe': inc['societe'], 'aaf_total': inc_total,
                    'aaf_refs': [x['ref'] for x in inc_group],
                    'ecart': round(inc_total - mq['attendu'], 2),
                    'raison': 'nom partiel',
                })
                for x in inc_group:
                    if x in inconnus_restants:
                        inconnus_restants.remove(x)
                found_match = True
                break
        # 2) Correspondance par montant exact
        if not found_match:
            for inc in inconnus_restants:
                inc_norm = _norm(inc['societe'])
                inc_group = [x for x in inconnus_restants if _norm(x['societe']) == inc_norm]
                inc_total = round(sum(x['verse'] for x in inc_group), 2)
                if abs(inc_total - mq['attendu']) <= 0.02:
                    a_verifier.append({
                        'sheet_ref': mq['ref'], 'sheet_societe': mq['societe'], 'sheet_attendu': mq['attendu'],
                        'aaf_societe': inc['societe'], 'aaf_total': inc_total,
                        'aaf_refs': [x['ref'] for x in inc_group],
                        'ecart': round(inc_total - mq['attendu'], 2),
                        'raison': 'montant identique',
                    })
                    for x in inc_group:
                        if x in inconnus_restants:
                            inconnus_restants.remove(x)
                    found_match = True
                    break
        if not found_match:
            manquants_restants.append(mq)

    ecarts = [r for r in rapproches if r['type'] == 'ECART']
    oks = [r for r in rapproches if r['type'] == 'OK']
    soldes = [r for r in rapproches if r['type'] == 'SOLDE']

    # Bandeau synthèse : production du mois
    total_commission = round(sum(v['montant'] for v in mes_ventes), 2)
    total_du = round(sum(v['attendu'] for v in mes_ventes), 2)
    total_verse_commissions = round(sum(l['verse'] for l in commissions_grouped), 2)

    return jsonify({
        'success': True,
        'aaf_mois': aaf_mois,
        'aaf_mois_source': aaf_mois_source,
        'periode_prod': periode_prod,
        'tunnel': tunnel,
        'nb_lignes_aaf': len(aaf_lignes),
        'nb_haute': nb_haute,
        'nb_total_rapproches': len(rapproches),
        'total_gen': total_gen,
        'synthese': {
            'nb_contrats': len(mes_ventes),
            'commission_totale': total_commission,
            'du': total_du,
            'verse': total_verse_commissions,
            'manque': round(total_du - total_verse_commissions, 2),
            'exclues_autre_tunnel': exclues_autre_tunnel,
        },
        'facture': facture,
        'has_part2': has_part2_col,
        'has_decomm': has_decomm_col,
        'unknown_cols': unknown_cols,
        'manquants': manquants_restants,
        'a_verifier': a_verifier,
        'ecarts': ecarts,
        'oks': oks,
        'decomms': decomms,
        'soldes': soldes,
        'challenges': challenges,
        'challenge_total': round(sum(l['verse'] for l in challenges), 2),
        'p2_anterieures': p2_anterieures,
        'inconnus': inconnus_restants,
        'mes_ventes_count': len(mes_ventes),
    })


# ===== ÉCHÉANCIER =====

@app.route('/api/echeancier')
@login_required
def api_echeancier():
    """Calcule l'échéancier encaissements / décaissements à la volée.

    Lecture seule — une seule lecture du Sheet par appel.
    """
    try:
        gc = get_sheets_client()
        if not gc:
            return jsonify({'success': False, 'error': 'Credentials Google non configurées'}), 500
        ws = gc.open_by_key(SUIVI_VENTES_SHEET_ID).sheet1
        rows = ws.get_all_values()

        def g(row, i):
            return row[i].strip() if len(row) > i else ''

        def _parse_montant(val, ref, societe, champ):
            """Parse un montant avec virgule/point/espaces. Remonte une anomalie si échec."""
            if not val:
                return 0.0, None
            cleaned = str(val).replace('\u202f', '').replace(' ', '').replace('€', '').replace(',', '.')
            try:
                return round(float(cleaned), 2), None
            except ValueError:
                return 0.0, {
                    'ref': ref, 'societe': societe,
                    'message': f'{champ} non numérique : « {val} »'
                }

        def _add_months(ym, n):
            if not ym or len(ym) < 7:
                return ''
            try:
                y, m = int(ym[:4]), int(ym[5:7])
            except ValueError:
                return ''
            m += n
            while m > 12:
                m -= 12; y += 1
            while m < 1:
                m += 12; y -= 1
            return f'{y:04d}-{m:02d}'

        # {mois_AAAA-MM: {enc: float, dec: float, details_enc: [], details_dec: []}}
        mois_data = {}
        anomalies = []
        annees_set = set()

        def _ensure_mois(ym):
            if ym and ym not in mois_data:
                mois_data[ym] = {'enc': 0.0, 'dec': 0.0, 'details_enc': [], 'details_dec': []}

        for row in rows[1:]:
            if len(row) < 18:
                continue
            ref     = g(row, 0)
            societe = g(row, 2)
            periode = g(row, 5)
            debut   = g(row, 6)
            statut  = g(row, 15)
            date_p1 = g(row, 16)
            date_p2 = g(row, 17)

            montant, err = _parse_montant(g(row, 11), ref, societe, 'MONTANT')
            if err:
                anomalies.append(err)
            comm_v, err = _parse_montant(g(row, 12), ref, societe, 'COMM_VENDEUR')
            if err:
                anomalies.append(err)
            comm_r, err = _parse_montant(g(row, 13), ref, societe, 'COMM_REFERENT')
            if err:
                anomalies.append(err)

            # --- Anomalies métier ---
            if montant > 0 and not date_p1:
                anomalies.append({
                    'ref': ref, 'societe': societe,
                    'message': 'DATE_P1 vide alors que MONTANT est renseigné'
                })
            if statut == '50-50' and not date_p2:
                anomalies.append({
                    'ref': ref, 'societe': societe,
                    'message': 'Statut 50-50 mais DATE_P2 vide — la moitié de la commission est invisible'
                })
            if date_p1 and date_p2 and date_p2 < date_p1:
                anomalies.append({
                    'ref': ref, 'societe': societe,
                    'message': f'DATE_P2 ({date_p2}) antérieure à DATE_P1 ({date_p1})'
                })
            if societe and ref and not g(row, 11):
                anomalies.append({
                    'ref': ref, 'societe': societe,
                    'message': 'MONTANT vide alors que le contrat est signé'
                })

            # --- Encaissements ---
            if montant > 0 and date_p1:
                if statut == '50-50':
                    moitie1 = round(montant / 2, 2)
                    moitie2 = round(montant - moitie1, 2)
                    _ensure_mois(date_p1)
                    mois_data[date_p1]['enc'] += moitie1
                    mois_data[date_p1]['details_enc'].append({
                        'ref': ref, 'societe': societe, 'montant': moitie1,
                        'part': '1re moitié (50-50)'
                    })
                    if date_p2:
                        _ensure_mois(date_p2)
                        mois_data[date_p2]['enc'] += moitie2
                        mois_data[date_p2]['details_enc'].append({
                            'ref': ref, 'societe': societe, 'montant': moitie2,
                            'part': '2e moitié (50-50)'
                        })
                else:
                    _ensure_mois(date_p1)
                    mois_data[date_p1]['enc'] += montant
                    mois_data[date_p1]['details_enc'].append({
                        'ref': ref, 'societe': societe, 'montant': montant,
                        'part': '100%'
                    })

            # --- Décaissements : COMM_VENDEUR + COMM_REFERENT sur PERIODE + 1 mois ---
            total_comm = round(comm_v + comm_r, 2)
            if total_comm > 0 and periode:
                mois_dec = _add_months(periode, 1)
                if mois_dec:
                    _ensure_mois(mois_dec)
                    mois_data[mois_dec]['dec'] += total_comm
                    mois_data[mois_dec]['details_dec'].append({
                        'ref': ref, 'societe': societe, 'montant': total_comm,
                        'part': f'Comm. vendeur {comm_v:.2f} + référent {comm_r:.2f}'
                    })

        # Collecter les années
        for ym in mois_data:
            if len(ym) >= 4:
                annees_set.add(ym[:4])

        # Arrondir les totaux
        for ym in mois_data:
            mois_data[ym]['enc'] = round(mois_data[ym]['enc'], 2)
            mois_data[ym]['dec'] = round(mois_data[ym]['dec'], 2)

        return jsonify({
            'success': True,
            'mois': mois_data,
            'anomalies': anomalies,
            'annees': sorted(annees_set),
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ===== SUIVI DES VENTES =====

def get_suivi_sheet_id():
    return SUIVI_VENTES_SHEET_ID

SUIVI_HEADERS = ['REF','REF_CLIENT','SOCIETE','VENDEUR','REFERENT','PERIODE','DEBUT',
    'FIN','TYPE','PDL_PCE','FOURNISSEUR','MONTANT','COMM_VENDEUR',
    'COMM_REFERENT','MARGE','STATUT_PAIEMENT','DATE_P1','DATE_P2','SEGMENT',
    'NOM_CLIENT','PRENOM_CLIENT','TEL_CLIENT','EMAIL_CLIENT',
    'VOLUME_ELEC_MWH','VOLUME_GAZ_MWH','LIEN_DRIVE',
    'REF_VENTE','SIREN','ADRESSE','SCORE','PAY_RANK',
    'TYPOLOGIE','NBR_SITES','COMMERCIAL_OHM','DATE_SIGNATURE',
    'PUISSANCE_KVA','DATE_ACTIVATION','OFFRE','CODE_NAF']

@app.route('/api/suivi-ventes/init-sheet')
@login_required
def init_suivi_sheet():
    try:
        gc = get_sheets_client()
        sh = gc.open_by_key(SUIVI_VENTES_SHEET_ID)
        ws = sh.sheet1
        _end_col = chr(64 + len(SUIVI_HEADERS)) if len(SUIVI_HEADERS) <= 26 else 'A' + chr(64 + len(SUIVI_HEADERS) - 26)
        ws.update(f'A1:{_end_col}1', [SUIVI_HEADERS])
        ws.format(f'A1:{_end_col}1', {
            'backgroundColor': {'red': 0.118, 'green': 0.106, 'blue': 0.294},
            'textFormat': {'foregroundColor': {'red': 1, 'green': 1, 'blue': 1}, 'bold': True, 'fontSize': 10}
        })
        ws.freeze(rows=1)
        sheet_url = f"https://docs.google.com/spreadsheets/d/{SUIVI_VENTES_SHEET_ID}"
        print(f"✅ Sheet Suivi Ventes initialisé: {SUIVI_VENTES_SHEET_ID}")
        return jsonify({'success': True, 'sheet_id': SUIVI_VENTES_SHEET_ID, 'sheet_url': sheet_url})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/suivi-ventes/ajouter', methods=['POST'])
@login_required
def ajouter_vente():
    try:
        d = request.get_json()
        sheet_id = get_suivi_sheet_id()
        if not sheet_id:
            return jsonify({'success': False, 'error': 'Sheet non initialisé'})

        gc = get_sheets_client()
        sh = gc.open_by_key(sheet_id)
        ws = sh.sheet1

        # Générer réf auto
        now = datetime.now()
        rows = ws.get_all_values()
        count = len(rows)
        ref = f"LW-{now.strftime('%Y%m')}-{count:03d}"

        montant = parse_float(d.get('montant_ht'))
        comm_v = parse_float(d.get('commission_vendeur'))
        comm_r = parse_float(d.get('commission_referent'))
        vol_elec = parse_float(d.get('volume_elec'))
        vol_gaz = parse_float(d.get('volume_gaz'))
        marge = montant - comm_v - comm_r

        # Normaliser le type
        type_val = (d.get('type_energie', '') or '').strip()
        if 'gaz' in type_val.lower():
            type_val = 'Gaz'
        elif 'lec' in type_val.lower():
            type_val = 'Électricité'
        print(f"📝 Vente type: {type_val} (reçu: {d.get('type_energie','')})")

        row_data = [
            ref, d.get('ref_client', ''), d.get('societe', ''), d.get('vendeur', ''), d.get('referent', ''),
            d.get('periode_prod', ''), d.get('date_debut_contrat', ''), d.get('date_fin_contrat', ''),
            type_val, d.get('pdl_pce', ''), d.get('fournisseur', ''),
            montant, comm_v, comm_r, marge,
            d.get('statut_paiement', ''), d.get('date_paiement_1', ''), d.get('date_paiement_2', ''),
            d.get('segment', ''), d.get('nom_client', ''), d.get('prenom_client', ''),
            d.get('tel_client', ''), d.get('email_client', ''),
            vol_elec, vol_gaz, d.get('lien_drive', ''),
            d.get('ref_vente', ''), d.get('siren', ''), d.get('adresse', ''),
            d.get('score', ''), d.get('pay_rank', ''), d.get('typologie', ''),
            d.get('nbr_sites', ''), d.get('commercial_ohm', ''), d.get('date_signature', ''),
            d.get('puissance_kva', ''), d.get('date_activation', ''), d.get('offre', ''), d.get('code_naf', ''),
        ]

        import time
        for attempt in range(3):
            try:
                ws.append_row(row_data, value_input_option='RAW')
                break
            except Exception:
                if attempt < 2: time.sleep(2)
                else: raise

        print(f"✅ Vente ajoutée: {ref} — {d.get('nom_client','')} — {montant}€")
        return jsonify({'success': True, 'ref': ref})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/suivi-ventes/liste')
@login_required
def liste_ventes():
    try:
        sheet_id = get_suivi_sheet_id()
        if not sheet_id:
            return jsonify({'success': False, 'error': 'Sheet non initialisé'})
        print(f"📊 SUIVI SHEET ID: {sheet_id}")
        gc = get_sheets_client()
        sh = gc.open_by_key(sheet_id)
        ws = sh.sheet1
        print(f"📊 Feuille: {ws.title} | ID feuille: {ws.id}")
        rows = ws.get_all_values()
        print(f"📊 Nb lignes: {len(rows)}")
        if rows:
            print(f"📊 Ligne 1 (en-têtes): {rows[0][:10]}")
        if len(rows) > 1:
            print(f"📊 Ligne 2 (1ère donnée): {rows[1][:10]}")
        if len(rows) < 2:
            return jsonify({'success': True, 'ventes': [], 'totaux': {'comm_vendeur': 0, 'comm_referent': 0, 'marge': 0, 'montant': 0, 'nb': 0}, 'anomalies_marge': []})

        vendeur_filter = request.args.get('vendeur', '')
        fournisseur_filter = request.args.get('fournisseur', '')
        annee_filter = request.args.get('annee', '')
        mois_filter = request.args.get('mois', '')
        search = request.args.get('search', '').lower()

        def g(row, i): return row[i] if len(row) > i else ''

        ventes = []
        total_cv, total_cr, total_m, total_montant = 0, 0, 0, 0
        anomalies_marge = []
        for row in rows[1:]:
            if len(row) < 14: continue
            if vendeur_filter and g(row,3) != vendeur_filter: continue
            if fournisseur_filter and g(row,10) != fournisseur_filter: continue
            periode = g(row,5)
            if annee_filter and not periode.startswith(annee_filter): continue
            if mois_filter and len(periode) >= 7 and periode[5:7] != mois_filter: continue
            if search:
                haystack = ' '.join([g(row,2),g(row,0),g(row,1),g(row,26),g(row,9),g(row,27),g(row,19),g(row,22)]).lower()
                # Normaliser accents pour la recherche
                import unicodedata as _ud
                haystack = _ud.normalize('NFD', haystack)
                haystack = ''.join(c for c in haystack if _ud.category(c) != 'Mn')
                search_norm = _ud.normalize('NFD', search)
                search_norm = ''.join(c for c in search_norm if _ud.category(c) != 'Mn')
                # Tous les mots doivent correspondre (ET)
                if not all(w in haystack for w in search_norm.split()):
                    continue
            montant = parse_float(g(row,11))
            cv = parse_float(g(row,12)); cr = parse_float(g(row,13))
            marge_stockee = parse_float(g(row,14))
            m = round(montant - cv - cr, 2)  # marge recalculée à la volée
            total_cv += cv; total_cr += cr; total_m += m; total_montant += montant
            # Détecter écart marge stockée
            if abs(marge_stockee - m) > 0.01 and montant > 0:
                anomalies_marge.append({'ref': g(row,0), 'societe': g(row,2), 'stockee': marge_stockee, 'calculee': m})
            ventes.append({
                'ref': g(row,0), 'ref_client': g(row,1), 'societe': g(row,2),
                'vendeur': g(row,3), 'referent': g(row,4),
                'periode_prod': g(row,5), 'date_debut': g(row,6), 'date_fin': g(row,7),
                'type': 'Gaz' if 'gaz' in g(row,18).lower() else g(row,8),
                'pdl_pce': g(row,9), 'fournisseur': g(row,10),
                'montant_ht': g(row,11), 'comm_vendeur': cv, 'comm_referent': cr, 'marge': m,
                'statut_paiement': g(row,15), 'date_p1': g(row,16), 'date_p2': g(row,17),
                'segment': g(row,18), 'nom_client': g(row,19), 'prenom_client': g(row,20),
                'tel_client': g(row,21), 'email_client': g(row,22),
                'volume_elec': g(row,23), 'volume_gaz': g(row,24), 'lien_drive': g(row,25),
                'ref_vente': g(row,26), 'siren': g(row,27), 'adresse': g(row,28),
                'score': g(row,29), 'pay_rank': g(row,30), 'typologie': g(row,31),
                'nbr_sites': g(row,32), 'commercial_ohm': g(row,33), 'date_signature': g(row,34),
                'puissance_kva': g(row,35), 'date_activation': g(row,36), 'offre': g(row,37), 'code_naf': g(row,38)
            })
        return jsonify({'success': True, 'ventes': ventes, 'totaux': {'comm_vendeur': total_cv, 'comm_referent': total_cr, 'marge': total_m, 'montant': total_montant, 'nb': len(ventes)}, 'anomalies_marge': anomalies_marge})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/suivi-ventes/export-vendeur')
@login_required
def export_vendeur():
    try:
        sheet_id = get_suivi_sheet_id()
        email = request.args.get('vendeur', '')
        periode = request.args.get('periode', '')
        annee = request.args.get('annee', '')
        if not sheet_id:
            return jsonify({'success': False, 'error': 'Sheet non initialisé'})
        gc = get_sheets_client()
        ws = gc.open_by_key(sheet_id).sheet1
        rows = ws.get_all_values()

        def g(row, i):
            return row[i] if len(row) > i else ''

        csv_lines = ['ROLE,REF,SOCIETE,VENDEUR,PERIODE,COMMISSION']
        total = 0.0

        for row in rows[1:]:
            if len(row) < 14:
                continue
            row_periode = g(row, 5)
            if periode and row_periode != periode:
                continue
            if annee and not row_periode.startswith(annee):
                continue

            ref = g(row, 0)
            societe = g(row, 2)
            vendeur = g(row, 3)
            referent = g(row, 4)
            comm_v = parse_float(g(row, 12))
            comm_r = parse_float(g(row, 13))

            # Ligne vendeur
            if email:
                if vendeur == email and comm_v > 0:
                    csv_lines.append(f'"Vendeur","{ref}","{societe}","{vendeur}","{row_periode}","{comm_v:.2f}"')
                    total += comm_v
                if referent == email and vendeur != email and comm_r > 0:
                    csv_lines.append(f'"Referent","{ref}","{societe}","{vendeur}","{row_periode}","{comm_r:.2f}"')
                    total += comm_r
            else:
                # Export global : une ligne par commission non nulle
                if comm_v > 0:
                    csv_lines.append(f'"Vendeur","{ref}","{societe}","{vendeur}","{row_periode}","{comm_v:.2f}"')
                    total += comm_v
                if comm_r > 0 and referent:
                    csv_lines.append(f'"Referent","{ref}","{societe}","{referent}","{row_periode}","{comm_r:.2f}"')
                    total += comm_r

        csv_lines.append(f'"TOTAL","","","","","{total:.2f}"')

        from flask import Response
        label = email.split('@')[0] if email else 'tous'
        csv_content = '\n'.join(csv_lines)
        return Response(csv_content, mimetype='text/csv',
            headers={'Content-Disposition': f'attachment;filename=export_{label}_{annee or periode or "all"}.csv'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/users/<email>', methods=['DELETE'])
@login_required
def delete_user(email):
    try:
        token = get_zoho_token()
        r = requests.delete(
            f'https://mail.zoho.eu/api/organization/{ZOHO_ORG_ID}/accounts',
            headers={'Authorization': f'Zoho-oauthtoken {token}', 'Content-Type': 'application/json'},
            json={'emailList': [email]}
        )
        return jsonify({'success': True, 'result': r.json()})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# ===== PROSPECTS MASTER — pagination serveur =====
import csv as csv_module
import math

MASTER_CSV = os.environ.get('MASTER_CSV_PATH', os.path.join(os.path.dirname(__file__), 'master_prospects_enrichi.csv'))

def iter_master_csv(filtres):
    """Générateur : lit le CSV ligne par ligne sans tout charger."""
    if not os.path.exists(MASTER_CSV):
        return
    with open(MASTER_CSV, encoding='utf-8', errors='ignore') as f:
        reader = csv_module.DictReader(f)
        for row in reader:
            # Filtres
            if filtres.get('segment'):
                seg = (row.get('typologie_contrat', '') or row.get('segments', '')).upper()
                if filtres['segment'].upper() not in seg:
                    continue
            if filtres.get('statut'):
                if (row.get('statut', '') or '') != filtres['statut']:
                    continue
            if filtres.get('annee_fin'):
                fin = row.get('date_fin_livraison', '') or ''
                if filtres['annee_fin'] not in fin:
                    continue
            if filtres.get('non_attribues'):
                if (row.get('vendeur_attribue', '') or '').strip():
                    continue
            if filtres.get('search'):
                hay = ' '.join([row.get('raison_sociale',''), row.get('siren',''), row.get('adresse',''), row.get('signataire',''), row.get('dirigeant','')]).lower()
                if filtres['search'].lower() not in hay:
                    continue
            yield row

@app.route('/api/prospects/master')
@login_required
def prospects_master():
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        filtres = {
            'segment': request.args.get('segment', ''),
            'statut': request.args.get('statut', ''),
            'annee_fin': request.args.get('annee_fin', ''),
            'non_attribues': request.args.get('non_attribues', '') == 'true',
            'search': request.args.get('search', ''),
        }
        skip = (page - 1) * per_page
        results = []
        total = 0
        for row in iter_master_csv(filtres):
            total += 1
            if total > skip and len(results) < per_page:
                results.append(row)
        return jsonify({
            'success': True,
            'prospects': results,
            'total': total,
            'page': page,
            'pages': math.ceil(total / per_page) if total else 0
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

# ===== PROSPECTION ADMIN =====
PROSPECTION_SHEET_ID = '1JFEAXFZbdvf40yDWZGVnuEgUN15XdOAx6WgqL69-AMA'

# ===== BASE PREMIUM =====
@app.route('/api/base-premium/stats')
@login_required
def base_premium_stats():
    try:
        gc = get_sheets_client()
        sh = gc.open_by_key(PROSPECTION_SHEET_ID)
        try:
            ws = sh.worksheet('LEADS OHM')
        except Exception:
            return jsonify({'success': True, 'stats': {'total': 0, 'attribues': 0, 'disponibles': 0, 'par_vendeur': []}})
        rows = ws.get_all_values()
        if len(rows) < 2:
            return jsonify({'success': True, 'stats': {'total': 0, 'attribues': 0, 'disponibles': 0, 'par_vendeur': []}})
        headers = rows[0]
        va_idx = next((i for i, h in enumerate(headers) if 'vendeur_attribue' in h.lower()), -1)
        total = len(rows) - 1
        vendeurs = {}
        attribues = 0
        for row in rows[1:]:
            attr = row[va_idx].strip() if va_idx >= 0 and va_idx < len(row) else ''
            if attr:
                attribues += 1
                if attr not in vendeurs:
                    vendeurs[attr] = 0
                vendeurs[attr] += 1
        # Enrichir noms
        par_vendeur = []
        try:
            gc2 = get_sheets_client()
            mdp_ws = gc2.open_by_key(os.environ.get('GOOGLE_SHEET_ID', '')).sheet1
            mdp_rows = mdp_ws.get_all_values()
            for email, nb in vendeurs.items():
                nom, prenom = email.split('@')[0], ''
                for mr in mdp_rows:
                    if len(mr) > 3 and mr[3].lower() == email.lower():
                        nom, prenom = mr[0], mr[1]
                        break
                par_vendeur.append({'email': email, 'nom': nom, 'prenom': prenom, 'nb_contacts': nb})
        except:
            par_vendeur = [{'email': e, 'nom': e.split('@')[0], 'prenom': '', 'nb_contacts': n} for e, n in vendeurs.items()]
        return jsonify({'success': True, 'stats': {'total': total, 'attribues': attribues, 'disponibles': total - attribues, 'par_vendeur': par_vendeur}})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/base-premium/liste')
@login_required
def base_premium_liste():
    try:
        gc = get_sheets_client()
        sh = gc.open_by_key(PROSPECTION_SHEET_ID)
        try:
            ws = sh.worksheet('LEADS OHM')
        except Exception:
            return jsonify({'success': True, 'prospects': [], 'total': 0, 'page': 1, 'pages': 0})
        rows = ws.get_all_values()
        if len(rows) < 2:
            return jsonify({'success': True, 'prospects': [], 'total': 0, 'page': 1, 'pages': 0})
        headers = rows[0]
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        seg_f = request.args.get('segment', '').upper()
        score_f = int(request.args.get('score_min', 0))
        statut_f = request.args.get('statut_ohm', '')
        annee_f = request.args.get('annee_fin', '')
        has_sign = request.args.get('has_signataire', '') == 'true'
        has_email = request.args.get('has_email', '') == 'true'
        non_attr = request.args.get('non_attribues', '') == 'true'

        def g(row, *names):
            for n in names:
                idx = next((i for i, h in enumerate(headers) if n.lower() in h.lower()), -1)
                if idx >= 0 and idx < len(row) and row[idx].strip():
                    return row[idx].strip()
            return ''

        filtered = []
        for i, row in enumerate(rows[1:], start=2):
            if seg_f and seg_f not in (g(row, 'segment', 'typologie') or '').upper():
                continue
            if score_f:
                try:
                    if int(g(row, 'score') or '0') < score_f: continue
                except: continue
            if statut_f and g(row, 'statut') != statut_f: continue
            if annee_f and annee_f not in (g(row, 'date_fin') or ''): continue
            if has_sign and not g(row, 'signataire'): continue
            if has_email and not g(row, 'email_signataire'): continue
            if non_attr and g(row, 'vendeur_attribue'): continue
            obj = {'_row': i}
            for j, h in enumerate(headers):
                obj[h] = row[j] if j < len(row) else ''
            filtered.append(obj)

        total = len(filtered)
        pages = math.ceil(total / per_page) if total else 0
        start = (page - 1) * per_page
        prospects = filtered[start:start + per_page]
        return jsonify({'success': True, 'prospects': prospects, 'total': total, 'page': page, 'pages': pages})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/base-premium/attribuer', methods=['POST'])
@login_required
def base_premium_attribuer():
    try:
        d = request.get_json()
        rows_list = d.get('rows', [])
        vendeur_email = d.get('vendeur_email', '')
        if not rows_list or not vendeur_email:
            return jsonify({'success': False, 'error': 'Rows et vendeur requis'})
        if len(rows_list) > 20:
            return jsonify({'success': False, 'error': 'Max 20 contacts'})
        gc = get_sheets_client()
        sh = gc.open_by_key(PROSPECTION_SHEET_ID)
        ws = sh.worksheet('LEADS OHM')
        headers = ws.row_values(1)
        col_idx = -1
        for i, h in enumerate(headers):
            if 'vendeur_attribue' in h.lower():
                col_idx = i + 1
                break
        if col_idx < 0:
            col_idx = len(headers) + 1
            ws.update_cell(1, col_idx, 'vendeur_attribue')
        import time
        for row_num in rows_list:
            ws.update_cell(row_num, col_idx, vendeur_email)
            time.sleep(0.3)
        print(f"💎 {len(rows_list)} contacts attribués à {vendeur_email}")
        return jsonify({'success': True, 'attribues': len(rows_list)})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/prospection/scraper', methods=['POST'])
@login_required
def prospection_scraper():
    try:
        import subprocess
        d = request.get_json()
        ville = d.get('ville', '').strip()
        secteurs = d.get('secteurs', [])
        if not ville or not secteurs:
            return jsonify({'success': False, 'error': 'Ville et secteurs requis'})

        script = os.path.join(os.path.dirname(__file__), '..', 'scraper_google_places.py')
        if not os.path.exists(script):
            script = '/Users/strategyglobal/Desktop/scraper_google_places.py'

        details = []
        total = 0
        for secteur in secteurs:
            try:
                result = subprocess.run(
                    ['python3', script, secteur, ville],
                    capture_output=True, text=True, timeout=120,
                    cwd=os.path.dirname(script)
                )
                output = result.stdout + result.stderr
                # Parse le résultat
                trouves = 0
                ajoutes = 0
                for line in output.split('\n'):
                    if 'trouvés' in line:
                        import re
                        m = re.search(r'(\d+) trouvés', line)
                        if m: trouves = int(m.group(1))
                    if 'nouvelles lignes' in line:
                        m = re.search(r'(\d+) nouvelles', line)
                        if m: ajoutes = int(m.group(1))
                details.append({'secteur': secteur, 'trouves': trouves, 'ajoutes': ajoutes})
                total += ajoutes
                print(f"🎯 Scraping {secteur}/{ville}: {trouves} trouvés, {ajoutes} ajoutés")
            except subprocess.TimeoutExpired:
                details.append({'secteur': secteur, 'trouves': 0, 'ajoutes': 0, 'error': 'timeout'})
            except Exception as e:
                details.append({'secteur': secteur, 'trouves': 0, 'ajoutes': 0, 'error': str(e)})

        return jsonify({'success': True, 'total_ajoutes': total, 'details': details})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/prospection/leads-ohm')
@login_required
def prospection_leads_ohm():
    try:
        gc = get_sheets_client()
        sh = gc.open_by_key(PROSPECTION_SHEET_ID)
        try:
            ws = sh.worksheet('LEADS OHM')
        except Exception:
            return jsonify({'success': True, 'prospects': []})
        rows = ws.get_all_values()
        if len(rows) < 2:
            return jsonify({'success': True, 'prospects': []})

        headers = rows[0]
        seg_f = request.args.get('segment', '')
        annee_f = request.args.get('annee_fin', '')
        statut_f = request.args.get('statut', '')
        non_attr = request.args.get('non_attribues', '') == 'true'
        per_page = int(request.args.get('per_page', 20))

        def g(row, name):
            idx = headers.index(name) if name in headers else -1
            return row[idx] if idx >= 0 and idx < len(row) else ''

        prospects = []
        for i, row in enumerate(rows[1:], start=2):
            if seg_f:
                s = g(row, 'segments') or g(row, 'typologie_contrat') or ''
                if seg_f.upper() not in s.upper(): continue
            if annee_f and annee_f not in (g(row, 'date_fin_livraison') or ''): continue
            if statut_f and (g(row, 'statut') or '') != statut_f: continue
            if non_attr and (g(row, 'vendeur_attribue') or '').strip(): continue
            obj = {'_row': i}
            for j, h in enumerate(headers):
                obj[h] = row[j] if j < len(row) else ''
            prospects.append(obj)
            if len(prospects) >= per_page: break

        return jsonify({'success': True, 'prospects': prospects})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/prospection/attribuer-leads', methods=['POST'])
@login_required
def attribuer_leads():
    try:
        d = request.get_json()
        rows_list = d.get('rows', [])
        email_vendeur = d.get('email_vendeur', '')
        if not rows_list or not email_vendeur:
            return jsonify({'success': False, 'error': 'Rows et email requis'})

        gc = get_sheets_client()
        sh = gc.open_by_key(PROSPECTION_SHEET_ID)
        try:
            ws = sh.worksheet('LEADS OHM')
        except Exception:
            return jsonify({'success': False, 'error': 'Feuille LEADS OHM non trouvée'})

        headers = ws.row_values(1)
        col_idx = -1
        for i, h in enumerate(headers):
            if 'vendeur_attribue' in h.lower():
                col_idx = i + 1
                break
        if col_idx < 0:
            col_idx = len(headers) + 1
            ws.update_cell(1, col_idx, 'vendeur_attribue')

        import time
        for row_num in rows_list:
            ws.update_cell(row_num, col_idx, email_vendeur)
            time.sleep(0.3)

        print(f"✅ {len(rows_list)} leads attribués à {email_vendeur}")
        return jsonify({'success': True, 'nb': len(rows_list)})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

def get_next_courtier_number():
    """Génère le prochain numéro de courtier auto-incrémenté"""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        creds_json = os.environ.get('GOOGLE_CREDS_JSON', '')
        if not creds_json:
            return None
        creds_dict = json.loads(creds_json)
        creds = Credentials.from_service_account_info(creds_dict, scopes=[
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ])
        gc = gspread.authorize(creds)
        sheet_id = os.environ.get('GOOGLE_SHEET_ID', '')
        sh = gc.open_by_key(sheet_id)
        ws = sh.sheet1
        all_values = ws.get_all_values()
        numbers = []
        for row in all_values[1:]:
            if len(row) > 11 and row[11].strip().isdigit():
                numbers.append(int(row[11]))
        next_num = max(numbers, default=46) + 1
        print(f"📊 Prochain numéro courtier: {next_num}")
        return next_num
    except Exception as e:
        print(f"⚠️ Erreur numéro courtier: {e}")
        return None


def update_role_in_sheets(email, new_role, sheet_id):
    """Met à jour le rôle d'un utilisateur dans un Google Sheet"""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        creds_json = os.environ.get('GOOGLE_CREDS_JSON', '')
        if not creds_json:
            return False
        creds_dict = json.loads(creds_json)
        creds = Credentials.from_service_account_info(creds_dict, scopes=[
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ])
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(sheet_id)
        ws = sh.sheet1
        all_values = ws.get_all_values()
        for i, row in enumerate(all_values):
            if len(row) > 3 and row[3].lower() == email.lower():
                # Colonne J = index 10 (0-based) = rôle
                ws.update_cell(i + 1, 10, new_role)
                print(f'✅ Sheets mis à jour: {email} → {new_role} (sheet {sheet_id[:8]}...)')
                return True
        print(f'⚠️ Email {email} non trouvé dans sheet {sheet_id[:8]}...')
        return False
    except Exception as e:
        print(f'❌ Sheets error: {e}')
        return False


@app.route('/api/promote-vendeur', methods=['POST'])
@login_required
def promote_vendeur():
    data = request.get_json()
    email = data.get('email')
    new_role = data.get('role')  # 'vendeur' ou 'referent'

    if not email or new_role not in ['vendeur', 'referent']:
        return jsonify({'error': 'Paramètres invalides'}), 400

    # 1. Google Sheets MDP ZOHO
    sheets_id = os.environ.get('GOOGLE_SHEET_ID', '')
    if sheets_id:
        update_role_in_sheets(email, new_role, sheets_id)

    # 2. CRM Neon (Vercel)
    crm_role = 'REFERENT' if new_role == 'referent' else 'VENDEUR'
    CRM_URL = os.environ.get('CRM_URL', 'https://liliwatt-crm-8ofi.vercel.app')
    CRM_API_KEY = os.environ.get('CRM_API_KEY', 'liliwatt-crm-api-key-2026')
    try:
        crm_res = requests.post(
            f'{CRM_URL}/api/crm/promote',
            headers={'X-API-Key': CRM_API_KEY, 'Content-Type': 'application/json'},
            json={'email': email, 'role': crm_role},
            timeout=10
        )
        print(f'CRM promote: {crm_res.status_code} — {crm_res.text}')
    except Exception as e:
        print(f'CRM promote error: {e}')

    # 3. Sheets Prospection
    master_sheet_id = os.environ.get('MASTER_SHEET_ID', '')
    if master_sheet_id:
        update_role_in_sheets(email, new_role, master_sheet_id)

    # 4. Création salle Meet si promu référent
    meet_link = None
    if new_role == 'referent':
        try:
            gc = get_sheets_client()
            sheet_id = os.environ.get('GOOGLE_SHEET_ID', '')
            ws = gc.open_by_key(sheet_id).sheet1
            rows = ws.get_all_values()
            prenom_ref = ''
            nom_ref = ''
            lien_actuel = ''
            row_idx = -1
            for i, row in enumerate(rows):
                if len(row) > 3 and row[3].strip().lower() == email.lower():
                    prenom_ref = row[1].strip() if len(row) > 1 else ''
                    nom_ref = row[0].strip() if len(row) > 0 else ''
                    lien_actuel = row[12].strip() if len(row) > 12 else ''
                    row_idx = i + 1
                    break
            if not lien_actuel or 'meet.google.com' not in lien_actuel:
                result = create_referent_meet_room(prenom_ref, nom_ref, email)
                meet_link = result['meet_link']
                if row_idx > 0:
                    ws.update_cell(row_idx, 13, meet_link)
                CRM_URL2 = os.environ.get('CRM_URL', 'https://liliwatt-crm-8ofi.vercel.app')
                CRM_KEY2 = os.environ.get('CRM_API_KEY', 'liliwatt-crm-api-key-2026')
                requests.post(f'{CRM_URL2}/api/crm/update-lien-visio',
                    headers={'X-API-Key': CRM_KEY2, 'Content-Type': 'application/json'},
                    json={'email': email, 'lien': meet_link}, timeout=10)
                print(f'✅ Salle Meet créée pour promotion: {meet_link}')
            else:
                print(f'⏭️ Salle Meet existe déjà: {lien_actuel}')
        except Exception as e:
            print(f'⚠️ Meet promotion error: {e}')

    # 5. Si rétrogradation referent → vendeur, orpheliner ses vendeurs
    orphelins = []
    if new_role == 'vendeur':
        orphelins = orpheliner_vendeurs(email)

    return jsonify({'success': True, 'email': email, 'role': new_role, 'meet_link': meet_link, 'orphelins': orphelins})


@app.route('/api/count-vendeurs-sous-referent')
@login_required
def count_vendeurs_sous_referent():
    email = request.args.get('email', '').strip().lower()
    if not email:
        return jsonify({'count': 0})
    try:
        gc = get_sheets_client()
        ws = gc.open_by_key(os.environ.get('GOOGLE_SHEET_ID', '')).sheet1
        rows = ws.get_all_values()
        count = sum(1 for row in rows if len(row) > 6 and row[6].strip().lower() == email)
        return jsonify({'count': count})
    except:
        return jsonify({'count': 0})


@app.route('/api/changer-referent', methods=['POST'])
@login_required
def changer_referent():
    data = request.get_json()
    vendeur_email = (data.get('vendeur_email') or '').strip().lower()
    referent_email = (data.get('referent_email') or '').strip().lower()

    if not vendeur_email:
        return jsonify({'success': False, 'error': 'Email vendeur manquant'}), 400

    results = {}

    # 1. Sheets colonne G (REFERANT) = index 7 en 1-based
    try:
        gc = get_sheets_client()
        sheet_id = os.environ.get('GOOGLE_SHEET_ID', '')
        ws = gc.open_by_key(sheet_id).sheet1
        rows = ws.get_all_values()
        found = False
        for i, row in enumerate(rows):
            if len(row) > 3 and row[3].strip().lower() == vendeur_email:
                ws.update_cell(i + 1, 7, referent_email)
                print(f'✅ Sheets: référent de {vendeur_email} → {referent_email or "aucun"}')
                results['sheets'] = 'ok'
                found = True
                break
        if not found:
            results['sheets'] = 'vendeur introuvable'
    except Exception as e:
        print(f'❌ Erreur Sheets changer-referent: {e}')
        results['sheets'] = str(e)

    # 2. CRM Neon
    try:
        crm_url = os.environ.get('CRM_URL', 'https://liliwatt-crm-8ofi.vercel.app')
        crm_key = os.environ.get('CRM_API_KEY', 'liliwatt-crm-api-key-2026')
        r = requests.post(
            f'{crm_url}/api/crm/assign-referent',
            headers={'X-API-Key': crm_key, 'Content-Type': 'application/json'},
            json={'vendeur_email': vendeur_email, 'referent_email': referent_email or None},
            timeout=15
        )
        print(f'CRM assign-referent: {r.status_code}')
        results['crm'] = 'ok' if r.ok else f'status {r.status_code}'
    except Exception as e:
        print(f'⚠️ CRM assign-referent error: {e}')
        results['crm'] = str(e)

    return jsonify({'success': True, 'results': results})


@app.route('/api/toggle-vendeur', methods=['POST'])
@login_required
def toggle_vendeur():
    data = request.get_json()
    email = data.get('email')
    actif = data.get('actif')  # True ou False

    if not email or actif is None:
        return jsonify({'error': 'Paramètres invalides'}), 400

    # 1. Google Sheets — colonne K = STATUT
    gc = get_sheets_client()
    if gc:
        try:
            sheet_id = os.environ.get('GOOGLE_SHEET_ID', '')
            sh = gc.open_by_key(sheet_id)
            ws = sh.sheet1
            all_values = ws.get_all_values()
            for i, row in enumerate(all_values):
                if len(row) > 3 and row[3].lower() == email.lower():
                    ws.update_cell(i + 1, 11, 'actif' if actif else 'inactif')
                    print(f'✅ Sheets toggle: {email} → {"actif" if actif else "inactif"}')
                    break
        except Exception as e:
            print(f'❌ Sheets toggle error: {e}')

    # 2. CRM Neon
    CRM_URL = os.environ.get('CRM_URL', 'https://liliwatt-crm-8ofi.vercel.app')
    CRM_API_KEY = os.environ.get('CRM_API_KEY', 'liliwatt-crm-api-key-2026')
    try:
        requests.post(
            f'{CRM_URL}/api/crm/toggle',
            headers={'X-API-Key': CRM_API_KEY, 'Content-Type': 'application/json'},
            json={'email': email, 'actif': actif},
            timeout=10
        )
    except Exception as e:
        print(f'CRM toggle error: {e}')

    # 3. Orpheliner si désactivation d'un référent
    orphelins = []
    if not actif:
        try:
            gc2 = get_sheets_client()
            ws2 = gc2.open_by_key(os.environ.get('GOOGLE_SHEET_ID', '')).sheet1
            for row in ws2.get_all_values():
                if len(row) > 9 and row[3].strip().lower() == email.lower() and row[9].strip().lower() == 'referent':
                    orphelins = orpheliner_vendeurs(email)
                    break
        except: pass

    return jsonify({'success': True, 'email': email, 'actif': actif, 'orphelins': orphelins})


@app.route('/api/supprimer-vendeur', methods=['POST'])
@login_required
def supprimer_vendeur():
    data = request.get_json()
    email = data.get('email')

    if not email:
        return jsonify({'error': 'Email requis'}), 400

    # 1. Google Sheets — marque SUPPRIMÉ dans colonne K
    gc = get_sheets_client()
    if gc:
        try:
            sheet_id = os.environ.get('GOOGLE_SHEET_ID', '')
            sh = gc.open_by_key(sheet_id)
            ws = sh.sheet1
            all_values = ws.get_all_values()
            for i, row in enumerate(all_values):
                if len(row) > 3 and row[3].lower() == email.lower():
                    ws.update_cell(i + 1, 11, 'supprime')
                    print(f'✅ Sheets supprimé: {email}')
                    break
        except Exception as e:
            print(f'❌ Sheets suppress error: {e}')

    # 2. CRM Neon — marque inactif + deletedAt
    CRM_URL = os.environ.get('CRM_URL', 'https://liliwatt-crm-8ofi.vercel.app')
    CRM_API_KEY = os.environ.get('CRM_API_KEY', 'liliwatt-crm-api-key-2026')
    try:
        requests.post(
            f'{CRM_URL}/api/crm/delete-user',
            headers={'X-API-Key': CRM_API_KEY, 'Content-Type': 'application/json'},
            json={'email': email},
            timeout=10
        )
    except Exception as e:
        print(f'CRM delete error: {e}')

    # 3. Orpheliner si c'est un référent
    orphelins = orpheliner_vendeurs(email)

    return jsonify({'success': True, 'email': email, 'orphelins': orphelins})


# ===== NEWSLETTER =====

def _nl_unsub_token(email):
    """HMAC token pour désinscription (32 chars)."""
    return hmac.new(NEWSLETTER_SECRET.encode(), email.strip().lower().encode(), hashlib.sha256).hexdigest()[:32]

def _nl_get_or_create_sheet(tab_name, headers):
    """Retourne un worksheet, le crée s'il n'existe pas."""
    gc = get_sheets_client()
    sh = gc.open_by_key(SUIVI_VENTES_SHEET_ID)
    try:
        ws = sh.worksheet(tab_name)
    except Exception:
        ws = sh.add_worksheet(title=tab_name, rows=500, cols=len(headers))
        ws.update('A1:' + chr(64 + len(headers)) + '1', [headers])
    return ws

def _nl_get_unsub_set():
    """Retourne le set des emails désinscrits."""
    try:
        ws = _nl_get_or_create_sheet('NEWSLETTER_UNSUB', ['EMAIL', 'DATE', 'ORIGINE'])
        rows = ws.get_all_values()
        return {r[0].strip().lower() for r in rows[1:] if r and r[0]}
    except Exception as e:
        print(f"⚠️ Erreur lecture NEWSLETTER_UNSUB: {e}")
        return set()

_NL_PARTICLES = frozenset({
    'de', 'du', 'des', 'le', 'la', 'les', 'van', 'von', 'der', 'den', 'el', 'al'
})
_NL_BLACKLIST_PRENOMS = frozenset({
    'm', 'mr', 'mme', 'dr',
    'sarl', 'sas', 'sasu', 'eurl', 'sci',
    'societe', 'entreprise', 'service', 'contact',
    'compta', 'direction', 'gerant'
})

def _nl_strip_accents(t):
    import unicodedata as _ud
    return ''.join(c for c in _ud.normalize('NFD', t) if _ud.category(c) != 'Mn').lower()

def _nl_normalise_prenom(raw):
    """Normalise un prénom brut. Retourne '' si absent ou invalide."""
    s = raw.strip()
    if not s:
        return ''
    # Longueur
    if len(s) < 2 or len(s) > 25:
        return ''
    # Caractères autorisés : lettres unicode, espace, tiret, apostrophe
    for ch in s:
        if ch.isdigit() or ch == '@':
            return ''
        if not (ch.isalpha() or ch in (' ', '-', "'")):
            return ''
    # Blacklist (token par token, sans accents)
    tokens = _nl_strip_accents(s).replace('-', ' ').replace("'", ' ').split()
    for tok in tokens:
        if tok in _NL_BLACKLIST_PRENOMS:
            return ''
    # Pas de normalisation de casse si < 3 caractères (préserve initiales "JC")
    if len(s) < 3:
        return s

    def _cap_word(w):
        """Capitalise un mot, gère les apostrophes internes et les particules."""
        if not w:
            return w
        if "'" in w:
            pre, _, suf = w.partition("'")
            pre_l = _nl_strip_accents(pre)
            if pre_l in ('d', 'l'):         # d'Alembert, l'Herminier
                return pre.lower() + "'" + (_cap_word(suf) if suf else '')
            pre_c = (pre[0].upper() + pre[1:].lower()) if len(pre) >= 3 else pre
            return pre_c + "'" + (_cap_word(suf) if suf else '')
        wl = _nl_strip_accents(w)
        if wl in _NL_PARTICLES:
            return w.lower()
        if len(w) < 3:
            return w                        # préserve initiales courtes
        return w[0].upper() + w[1:].lower()

    # Split sur tiret, puis sur espace dans chaque segment
    return '-'.join(
        ' '.join(_cap_word(sp) for sp in hp.split(' '))
        for hp in s.split('-')
    )

def _nl_parse_balises(text):
    """Parse [bouton]texte|url[/bouton] et [lien]texte|url[/lien] dans du HTML échappé."""
    # Boutons
    def repl_bouton(m):
        parts = m.group(1).split('|', 1)
        if len(parts) != 2:
            return m.group(0)
        txt, url = parts
        return (f'<table role="presentation" cellpadding="0" cellspacing="0" style="margin:24px 0;">'
                f'<tr><td>'
                f'<a href="{url.strip()}" style="display:inline-block;background:#7c3aed;color:#ffffff;text-decoration:none;font-family:Inter,Arial,sans-serif;font-weight:600;font-size:15px;padding:14px 28px;border-radius:8px;" target="_blank">{txt.strip()}</a>'
                f'</td></tr></table>')
    text = re.sub(r'\[bouton\](.*?)\[/bouton\]', repl_bouton, text, flags=re.DOTALL)
    # Supprimer les <br> immédiatement avant/après les wrappers bouton
    text = re.sub(r'(<br>)+(<table role="presentation" cellpadding="0" cellspacing="0" style="margin:24px 0;">)', r'\2', text)
    text = re.sub(r'(</table>)(<br>)+', r'\1', text)
    # Liens
    def repl_lien(m):
        parts = m.group(1).split('|', 1)
        if len(parts) != 2:
            return m.group(0)
        txt, url = parts
        return f'<a href="{url.strip()}" style="color:#7c3aed;text-decoration:underline;font-weight:600;" target="_blank">{txt.strip()}</a>'
    text = re.sub(r'\[lien\](.*?)\[/lien\]', repl_lien, text, flags=re.DOTALL)
    # Raccourcis
    if GOOGLE_AVIS_URL:
        text = text.replace('[avis]', GOOGLE_AVIS_URL)
    if PARRAINAGE_URL:
        text = text.replace('[parrainage]', PARRAINAGE_URL)
    return text

def _nl_build_html(objet, titre, body_html, unsub_url, fmt='newsletter', cta_texte='', cta_lien='', prenom=''):
    """Construit le template email newsletter complet — 9 blocs fidèles à la maquette."""
    ml = MENTIONS_LEGALES
    avis_url = GOOGLE_AVIS_URL or '#'
    parr_url = PARRAINAGE_URL or '#'
    titre_html = (f'<div style="font-size:24px;font-weight:800;color:#1e1b4b;line-height:1.2;margin:16px 0 0;">'
                  f'{html_mod.escape(titre)}</div>') if titre else ''
    # CTA centré en bas de BLOC 5 : obligatoire en Newsletter (validé côté UI), facultatif sinon
    cta_html = ''
    if cta_texte and cta_lien:
        cta_html = (
            f'<table role="presentation" cellpadding="0" cellspacing="0" align="center" style="margin-top:32px;">'
            f'<tr><td style="border:2px solid #7c3aed;border-radius:8px;">'
            f'<a href="{html_mod.escape(cta_lien)}" '
            f'style="display:inline-block;padding:14px 32px;color:#7c3aed;background:transparent;'
            f'text-decoration:none;font-weight:700;font-size:14px;border-radius:6px;" '
            f'target="_blank">{html_mod.escape(cta_texte)}</a>'
            f'</td></tr></table>'
        )

    salutation = f'Bonjour&nbsp;{html_mod.escape(prenom)},' if prenom else 'Bonjour,'

    def _social_cells(spacing=4):
        c = ''
        for name, url, img in SOCIAL_LINKS:
            c += (f'<td style="padding:0 {spacing}px;">'
                  f'<a href="{url}" title="{name.capitalize()}" target="_blank">'
                  f'<img src="{_nl_asset(img)}" alt="{name.capitalize()}" width="32" height="32" '
                  f'style="display:block;border:0;border-radius:50%;" /></a></td>')
        return c

    social_hdr = f'<table role="presentation" cellpadding="0" cellspacing="0" align="right"><tr>{_social_cells()}</tr></table>'
    social_ftr = f'<table role="presentation" cellpadding="0" cellspacing="0" align="center"><tr>{_social_cells()}</tr></table>'

    _sep = '<span style="color:#4c4a7a;margin:0 8px;">|</span>'
    _link_tpl = '<a href="{u}" style="color:#ffffff;text-decoration:none;font-size:13px;" target="_blank">{l}</a>'
    footer_nav_html = _sep.join(
        _link_tpl.format(u=u, l=html_mod.escape(l)) for l, u in FOOTER_NAV
    )

    parts = []
    parts.append(f'''<!DOCTYPE html><html lang="fr"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head><body style="margin:0;padding:0;">
<!-- BLOC 1 : Preheader -->
<div style="display:none;max-height:0;overflow:hidden;mso-hide:all;">Newsletter LILIWATT &mdash; {objet}</div>
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#f5f3ff;font-family:Inter,Arial,sans-serif;color:#241f47;">
<tr><td style="background:#ffffff;padding:8px 0;">
  <table role="presentation" width="640" cellpadding="0" cellspacing="0" align="center"><tr>
    <td style="font-size:12px;color:#6b7280;padding:0 20px;text-align:center;">Newsletter &#8211; Ao&#251;t 2026</td>
  </tr></table>
</td></tr>''')

    parts.append(f'''<!-- BLOC 2 : Bandeau navy -->
<tr><td align="center">
<table role="presentation" width="640" cellpadding="0" cellspacing="0" style="background:#1e1b4b;">
  <tr><td style="padding:18px 28px;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"><tr>
      <td style="vertical-align:middle;">
        <img src="{_nl_asset('logo_blanc.png')}" alt="LILIWATT" width="140" height="36" style="display:block;border:0;" />
      </td>
      <td align="right" style="vertical-align:middle;">{social_hdr}</td>
    </tr></table>
  </td></tr>
</table>
</td></tr>''')

    _hero_img  = 'hero_bienvenue.jpg' if fmt == 'bienvenue' else 'hero.jpg'
    _hero_alt  = 'Bienvenue chez LILIWATT' if fmt == 'bienvenue' else 'L&#39;&#233;nergie &#233;volue, nous vous accompagnons'
    parts.append(f'''<!-- BLOC 3 : Hero -->
<tr><td align="center">
<table role="presentation" width="640" cellpadding="0" cellspacing="0">
  <tr><td bgcolor="#4c1d95">
    <img src="{_nl_asset(_hero_img)}" alt="{_hero_alt}" width="640" style="display:block;width:100%;max-width:640px;height:auto;border:0;" />
  </td></tr>
</table>
</td></tr>''')

    parts.append(f'''<!-- BLOC 4 : Salutation -->
<tr><td align="center">
<table role="presentation" width="640" cellpadding="0" cellspacing="0" style="background:#ffffff;">
  <tr><td style="padding:28px 40px 12px;">
    <div style="font-size:16px;font-weight:700;color:#1e1b4b;margin-bottom:10px;">{salutation}</div>
    <div style="font-size:14px;color:#4b5563;line-height:1.65;">{html_mod.escape(NL_INTRO.get(fmt, NL_INTRO["newsletter"]))}</div>
  </td></tr>
</table>
</td></tr>''')

    parts.append(f'''<!-- BLOC 5 : Contenu editable -->
<tr><td align="center">
<table role="presentation" width="640" cellpadding="0" cellspacing="0" style="background:#ffffff;">
  <tr><td style="padding:8px 40px 12px;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#f5f3ff;border-radius:12px;">
      <tr><td style="padding:32px;">
        <table role="presentation" cellpadding="0" cellspacing="0"><tr>
          <td style="vertical-align:top;width:56px;">
            <img src="{_nl_asset('icone_actu.png' if fmt == 'newsletter' else 'icone_contact.png')}" alt="" width="56" height="56" style="display:block;border:0;border-radius:50%;" />
          </td>
          <td style="padding-left:16px;vertical-align:top;">
            <div style="font-size:13px;font-weight:800;letter-spacing:1px;color:#7c3aed;text-transform:uppercase;margin-bottom:6px;">{'ACTUALIT\u00c9 DU MOIS' if fmt == 'newsletter' else ('BIENVENUE' if fmt == 'bienvenue' else 'RESTONS EN CONTACT')}</div>
          </td>
        </tr></table>
        {titre_html}
        <div style="font-size:14px;line-height:1.65;color:#241f47;margin-top:16px;">{body_html}</div>
        {cta_html}
      </td></tr>
    </table>
  </td></tr>
</table>
</td></tr>''')

    parts.append(f'''<!-- BLOC 6 : Double colonne Avis + Parrainage -->
<tr><td align="center">
<table role="presentation" width="640" cellpadding="0" cellspacing="0" style="background:#ffffff;">
  <tr><td style="padding:12px 32px 28px;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"><tr>
      <td width="50%" style="vertical-align:top;padding:0 8px 0 0;">
        <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#f7f2ff;border-radius:12px;">
          <tr><td style="padding:22px 18px;">
            <table role="presentation" cellpadding="0" cellspacing="0"><tr>
              <td style="vertical-align:middle;width:40px;"><img src="{_nl_asset('icone_avis.png')}" alt="" width="32" height="32" style="display:block;border:0;" /></td>
              <td style="padding-left:10px;font-size:12px;font-weight:800;letter-spacing:1px;color:#1e1b4b;text-transform:uppercase;vertical-align:middle;">Votre avis nous est important</td>
            </tr></table>
            <div style="font-size:12px;color:#5b5486;line-height:1.5;margin:10px 0 14px;">Vous &#234;tes satisfait ? Un avis Google nous aide beaucoup !</div>
            <a href="{avis_url}" style="display:inline-block;background:#7c3aed;color:#ffffff;text-decoration:none;font-weight:600;font-size:13px;padding:12px 22px;border-radius:8px;white-space:nowrap;" target="_blank">Laisser un avis Google</a>
          </td></tr>
        </table>
      </td>
      <td width="50%" style="vertical-align:top;padding:0 0 0 8px;">
        <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#f7f2ff;border-radius:12px;">
          <tr><td style="padding:22px 18px;">
            <table role="presentation" cellpadding="0" cellspacing="0"><tr>
              <td style="vertical-align:middle;width:40px;"><img src="{_nl_asset('icone_parrainage.png')}" alt="" width="32" height="32" style="display:block;border:0;" /></td>
              <td style="padding-left:10px;font-size:12px;font-weight:800;letter-spacing:1px;color:#1e1b4b;text-transform:uppercase;vertical-align:middle;">Programme parrainage</td>
            </tr></table>
            <div style="font-size:12px;color:#5b5486;line-height:1.5;margin:10px 0 14px;">Recommandez une entreprise, gagnez jusqu&#39;&#224; 550&#8364; en cartes Amazon !</div>
            <a href="{parr_url}" style="display:inline-block;background:#7c3aed;color:#ffffff;text-decoration:none;font-weight:600;font-size:13px;padding:12px 22px;border-radius:8px;white-space:nowrap;" target="_blank">Je parraine une entreprise</a>
          </td></tr>
        </table>
      </td>
    </tr></table>
  </td></tr>
</table>
</td></tr>''')

    parts.append(f'''<!-- BLOC 7 : Signature / Contact -->
<tr><td align="center">
<table role="presentation" width="640" cellpadding="0" cellspacing="0" style="background:#ffffff;">
  <tr><td style="padding:24px 32px 28px;border-top:1px solid #ece7fb;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#f5f3ff;border-radius:12px;">
      <tr><td style="padding:20px 22px 10px;">
        <span style="font-family:Syne,Inter,Arial,sans-serif;font-size:20px;font-weight:700;color:#7c3aed;font-style:italic;">Besoin d&#39;un conseil&nbsp;?</span>
      </td></tr>
      <tr><td style="padding:0 22px 22px;">
        <table role="presentation" width="100%" cellpadding="0" cellspacing="0"><tr>
          <td style="vertical-align:middle;width:80px;">
            <img src="{_nl_asset('conseillere.png')}" alt="Conseill&#232;re" width="80" height="80" style="display:block;border:0;border-radius:50%;" />
          </td>
          <td style="padding-left:16px;vertical-align:middle;font-size:13px;line-height:1.5;">
            <strong style="color:#1e1b4b;">Notre &#233;quipe est &#224; votre &#233;coute.</strong><br>
            <span style="color:#5b5486;font-size:12px;">Une question sur votre contrat ? Contactez-nous.</span>
          </td>
          <td style="vertical-align:middle;padding-left:16px;text-align:right;white-space:nowrap;">
            <div style="font-size:11px;font-weight:800;letter-spacing:1px;color:#1e1b4b;text-transform:uppercase;margin-bottom:6px;">Service Relations Client</div>
            <a href="mailto:{ml['email']}" style="color:#7c3aed;text-decoration:none;font-weight:600;font-size:13px;">{ml['email']}</a><br>
            <span style="color:#1e1b4b;font-weight:600;font-size:13px;">{ml['telephone']}</span>
          </td>
        </tr></table>
      </td></tr>
    </table>
  </td></tr>
</table>
</td></tr>''')

    parts.append(f'''<!-- BLOC 8 : Footer navy -->
<tr><td align="center">
<table role="presentation" width="640" cellpadding="0" cellspacing="0" style="background:#1e1b4b;">
  <tr><td style="padding:28px 30px 12px;text-align:center;">
    <img src="{_nl_asset('logo_blanc.png')}" alt="LILIWATT" width="120" height="30" style="display:inline-block;border:0;margin-bottom:16px;" /><br>
    {social_ftr}
  </td></tr>
  <tr><td style="padding:16px 30px 8px;text-align:center;">
    {footer_nav_html}
  </td></tr>
  <tr><td style="padding:8px 30px 20px;text-align:center;font-size:11px;color:#b7aee0;line-height:1.6;">
    {ml['marque']} &#183; {ml['adresse']} &#183; {ml['email']} &#183; {ml['telephone']}
  </td></tr>
</table>
</td></tr>''')

    parts.append(f'''<!-- BLOC 9 : Mentions legales + desinscription -->
<tr><td align="center">
<table role="presentation" width="640" cellpadding="0" cellspacing="0" style="background:#f5f3ff;">
  <tr><td style="padding:16px 30px;text-align:center;font-size:10px;color:#8a83ad;line-height:1.7;">
    {ml['raison_sociale']} {ml['forme']} &#8211; {ml['siren']}<br>
    <a href="{unsub_url}" style="color:#7c3aed;text-decoration:underline;font-size:10px;">Se d&#233;sinscrire de cette newsletter</a>
  </td></tr>
</table>
</td></tr>

</table>
</body></html>''')

    return '\n'.join(parts)

def _nl_send_thread(objet, titre, message, dest_map, fmt='newsletter', cta_texte='', cta_lien=''):
    """Thread d'envoi échelonné. dest_map = {email_normalise: prenom_normalise}."""
    global _nl_status
    destinataires = sorted(dest_map.keys())
    _nl_status = {"en_cours": True, "total": len(destinataires), "envoyes": 0, "erreurs": [], "objet": objet}
    # Préparer le HTML du corps (invariant par destinataire)
    escaped = html_mod.escape(message)
    escaped = re.sub(r'\n{2,}', '<br>', escaped).replace('\n', '<br>')
    body_html = _nl_parse_balises(escaped)
    token_zoho = get_zoho_token()
    account_id = os.environ.get('ZOHO_ACCOUNT_ID', '8439060000000002002')
    base_url = os.environ.get('RENDER_EXTERNAL_URL', 'https://liliwatt-admin.onrender.com')
    for i, email in enumerate(destinataires):
        if not _nl_status["en_cours"]:
            break
        prenom = dest_map.get(email, '')
        unsub_token = _nl_unsub_token(email)
        unsub_url = f"{base_url}/newsletter/unsubscribe?email={email}&token={unsub_token}"
        full_html = _nl_build_html(objet, titre, body_html, unsub_url, fmt, cta_texte, cta_lien, prenom)
        try:
            # Renouveler le token tous les 40 envois
            if i > 0 and i % 40 == 0:
                token_zoho = get_zoho_token()
            r = requests.post(
                f'https://mail.zoho.eu/api/accounts/{account_id}/messages',
                headers={'Authorization': f'Zoho-oauthtoken {token_zoho}', 'Content-Type': 'application/json'},
                json={'fromAddress': 'contact@liliwatt.fr', 'toAddress': email,
                      'subject': objet, 'content': full_html, 'mailFormat': 'html'},
                timeout=15
            )
            if r.status_code >= 400:
                _nl_status["erreurs"].append(email)
                print(f"❌ Newsletter → {email}: HTTP {r.status_code}")
            else:
                _nl_status["envoyes"] += 1
                print(f"✅ Newsletter → {email} ({_nl_status['envoyes']}/{_nl_status['total']})")
        except Exception as e:
            _nl_status["erreurs"].append(email)
            print(f"❌ Newsletter → {email}: {e}")
        # Pause
        if (i + 1) % NL_LOT == 0 and (i + 1) < len(destinataires):
            print(f"⏸ Pause 60s après {i+1} envois…")
            time.sleep(60)
        else:
            time.sleep(NL_DELAI_S)
    _nl_status["en_cours"] = False
    # Log dans NEWSLETTER_LOG
    try:
        ws = _nl_get_or_create_sheet('NEWSLETTER_LOG', ['DATE', 'OBJET', 'NB_DEST', 'SUCCES', 'ERREURS'])
        ws.append_row([
            datetime.now().strftime('%Y-%m-%d %H:%M'),
            objet,
            str(_nl_status["total"]),
            str(_nl_status["envoyes"]),
            str(len(_nl_status["erreurs"]))
        ])
    except Exception as e:
        print(f"⚠️ Log newsletter: {e}")
    print(f"📊 Newsletter terminée: {_nl_status['envoyes']}/{_nl_status['total']} envoyés, {len(_nl_status['erreurs'])} erreurs")

@app.route('/newsletter/send', methods=['POST'])
@login_required
def newsletter_send():
    global _nl_status
    if _nl_status.get("en_cours"):
        return jsonify({"success": False, "error": "Une campagne est déjà en cours"}), 409
    d = request.get_json()
    objet = (d.get('objet') or '').strip()
    message = (d.get('message') or '').strip()
    cible = (d.get('cible') or 'tous').strip()
    email_unique = (d.get('email_unique') or '').strip().lower()
    if not objet or not message:
        return jsonify({"success": False, "error": "Objet et message requis"}), 400
    titre = (d.get('titre') or '').strip()
    fmt = (d.get('fmt') or 'newsletter').strip()
    cta_texte = (d.get('cta_texte') or '').strip()
    cta_lien = (d.get('cta_lien') or '').strip()
    # Construire dest_map {email: prenom}
    if cible == 'un' and email_unique:
        dest_map = {email_unique: ''}
    else:
        try:
            gc = get_sheets_client()
            ws = gc.open_by_key(SUIVI_VENTES_SHEET_ID).sheet1
            rows = ws.get_all_values()
            email_re = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')
            email_prenom = {}
            for row in rows[1:]:
                if len(row) > 22:
                    e = row[22].strip().lower()
                    if e and email_re.match(e) and e not in email_prenom:
                        raw_p = row[20].strip() if len(row) > 20 else ''
                        email_prenom[e] = _nl_normalise_prenom(raw_p)
            unsub = _nl_get_unsub_set()
            dest_map = {e: p for e, p in email_prenom.items() if e not in unsub}
        except Exception as e:
            return jsonify({"success": False, "error": f"Erreur lecture Sheet: {e}"}), 500
    if not dest_map:
        return jsonify({"success": False, "error": "Aucun destinataire trouvé"}), 400
    t = threading.Thread(target=_nl_send_thread, args=(objet, titre, message, dest_map, fmt, cta_texte, cta_lien), daemon=True)
    t.start()
    return jsonify({"success": True, "started": True, "total": len(dest_map)})

@app.route('/newsletter/status')
@login_required
def newsletter_status():
    return jsonify(_nl_status)

@app.route('/newsletter/count')
@login_required
def newsletter_count():
    try:
        gc = get_sheets_client()
        ws = gc.open_by_key(SUIVI_VENTES_SHEET_ID).sheet1
        rows = ws.get_all_values()
        email_re = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')
        brut = []
        for row in rows[1:]:
            if len(row) > 22:
                e = row[22].strip().lower()
                if e and email_re.match(e):
                    brut.append(e)
        total_brut = len(brut)
        uniques_set = set(brut)
        doublons = total_brut - len(uniques_set)
        unsub = _nl_get_unsub_set()
        desinscrits = len(uniques_set & unsub)
        destinataires = sorted(uniques_set - unsub)
        return jsonify({
            "success": True,
            "total_brut": total_brut,
            "uniques": len(uniques_set),
            "doublons": doublons,
            "desinscrits": desinscrits,
            "destinataires": len(destinataires)
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/newsletter/clients')
@login_required
def newsletter_clients():
    try:
        gc = get_sheets_client()
        ws = gc.open_by_key(SUIVI_VENTES_SHEET_ID).sheet1
        rows = ws.get_all_values()
        email_re = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')
        unsub = _nl_get_unsub_set()
        seen = set()
        clients = []
        for row in rows[1:]:
            if len(row) > 22:
                e = row[22].strip().lower()
                if not e or not email_re.match(e) or e in seen or e in unsub:
                    continue
                seen.add(e)
                nom    = row[19].strip() if len(row) > 19 else ''
                prenom = row[20].strip() if len(row) > 20 else ''
                parts  = [p for p in [prenom, nom] if p]
                label  = (' '.join(parts) + ' - ' + e) if parts else e
                clients.append({"email": e, "label": label, "nom": nom, "prenom": prenom})
        clients.sort(key=lambda c: (c['nom'].lower() or c['email'], c['prenom'].lower()))
        return jsonify({"success": True, "clients": clients})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/newsletter/sheet-headers')
@login_required
def newsletter_sheet_headers():
    try:
        gc = get_sheets_client()
        ws = gc.open_by_key(SUIVI_VENTES_SHEET_ID).sheet1
        headers = ws.row_values(1)
        return jsonify({"success": True, "headers": [{"index": i, "name": h} for i, h in enumerate(headers)]})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/newsletter/preview', methods=['POST'])
@login_required
def newsletter_preview():
    d = request.get_json()
    message = (d.get('message') or '').strip()
    objet = (d.get('objet') or 'Aperçu').strip()
    titre = (d.get('titre') or '').strip()
    fmt = (d.get('fmt') or 'newsletter').strip()
    cta_texte = (d.get('cta_texte') or '').strip()
    cta_lien = (d.get('cta_lien') or '').strip()
    escaped = html_mod.escape(message)
    escaped = re.sub(r'\n{2,}', '<br>', escaped).replace('\n', '<br>')
    body_html = _nl_parse_balises(escaped)
    preview_html = _nl_build_html(objet, titre, body_html, '#', fmt, cta_texte, cta_lien)
    return jsonify({"success": True, "html": preview_html})

@app.route('/newsletter/test-send', methods=['POST'])
@login_required
def newsletter_test_send():
    d = request.get_json()
    objet = (d.get('objet') or '').strip()
    titre = (d.get('titre') or '').strip()
    message = (d.get('message') or '').strip()
    test_email = (d.get('test_email') or '').strip().lower()
    fmt = (d.get('fmt') or 'newsletter').strip()
    cta_texte = (d.get('cta_texte') or '').strip()
    cta_lien = (d.get('cta_lien') or '').strip()
    if not objet or not message:
        return jsonify({"success": False, "error": "Objet et message requis"}), 400
    if not test_email:
        return jsonify({"success": False, "error": "Email de test requis"}), 400
    # Chercher le prénom dans le Sheet ; sinon "Test" (pour voir le rendu personnalisé)
    prenom_test = 'Test'
    prenom_source = 'fallback'
    try:
        gc = get_sheets_client()
        ws = gc.open_by_key(SUIVI_VENTES_SHEET_ID).sheet1
        rows = ws.get_all_values()
        email_re = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')
        for row in rows[1:]:
            if len(row) > 22 and row[22].strip().lower() == test_email:
                raw_p = row[20].strip() if len(row) > 20 else ''
                p = _nl_normalise_prenom(raw_p)
                if p:
                    prenom_test = p
                    prenom_source = 'sheet'
                break
    except Exception:
        pass  # Sheet inaccessible : on garde "Test"
    escaped = html_mod.escape(message)
    escaped = re.sub(r'\n{2,}', '<br>', escaped).replace('\n', '<br>')
    body_html = _nl_parse_balises(escaped)
    full_html = _nl_build_html(objet, titre, body_html, '#', fmt, cta_texte, cta_lien, prenom_test)
    try:
        token_zoho = get_zoho_token()
        account_id = os.environ.get('ZOHO_ACCOUNT_ID', '8439060000000002002')
        r = requests.post(
            f'https://mail.zoho.eu/api/accounts/{account_id}/messages',
            headers={'Authorization': f'Zoho-oauthtoken {token_zoho}', 'Content-Type': 'application/json'},
            json={'fromAddress': 'contact@liliwatt.fr', 'toAddress': test_email,
                  'subject': f'[TEST] {objet}', 'content': full_html, 'mailFormat': 'html'},
            timeout=15
        )
        if r.status_code >= 400:
            return jsonify({"success": False, "error": f"Zoho HTTP {r.status_code}"}), 500
        return jsonify({"success": True, "prenom": prenom_test, "prenom_source": prenom_source})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/bienvenue/defaults')
@login_required
def bienvenue_defaults():
    return jsonify({"success": True, "defaults": BIENVENUE_DEFAUT})

@app.route('/bienvenue/preview', methods=['POST'])
@login_required
def bienvenue_preview():
    d = request.get_json()
    objet   = (d.get('objet')   or BIENVENUE_DEFAUT['objet']).strip()
    titre   = (d.get('titre')   or BIENVENUE_DEFAUT['titre']).strip()
    message = (d.get('message') or BIENVENUE_DEFAUT['message']).strip()
    cta_texte = (d.get('cta_texte') or '').strip()
    cta_lien  = (d.get('cta_lien')  or '').strip()
    escaped = html_mod.escape(message)
    escaped = re.sub(r'\n{2,}', '<br>', escaped).replace('\n', '<br>')
    body_html = _nl_parse_balises(escaped)
    preview_html = _nl_build_html(objet, titre, body_html, '#', 'bienvenue', cta_texte, cta_lien)
    return jsonify({"success": True, "html": preview_html})

@app.route('/bienvenue/send', methods=['POST'])
@login_required
def bienvenue_send():
    d = request.get_json()
    objet     = (d.get('objet')     or BIENVENUE_DEFAUT['objet']).strip()
    titre     = (d.get('titre')     or BIENVENUE_DEFAUT['titre']).strip()
    message   = (d.get('message')   or BIENVENUE_DEFAUT['message']).strip()
    email_dest = (d.get('email')    or '').strip().lower()
    cta_texte = (d.get('cta_texte') or '').strip()
    cta_lien  = (d.get('cta_lien')  or '').strip()
    if not email_dest:
        return jsonify({"success": False, "error": "Email destinataire requis"}), 400
    # Prénom depuis le Sheet
    prenom = ''
    try:
        gc = get_sheets_client()
        ws = gc.open_by_key(SUIVI_VENTES_SHEET_ID).sheet1
        for row in ws.get_all_values()[1:]:
            if len(row) > 22 and row[22].strip().lower() == email_dest:
                raw_p = row[20].strip() if len(row) > 20 else ''
                prenom = _nl_normalise_prenom(raw_p)
                break
    except Exception:
        pass
    escaped = html_mod.escape(message)
    escaped = re.sub(r'\n{2,}', '<br>', escaped).replace('\n', '<br>')
    body_html = _nl_parse_balises(escaped)
    unsub_url = url_for('newsletter_unsubscribe', email=email_dest,
                        token=_nl_unsub_token(email_dest), _external=True)
    full_html = _nl_build_html(objet, titre, body_html, unsub_url, 'bienvenue', cta_texte, cta_lien, prenom)
    try:
        token_zoho = get_zoho_token()
        account_id = os.environ.get('ZOHO_ACCOUNT_ID', '8439060000000002002')
        r = requests.post(
            f'https://mail.zoho.eu/api/accounts/{account_id}/messages',
            headers={'Authorization': f'Zoho-oauthtoken {token_zoho}', 'Content-Type': 'application/json'},
            json={'fromAddress': 'contact@liliwatt.fr', 'toAddress': email_dest,
                  'subject': objet, 'content': full_html, 'mailFormat': 'html'},
            timeout=15
        )
        if r.status_code >= 400:
            return jsonify({"success": False, "error": f"Zoho HTTP {r.status_code}"}), 500
        return jsonify({"success": True, "prenom": prenom or ''})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/bienvenue/test-send', methods=['POST'])
@login_required
def bienvenue_test_send():
    d = request.get_json()
    objet      = (d.get('objet')      or BIENVENUE_DEFAUT['objet']).strip()
    titre      = (d.get('titre')      or BIENVENUE_DEFAUT['titre']).strip()
    message    = (d.get('message')    or BIENVENUE_DEFAUT['message']).strip()
    test_email = (d.get('test_email') or '').strip().lower()
    cta_texte  = (d.get('cta_texte') or '').strip()
    cta_lien   = (d.get('cta_lien')  or '').strip()
    if not test_email:
        return jsonify({"success": False, "error": "Email de test requis"}), 400
    prenom_test   = 'Test'
    prenom_source = 'fallback'
    try:
        gc = get_sheets_client()
        ws = gc.open_by_key(SUIVI_VENTES_SHEET_ID).sheet1
        for row in ws.get_all_values()[1:]:
            if len(row) > 22 and row[22].strip().lower() == test_email:
                raw_p = row[20].strip() if len(row) > 20 else ''
                p = _nl_normalise_prenom(raw_p)
                if p:
                    prenom_test   = p
                    prenom_source = 'sheet'
                break
    except Exception:
        pass
    escaped = html_mod.escape(message)
    escaped = re.sub(r'\n{2,}', '<br>', escaped).replace('\n', '<br>')
    body_html = _nl_parse_balises(escaped)
    full_html = _nl_build_html(objet, titre, body_html, '#', 'bienvenue', cta_texte, cta_lien, prenom_test)
    try:
        token_zoho = get_zoho_token()
        account_id = os.environ.get('ZOHO_ACCOUNT_ID', '8439060000000002002')
        r = requests.post(
            f'https://mail.zoho.eu/api/accounts/{account_id}/messages',
            headers={'Authorization': f'Zoho-oauthtoken {token_zoho}', 'Content-Type': 'application/json'},
            json={'fromAddress': 'contact@liliwatt.fr', 'toAddress': test_email,
                  'subject': f'[TEST] {objet}', 'content': full_html, 'mailFormat': 'html'},
            timeout=15
        )
        if r.status_code >= 400:
            return jsonify({"success": False, "error": f"Zoho HTTP {r.status_code}"}), 500
        return jsonify({"success": True, "prenom": prenom_test, "prenom_source": prenom_source})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/newsletter/unsubscribe')
def newsletter_unsubscribe():
    email = (request.args.get('email') or '').strip().lower()
    token = (request.args.get('token') or '').strip()
    if not email or not token:
        return '<h1>Lien invalide</h1>', 400
    expected = _nl_unsub_token(email)
    if not hmac.compare_digest(token, expected):
        return '<h1>Lien invalide</h1>', 400
    # Ajouter dans NEWSLETTER_UNSUB
    try:
        ws = _nl_get_or_create_sheet('NEWSLETTER_UNSUB', ['EMAIL', 'DATE', 'ORIGINE'])
        existing = {r[0].strip().lower() for r in ws.get_all_values()[1:] if r and r[0]}
        if email not in existing:
            ws.append_row([email, datetime.now().strftime('%Y-%m-%d %H:%M'), 'lien_mail'])
    except Exception as e:
        print(f"⚠️ Erreur désinscription sheet: {e}")
    # Notifier contact@liliwatt.fr
    try:
        tk = get_zoho_token()
        if tk:
            account_id = os.environ.get('ZOHO_ACCOUNT_ID', '8439060000000002002')
            requests.post(
                f'https://mail.zoho.eu/api/accounts/{account_id}/messages',
                headers={'Authorization': f'Zoho-oauthtoken {tk}', 'Content-Type': 'application/json'},
                json={'fromAddress': 'contact@liliwatt.fr', 'toAddress': 'contact@liliwatt.fr',
                      'subject': f'Désinscription newsletter : {email}',
                      'content': f'<p>{email} s\'est désinscrit de la newsletter LILIWATT.</p>', 'mailFormat': 'html'},
                timeout=10
            )
    except Exception as e:
        print(f"⚠️ Notif désinscription: {e}")
    ml = MENTIONS_LEGALES
    unsub_page = (
        '<!DOCTYPE html><html lang="fr"><head><meta charset="UTF-8">'
        '<meta name="viewport" content="width=device-width,initial-scale=1">'
        '<title>D\u00e9sinscription \u2014 LILIWATT</title>'
        '<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=Inter:wght@400;600;700&display=swap" rel="stylesheet">'
        '<style>'
        '*{margin:0;padding:0;box-sizing:border-box}'
        'body{font-family:Inter,sans-serif;background:#f4f1fb;min-height:100vh;display:flex;align-items:center;justify-content:center;padding:20px}'
        '.card{background:#fff;border-radius:22px;max-width:480px;width:100%;overflow:hidden;box-shadow:0 20px 60px rgba(124,58,237,.12)}'
        '.head{background:linear-gradient(135deg,#1e1b4b,#7c3aed 60%,#d946ef);padding:28px;text-align:center;color:#fff}'
        '.head h1{font-family:Syne,sans-serif;font-size:24px;font-weight:800;letter-spacing:1px}'
        '.head p{font-size:12px;opacity:.8;margin-top:4px;letter-spacing:1px}'
        '.body{padding:32px 28px;text-align:center}'
        '.check{width:64px;height:64px;border-radius:50%;background:linear-gradient(135deg,#22c55e,#16a34a);color:#fff;font-size:32px;display:flex;align-items:center;justify-content:center;margin:0 auto 18px}'
        '.body h2{font-size:20px;color:#1e1b4b;margin-bottom:10px}'
        '.body p{color:#6b6591;font-size:14px;line-height:1.6}'
        '.foot{background:#f4f1fb;padding:16px;text-align:center;font-size:11px;color:#8a83ad}'
        '</style></head><body>'
        '<div class="card">'
        '<div class="head"><h1>LILIWATT</h1><p>NEWSLETTER</p></div>'
        '<div class="body">'
        '<div class="check">\u2713</div>'
        '<h2>D\u00e9sinscription confirm\u00e9e</h2>'
        '<p>L\'adresse <strong>' + email + '</strong> a \u00e9t\u00e9 retir\u00e9e de notre liste. '
        'Vous ne recevrez plus de newsletters.</p>'
        '</div>'
        '<div class="foot">' + ml['marque'] + ' \u2013 ' + ml['adresse'] + '</div>'
        '</div></body></html>'
    )
    return unsub_page, 200


if __name__ == '__main__':
    app.run(debug=True, port=5001)
